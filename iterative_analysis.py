print("Importing libraries...\n-------------------------------------")

# region Imports
from scipy.interpolate import RBFInterpolator
from sklearn.preprocessing import PolynomialFeatures
from sklearn.linear_model import LinearRegression
from mealpy import PSO, FloatVar
from sklearn import linear_model
import scipy.stats as stats
import sys
import os
import matplotlib
from pyNastran.bdf.bdf import BDF
from pyNastran.op2.op2 import OP2
from numpy.linalg import norm
import matplotlib.pyplot as plt
import numpy as np
import cv2
from scipy.interpolate import griddata
import time
import numpy as np
import image_process_tool_box as imgan
import file_paths as path
from speckle_pattern import generate_and_save, generate_lines, generate_checkerboard
import openpyxl
import sundic.sundic as sdic
import sundic.post_process as sdpp
import sundic.settings as sdset
import os
from pyDOE import lhs
import configparser
from scipy.optimize import minimize
import traceback
import numpy as np
import sundic.sundic as sdic
import sundic.post_process as sdpp
import sundic.settings as sdset
import cv2
import os
import sys
from datetime import datetime
import warnings
import subprocess
import scienceplots
# endregion

FLAGS = {
    'Create_new_excel': False,   
    'generate_pattern': False,
    'Perlin_images': False,
    'Pattern analysis': False,
    'Load FEA': False,
    'Deform speckles': False,
    'run_dic': False,               # Fixed
    'run_error': False,
    'error_dist': False,    
    'Excel plots': False,
    'Optimisation_poly': False
}
imgan.flag_status(FLAGS,2.5)    # Just in case

tic1 = time.time()
'=========================LOAD PATHS AND GLOBAL VARIABLES===================='

# Number of patterns
pattern_count = 1
# FEA data
op2_path = path.flat30_4mm_op2
bdf_path = path.flat30_4mm_bdf

# op2_path = path.msjul_30_op2
# bdf_path = path.msjul_30_bdf

image_width = 2000
image_height = 500                    # Incorrect input here might result in broadcasting errors downstream

# Translation values
u_min = 0.0
u_max = 1.05
uinterval = 0.05

# The extension for images we are currently working with
image_type = 'tif'
image_save_path = 2 # Explained in pattern generation. Used in pattern analysis too hence global

# Image deformation scenario
deform_type = 0

# DIC analysis scenario
study_type = 0

# Iterative process. Some variable will be altered and errors are plotted against
# said parameter at the end of the process

# changing_parameter = [1,5,9,13,17,21,25,29,33,37,41,45,49,53,57,61,65,69,71,75,79,83,87,91]

changing_parameter = [10,15,20,25,30,35,40,45,50,55,60,65,70,75,80,85,90,95,100,105,110,115,120,125,130,135,140]
# changing_parameter = [13,17,21,25,29,33,37,41,45,49,53,57,61,65,69,13,17,21,25,29,33,37,41,45,49,53,57,61,65,69]

# changing_parameter = [0,3,6,9,12,15,18,21,24,27,30,33,36,39,42,45,48,51]

shape_fun = ["Affine","Quadratic"]

iterations_limit = len(changing_parameter)

for iteration in range(iterations_limit):
    
    print(f'Changing parameter: {changing_parameter[iteration]}')
    # Testing different mesh element sizes
    # param = changing_parameter[iteration]
    # op2_path = getattr(path, f'msjul_{param}_op2')
    # bdf_path = getattr(path, f'msjul_{param}_bdf')

    # Change flags after first iteration
    if iteration > 0:
        FLAGS['generate_pattern'] = False
        FLAGS['Deform speckles'] = False
        FLAGS['Perlin_images'] = False

    # For shape function analysis
    # half_iter =  len(changing_parameter)//2
    # if iteration > half_iter:
    #     shappe = shape_fun[1]
    # else:
    #     shappe = shape_fun[0]

    # region Path variables and directories
    #-----------------------------------------------------------------------------------------------

    # Paths 
    error_hist_path         = rf"output\histograms\iterative_study\iter_{iteration}"
    DIC_contour_path        = rf"output\DIC_contour\iterative_study\iter_{iteration}"
    znssd_figure_path       = rf"output\DIC_contour\ZNSSD\iterative_study\iter_{iteration}"
    error_heatmap_path      = rf"output\Heatmaps\iterative_study\iter_{iteration}"
    debugg_folder           = rf"output\Debugging\iterative_study\iter_{iteration}"
    excel_path              = rf"output\excel_docs\iterative_study\iter_{iteration}"
    deformed_image_path     = r"data\speckle_pattern_img\deformed_im"
    reference_image_path    = r"data\speckle_pattern_img\reference_im"
    difference_images       = rf"output\plots\Difference_images\iterative_study\iter_{iteration}"
    DIC_settings_path       = r"settings.ini"
    Contour_path            = rf"output\DIC_contour\iterative_study\iter_{iteration}"
    spt_img_path            = rf"data\speckle_pattern_img\subpixel_translation"
    numpy_files             = rf"output\numpy_files\iterative_study\iter_{iteration}"
    sundic_binary_folder    = rf"output\sundic_bin\iter_{iteration}"
    plot_path               = rf"output\plots\iterative_study\iter_{iteration}"
    single_plot_path        = rf"output\plots"
    slice_path              = rf"output\Slices\iterative_study\iter_{iteration}"
    image_matrix            = r"output\image_matrix"
    autocorrelation_path    = rf"output\Autocorrelation\iterative_study\iter_{iteration}"
    power_spec              = rf"output\spectral_analysis\iterative_study\iter_{iteration}"
    optimised_save          = rf"data\speckle_pattern_img\Optimised"
    scatter_plots           = rf"output\plots\Scatter_plots"
    planar_images           = r"planar_images"



    # List of all directories to ensure exist
    dirs_to_create = [
        error_hist_path,
        DIC_contour_path,
        znssd_figure_path,
        error_heatmap_path,
        debugg_folder,
        excel_path,
        deformed_image_path,
        reference_image_path,
        difference_images,
        spt_img_path,
        sundic_binary_folder,
        plot_path,
        single_plot_path,
        slice_path,
        numpy_files,
        image_matrix,
        autocorrelation_path,
        power_spec,
        optimised_save,
        scatter_plots,
        planar_images
    ]

    # Create each directory if it doesn't exist
    for dir_path in dirs_to_create:
        if not os.path.exists(dir_path):
            os.makedirs(dir_path)
    #-----------------------------------------------------------------------------------------------
    # endregion

    '===========================CREATE EXCEL SHEET================================='
    if FLAGS['Create_new_excel']:
        """
        Creates a new excel document to store data. Must remember to turn off the flag
        to avoid overwriting. The document number also changes everytime this block is
        executed.
        """
        print("\n1. Creating excel document...\n")

        if not os.path.exists(excel_path):
            os.makedirs(excel_path)
            doc_number = 1
        else:
            # Use existing files to determine name for new file
            existing_files = [f for f in os.listdir(excel_path)
                                if f.startswith("Pattern_evaluation_") and f.endswith(".xlsx")]
            doc_number = len(existing_files) + 1
        print("Current doc_number:", doc_number)
        exl_file_name = f"Pattern_evaluation_{doc_number}.xlsx"
        exl_file_path = os.path.join(excel_path, exl_file_name)

        # Create new workbook and save it
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        workbook.save(exl_file_path)
        del workbook


    '===========================GENERATE PATTERN IMAGES============================'
    if FLAGS['generate_pattern']:
        print('\n2. Generating speckle pattern images...\n')

        # latest excel file
        exl_file_path = imgan.excel_doc_path(excel_path)

        # Generate and save     
        pattern_method = 1
        with_turing = False

        if pattern_method == 1:
            # Speckles
            imgan.ladisk_generator(exl_file_path,reference_image_path,pattern_count,image_width=image_width,
                                image_height=image_height)
            
        elif pattern_method == 2:
            # Straight lines
            imgan.ladisk_generator_lines(exl_file_path,reference_image_path,pattern_count,image_width=image_width,
                                image_height=image_height)
            
        elif pattern_method == 3:
            # Checkerboard
            imgan.ladisk_generator_cb(exl_file_path,reference_image_path,pattern_count,image_width=image_width,
                                image_height=image_height)
            
        elif pattern_method == 4:
            # Perlin images with textures
            '''
            Texture options:    none*
                            thresholded
                            sinusoidal
                            bimodal
                            logarithmic
                            cubic*
                            leopard
                            perlin_blobs*
                            gaussian_spots
                            worley
            '''
            imgan.single_perlin(image_height, 
                                image_width, 
                                excel_path=exl_file_path, 
                                ref_image_save=reference_image_path, 
                                texture_function="none",
                                number_of_images=pattern_count
                                )
            
        if with_turing:
            # Generate turing look-alikes from existing images
            imgan.make_turing(reference_image_path, 
                              rep=40, 
                              radius=1, 
                              sharpen_percent=250, 
                              size=None, 
                              replace=True
                              )

        plt.close('all')


    '=============================PERLIN NOISE IMAGES=============================='
    if FLAGS['Perlin_images']:
        # Generate a new excel sheet to avoid overriding the old one if/incase it has important
        # results

        # Check for existing excel files and get the path to the most recent one
        exl_file_path = imgan.excel_doc_path(excel_path)

        # Load the Fe nodes and displaced nodes
        nodes_2d, deformed_nodes_2d = imgan.load_fe_nodes(bdf_path, op2_path)

        # Scale the data to fit the image
        nodes_2d = nodes_2d * 1000
        deformed_nodes_2d = deformed_nodes_2d * 1000

        # Create matrix for sizing boundaries of the displacement field
        matrix = np.full((image_width,image_height), np.nan)
        _,_,dx, dy = imgan.smooth_field(matrix, nodes_2d, deformed_nodes_2d, 3)

        '''
        Texture options:    none
                            thresholded
                            sinusoidal
                            bimodal
                            logarithmic
                            cubic
                            leopard
                            perlin_blobs
                            gaussian_spots
                            binary_blur
                            worley
        '''
        imgan.generate_perlin_pair(image_height,image_width, dx, dy,
                                exl_file_path,reference_image_path,deformed_image_path,
                                number_of_images=pattern_count, texture_fun='sinusoidal')
        

    '=============================ANALYSE PATTERNS================================='
    if FLAGS['Pattern analysis']:
        print('\n3. Analysing speckle patterns...\n')
        """
        This block reads reference speckle pattern images from the selected folder (data or planar images)
        and applies the defined speckle pattern metrics. The code will be adjusted to be able to read
        arbitrary image files as well. Metric values are saved in relevant cells in the active Excel sheet.
        """
        try:
            # Most recent excel
            exl_file_path = imgan.excel_doc_path(excel_path)
            # Create workbook if it doesn't exist
            try:
                workbook = openpyxl.load_workbook(exl_file_path)
            except FileNotFoundError:
                workbook = openpyxl.Workbook()

            sheet = workbook.active
            sheet['A1'] = 'MSF'                 # Gradient-based 1st order
            sheet['B1'] = 'MIG'                 # Kind of gradient-based
            sheet['C1'] = 'E_f'                 # Gradient based 1st and 2nd order
            sheet['D1'] = 'MIOSD'               # Gradient based 2nd order
            sheet['E1'] = 'Shannon'             # Information content
            sheet['F1'] = 'Power spectrum area'      # Information content
            sheet['G1'] = 'SSSIG globe'         # Gradient based
            sheet['H1'] = 'Correlation peak radius'         # Morphology and feature content/distribution
            # Determine the image folder based on the save path

            imgan.rename_img(reference_image_path)      # Rename images so that they follow the same convention

            # Get reference images (even-numbered)
            image_files = imgan.get_image_strings(reference_image_path,imagetype=image_type)

            print("\n------\nimage files",image_files)
            # Assume that all the image files follow the same convention as the first file
            first_file = image_files[0]
            # Extract the string parts following the prefix. This will be used for checking if files exist in the next loop.
            image_file_endswith = '_'.join(first_file.split('_')[1:])

            reference_image_prefixes = imgan.expected_prefixes(reference_image_path)
            prefix_index = {prefix: i for i, prefix in enumerate(reference_image_prefixes)}

            for i, prefix_num in enumerate(reference_image_prefixes):
                # excel index
                j = prefix_index[prefix_num] + 2

                # Assume reference file name and skip index if it does not exist
                ref_image_file = f'{prefix_num}_{image_file_endswith}'

                # Read reference image (even number)
                ref_img_path = os.path.join(reference_image_path, ref_image_file)

                try: 

                    if not os.path.exists(ref_img_path):
                        raise Exception(f'Image file {ref_image_file} not found. Printing zeros ')

                    # Run pattern analysis (flag)
                    # Generated_speckles = cv2.imread(ref_img_path, cv2.IMREAD_GRAYSCALE)
                    Generated_speckles = imgan.readImage(ref_img_path)

                    print(f'{ref_image_file}:')
                    msf = imgan.MSF(Generated_speckles,33,5)
                    mean_intensity_gradient = imgan.MIG(Generated_speckles)

                    #---------------------
                    # change
                    subset_entropy = imgan.meanSE(Generated_speckles, 33, 5)
                    
                    mag, powr, fx, fy = imgan.compute_fft_and_freq(Generated_speckles)
                    power_area = imgan.integrate_power_2d(powr, fx, fy)
                    power_fig = imgan.plot_1d_spectra(mag, powr, fx, fy, show=False)

                    power_figsave_path = os.path.join(power_spec, f'Power_spectrum_{ref_image_file}.png')
                    power_fig.savefig(power_figsave_path)
                    plt.close(power_fig)

                    #---------------------

                    sss = imgan.globalSSSIG(Generated_speckles, 33, 5)
                    shannon = imgan.ShannonEnt(Generated_speckles)
                    miosd = imgan.miosd(Generated_speckles)
                    ef = imgan.Ef(Generated_speckles)
                    _,figR,R_peak = imgan.autocorr(Generated_speckles, meth=6, cardinality=100, autype='2d')
                    # Save autocorrelation figure
                    plt.figure(figsize=(10, 8))
                    figsave_path = os.path.join(autocorrelation_path, f'Autocorrelation_{ref_image_file}.png')
                    figR.savefig(figsave_path)
                    plt.close(figR)
                    plt.close('all')


                    # Skip image indeces that do not exist. 
                    # Check image prefixes
                    sheet[f'A{j}'] = msf
                    sheet[f'B{j}'] = mean_intensity_gradient
                    sheet[f'C{j}'] = ef
                    sheet[f'D{j}'] = miosd
                    sheet[f'E{j}'] = shannon

                    #---------------------------
                    sheet[f'F{j}'] = power_area
                    #--------------------------

                    sheet[f'G{j}'] = sss
                    sheet[f'H{j}'] = R_peak

                    
                except Exception:

                    sheet[f'A{j}'] = 0
                    sheet[f'B{j}'] = 0
                    sheet[f'C{j}'] = 0
                    sheet[f'D{j}'] = 0
                    sheet[f'E{j}'] = 0
                    sheet[f'F{j}'] = 0
                    sheet[f'G{j}'] = 0
                    sheet[f'H{j}'] = 0


            # Save the workbook
            workbook.save(exl_file_path)
            print(f"Workbook saved at {exl_file_path}")
        except FileNotFoundError as fnf_error:
            print(f"File or folder not found: {fnf_error}")
        except Exception as e:
            tb = traceback.extract_tb(sys.exc_info()[2])  # traceback info
            filename, line_number, function_name, text = tb[-1]  # Last traceback entry
            print(f'Error in pattern analysis on line {line_number}: {str(e)}')
            print(f'Look at {filename}')

        finally:
            if 'workbook' in locals():
                workbook.close()

        plt.close('all')


    '==========================IMPORT FINITE ELEMENT MESH=========================='
    if FLAGS['Load FEA']:
        # Display figure of mesh nodes if necessary
        FE_flags = {'show meshes':False,
                    'show_displacement_field': False}

        print('\n4. Loading FEA data...\n')

        model = OP2()
        model.read_op2(op2_path)
        bdf = BDF()
        bdf.read_bdf(bdf_path)

        # Extract the displacement data from the op2 file. Assuming isubcase = 1 and static analysis
        itime = 0
        isubcase = 1
        disp = model.displacements[isubcase]
        # Extract the translational displacements
        txyz = disp.data[itime, :, :3]

        # Calculate the total deflection of the vector (from documentation)
        total_xyz = norm(txyz, axis=1)
        nnodes = disp.data.shape[1]

        # Get the node positions from the BDF file
        nodes = np.array([bdf.nodes[nid].get_position() for nid in bdf.nodes])
        # Extract only the x and y components for 2D visualization (2d dic)
        nodes_2d = nodes[:, :2]
        print(f'\n-------\nNodes 2D shape: {nodes_2d.shape}\n-------')
        displacements_2d = txyz[:, :2]
        # Calculate the deformed positions by adding displacements to original node positions [The shapes of the objects from the bdf and the op2 are (nnodes, 3)]
        # deformed_nodes = nodes + txyz
        # Calculate the deformed positions by adding displacements to original node positions (2D)
        deformed_nodes_2d = nodes_2d + displacements_2d

        # Plotting the original and deformed mesh in 2D
        if FE_flags['show meshes']:
            plt.figure(figsize=(12, 6))
            # Original mesh
            plt.subplot(1, 2, 1)
            plt.scatter(nodes_2d[:, 0] , nodes_2d[:, 1], c='b', marker='o', label='Original')
            plt.title('Original Mesh')
            plt.xlabel('X')
            plt.ylabel('Y')
            plt.grid(True)
            plt.axis('equal')
            # Deformed mesh
            plt.subplot(1, 2, 2)
            plt.scatter(deformed_nodes_2d[:, 0], deformed_nodes_2d[:, 1], c='r', marker='o', label='Deformed')
            plt.title('Deformed Mesh')
            plt.xlabel('X')
            plt.ylabel('Y')
            plt.grid(True)
            plt.axis('equal')
            plt.show(block=False)
            cfm = plt.get_current_fig_manager()     # Bring window to foreground
            cfm.window.activateWindow()
            cfm.window.raise_()
            plt.pause(2.5)
            plt.close()

        # FEA points and alignment
        # 1 unit change in FE x,y results in 0.001 pixel change in pixel coordinates
        # the FE coordinates and associated displacements are scaled up by 1000 to 
        # have a 1 to 1 scale between the FE coordinates and the image coordinates
        # This was decided after the first time I tried to align a mesh to an image:
        # The (2000 by 500 mm) mesh was 1000 times smaller than the (2000 by 500) image.
        # Scale data up to fit image. Assume both are rectangular and aligned at origin,
        # and that they are parallel

        max_x_FE = (np.max(nodes_2d[:,0]))
        max_y_FE = (np.max(nodes_2d[:,1]))
        print(f'max x = {max_x_FE}\nmax y = {max_y_FE}')

        # Read an image to scale FE data
        read_image = imgan.get_image_strings(reference_image_path)
        # read_first_image_path = os.path.join(reference_image_path,read_image[0])
        # reference_image_for_scale = cv2.imread(read_first_image_path,cv2.IMREAD_GRAYSCALE)
        image_x_length = image_width
        print(f'Image x length = {image_x_length}')

        # Scale coordinates
        # FE_to_img = image_width / max_x_FE    
        x_scale = image_x_length / max_x_FE       
        print(f'\nscale = {x_scale}')

    
        original_points = nodes_2d * x_scale
        new_points = deformed_nodes_2d * x_scale

        

        print('\nFEA data loaded successfully.')
        dx, dy, FEx_rbf_interp, FEy_rbf_interp = imgan.smooth_field(
            np.zeros((image_height, image_width)),
            original_points,
            new_points,
            3
        )

        if FE_flags['show_displacement_field']:
            matrix = np.zeros((image_height,image_width))
            field = imgan.show_field(matrix, original_points,new_points,debugg_folder)
            plt.draw()
            plt.pause(2.5)
            plt.close('all')


    '==============================DEFORM IMAGES==================================='
    if FLAGS['Deform speckles']:
        """
        The generated images files are accessed to conduct the deformation procedure
        using radial basis function interpolation. The new files are saved separately
        and will be accessed alongside the reference images by the DIC program.
        
        Image deformation occurs through a series of steps:
            1. Corresponding BDF and OP2 files are accessed.
            2. Node coordinates are accessed from the BDF files
            3. Node displacement data is accessed from the op2 files
            4. The two sets of data are interpolated using RF functions and are used
            to remap the image coordinates using scikit-ndimage.map_coordinates
        """
        deform_flag = {
            'gaussian_blur_ref_img': True,
            'gaussian_noise_ref_img': False,
            'show process': False,
            'Inverse consistency': False
        }

        if iteration > 0:
            deform_flag['gaussian_blur_ref_img'] = False

        imgan.flag_status(deform_flag, wait_time=1.5)
               
        # Add gaussian noise
        if deform_flag['gaussian_noise_ref_img']:
            print('\nReference image prefilter')
            imgan.gaussian_noise_images(
                reference_image_path,
                par='even'
            )

        # Add Gaussian blur
        if deform_flag['gaussian_blur_ref_img']:
            print('\nReference image prefilter')
            imgan.gaussian_blur_images(
                reference_image_path,
                size=5,
                sig_y=1.0,
                par='even'
            )


        
        # If 'show process'
        wait_time = 0.025

        """
        Deformation type:
            0 -> FE-based
            1 -> Rigid body translation
        """

        print('\n5. Deforming images...')
        if deform_type == 0:

            print('\nFE-based deformation\n')
            image_files = imgan.get_image_strings(reference_image_path)

            for k, ref_image_file in enumerate(image_files):
                # Read reference image (even number). For each image, create a path
                # to the image by
                ref_img_path = os.path.join(reference_image_path, ref_image_file)
                reference_image = cv2.imread(ref_img_path)
                if reference_image is None:
                    print(f"Warning: Could not read reference image {ref_img_path}")
                    continue

                # Create deformed image name (odd number).
                ref_num = int(ref_image_file.split('_')[0])
                deform_num = ref_num + 1
                deformed_name = f"{deform_num}_Generated_spec_image.tif"

                # Deform images
                tic_def = time.time()
                print(f'\nApplying deformation: Image {k+1}')
                
                #---------------------------------------------------------------
                # DEFORM
                # transformed_image, difference_image = imgan.img_deform(
                #     reference_image, original_points, new_points, 3)
                transformed_image, difference_image = imgan.img_deform(
                    reference_image,
                    dx=dx,
                    dy=dy
                )
                #--------------------------------------------------------------

                deformed_path = os.path.join(deformed_image_path, deformed_name)
                cv2.imwrite(deformed_path, transformed_image)
                toc_def = time.time()

                # Showing the difference image
                if difference_image.ndim == 3:
                    difference_image = np.mean(difference_image, axis=2)

                normalised_difference = difference_image / difference_image.max()

                # This is where the difference images are created and saved
                plt.figure(figsize=(6, 5))
                plt.title("Difference image")
                img = plt.imshow(normalised_difference, cmap='jet')
                plt.colorbar(label="Difference")
                plt.axis('off')

                plt.tight_layout()

                # Save the figure exactly as displayed (including colormap)
                difference_img_save_path = os.path.join(
                    difference_images,
                    f'{ref_num}_difference_image.png'
                )
                plt.savefig(
                    difference_img_save_path,
                    bbox_inches='tight',
                    pad_inches=0,
                    dpi=300
                )

                # Show and then close
                plt.show(block=False)
                plt.pause(wait_time)
                plt.close()

                # Collapsed differences
                collapsed_difference = np.mean(difference_image, axis=0)
                # Generate the x-line (pixel positions along the columns of the image)
                x_line = np.arange(difference_image.shape[1])

                # Plotting the collapsed error (difference) along axis 0 as a line graph
                plt.figure(figsize=(10, 6))
                plt.plot(x_line, collapsed_difference, color = 'blue', linewidth = 1.5)
                plt.title("Collapsed Difference Image Along Axis 0")
                plt.xlabel("Pixel Column Index (X)")
                plt.ylabel("Mean Pixel Difference")
                plt.grid(True)

                # Save the plot before closing it
                slice_save = os.path.join(
                    difference_images,
                    f'{ref_num}_collapsed_difference_image_axis0.png'
                )
                plt.savefig(slice_save)
                plt.close()

                # Slice through
                row_index = difference_image.shape[0] // 2
                # Slice the difference image at the specific row
                row_difference = difference_image[row_index, :]
                # Generate the x-line (pixel positions along the columns of the
                # selected row)
                x_line = np.arange(difference_image.shape[1])

                # Plotting the pixel differences along the selected row
                plt.figure(figsize=(10, 6))
                plt.plot(x_line, row_difference, color='blue', linewidth=1.5)
                plt.title(f"Difference Image Slice at Row {row_index}")
                plt.xlabel("Pixel Column Index (X)")
                plt.ylabel("Pixel Difference")
                plt.grid(True)

                # Save the plot before closing it
                slice_save = os.path.join(
                    difference_images,
                    f'{ref_num}_slice_row_{row_index}_difference_image.png'
                )
                plt.savefig(slice_save)
                plt.close()

                # Inverse consistency
                if deform_flag['Inverse consistency']:
                    # Reverse the deformation to analyse inverse consistency
                    if k < 999:
                        print('Applying reverse deformation...')
                        reversed_deformation = imgan.img_deform(
                            transformed_image,
                            new_points,
                            original_points, 3
                        )
                        test_image_path = os.path.join(
                            debugg_folder,
                            f'{ref_num}_reversed_deformation.tif')
                        cv2.imwrite(test_image_path, reversed_deformation)

                # Show before and after picture of deformation
                if deform_flag['show process']:
                    # Display the original and transformed images using matplotlib
                    # (with the FE mesh imposed)
                    fig, axs = plt.subplots(2, 1, figsize=(12, 6))
                    axs[0].imshow(reference_image)
                    axs[0].set_title(f"{ref_num}_Generated_spec_image.tif")
                    axs[0].scatter(*zip(*original_points), color='red', s=5)  # Mark original points

                    axs[1].imshow(transformed_image)
                    axs[1].set_title(f"{deform_num}_Generated_spec_image.tif")
                    axs[1].scatter(*zip(*new_points), color='red', s=5)  # Mark new points

                    for ax in axs:
                        ax.axis('on')

                    text = plt.figtext(
                        0.5,
                        0.02,
                        f'First image size: {reference_image.shape}, '
                        f'Second image: {transformed_image.shape}',
                        ha='center'
                    )
                    plt.tight_layout()
                    plt.draw()
                    plt.pause(wait_time)

                    save_path = os.path.join(
                        debugg_folder,
                        f"comparison_{ref_num}_{deform_num}.png"
                    )
                    plt.savefig(save_path, dpi=300, bbox_inches='tight')
                    plt.close(fig)

                print(f'Image {k+1} deformation complete. '
                    f'{toc_def - tic_def:.3f} seconds\n')
                
        elif deform_type == 1:   
            # Performs sub-pixel translations and saves the results in translation
            # folder
            print('Rigid-body translation')
            imgan.subpixel_translation(
                reference_image_path=reference_image_path,
                savepath=spt_img_path,
                shift_method='grid',  # Fourier or Grid
                imagetype=image_type,                 
                umin=u_min,
                umax=u_max,
                intervals=uinterval
            )
        else:
            print('Invalid input')

        plt.close('all')


    '========================RUN DIGITAL IMAGE CORRELATION========================='
    if FLAGS['run_dic']:

        # Toggle between running analysis and/or processing the results
        DIC_flags = {   "Image_blur": False,
                        'sub_pixel_noise': False,
                        "run_analysis": True,
                        "process_subpixel_data": True}
        
        imgan.flag_status(DIC_flags,wait_time=1.5)

        # Move the latest log file to a new location when this section
        # is executed.

        # Add Gaussian blur
        if DIC_flags['Image_blur']:
            print('\nDeformed image prefilter')
            imgan.gaussian_blur_images(reference_image_path, size = 5, sig_y = 1.0, par = 'even')
            imgan.gaussian_blur_images(deformed_image_path, size = 5, sig_y = 1.0, par = 'odd')
            if study_type == 1 and DIC_flags['sub_pixel_noise']:
                imgan.gaussian_blur_images(spt_img_path, size = 5, sig_y = 1.0, par = 'none')
        
        
        print("\n6. Running DIC analysis...\n")

        if not os.path.exists(DIC_settings_path):
            print(f"Settings file not found at: {DIC_settings_path}")
                
        settings = sdset.Settings.fromSettingsFile(DIC_settings_path)

        roi = 'Auto'
        # Automatic ROI size based on image dimensions.
        start_x = changing_parameter[iteration]
        start_y = changing_parameter[iteration]

        if roi.lower() == 'auto':
            width = image_width - (2 * start_x)
            height = image_height - (2 * start_y)
        else:
            width = 1990
            height = 490

        settings.ROI = [start_x, start_y, width, height]
        settings.SubsetSize = 33
        settings.StepSize = 11
        settings.GaussianBlurSize = 5
        settings.GaussianBlurStdDev = 0.0
        settings.DebugLevel = 2
        settings.CPUCount = 6
        settings.OptimizationAlgorithm = "IC-GN"
        settings.ShapeFunctions = 'Affine'
        settings.ReferenceStrategy = 'absolute'

        if study_type == 0:
            Deformed_images = deformed_image_path
            sundic_binary_file_loc = sundic_binary_folder

        elif study_type == 1:
            Deformed_images = spt_img_path
            sundic_binary_file_loc = os.path.join(sundic_binary_folder, "subpixel")
            if not os.path.exists(sundic_binary_file_loc):
                os.makedirs(sundic_binary_file_loc)         # Create folder if it does not exist

        print(settings.__repr__())
        print("Settings loaded successfully\n")

        # Check ray status and create headnode if uninitialised
        try:
            print('Checking ray...')
            subprocess.run('ray status', shell=True, check=True)
            time.sleep(5)
        # If not initialised
        except subprocess.CalledProcessError:
            print('Starting Ray...')
            subprocess.run('ray start --head --num-cpus=8', shell=True)
            time.sleep(5)
            subprocess.run('ray status', shell=True)
        
        
        if DIC_flags["run_analysis"]:

            imgan.run_dic(
                settings,
                reference_image_path,
                Deformed_images,
                sundic_binary_file_loc,
                debugg_folder,
                DIC_contour_path,
                znssd_figure_path,
                study_type,
                image_type,
                start_index=0, 
                umin=u_min, 
                umax=u_max,
                u_interval=uinterval
            )
        
        if study_type == 1 and DIC_flags["process_subpixel_data"]:
            error_matrix,subpixel_trans_figure = imgan.subpixel_analysis(sundic_binary_file_loc,plot_path,
                                                                        u_min=u_min, u_max=u_max,u_interval=uinterval)
            #save matrix
            matrix_save = os.path.join(numpy_files,'error_matrix.npy')
            np.save(matrix_save, error_matrix)

        plt.close('all')


    '=========================ERROR ANALYSIS TO EXCEL=============================='
    if FLAGS['run_error']:

        print('\n7. Running error analysis...')

        # Ensure the Excel directory exists
        if not os.path.exists(excel_path):
            os.makedirs(excel_path)

        # Check for existing Excel files and increment doc_number
        existing_files = [
            f for f in os.listdir(excel_path)
            if f.startswith("Pattern_evaluation_") and f.endswith(".xlsx")
        ]
        doc_number = len(existing_files)

        print("Current doc_number (Error analysis):", doc_number)

        exl_file_name = f"Pattern_evaluation_{doc_number}.xlsx"
        exl_file_path = os.path.join(excel_path, exl_file_name)

        # Load workbook if it exists, otherwise create a new one
        try:
            workbook = openpyxl.load_workbook(exl_file_path)
        except FileNotFoundError:
            workbook = openpyxl.Workbook()

        sheet = workbook.active

        # Write column headers
        sheet['I1'] = 'RMSE_d2f'
        sheet['J1'] = 'MAPE_d2f'
        sheet['K1'] = 'RMSE_f2d'
        sheet['L1'] = 'MAPE_f2d'
        sheet['M1'] = 'MBE_d2f'
        sheet['N1'] = 'MBE_f2d'
        sheet['O1'] = 'SDE_d2f'
        sheet['P1'] = 'SDE_f2d'

        # === FEM Data ===
        fem_xcoord = nodes_2d[:, 0] * x_scale
        fem_ycoord = nodes_2d[:, 1] * x_scale
        fem_x_disp = displacements_2d[:, 0] * x_scale  # x-direction displacement
        fem_y_disp = displacements_2d[:, 1] * x_scale
        # Compute magnitude of displacement
        fem_mag = np.sqrt(fem_x_disp**2 + fem_y_disp**2)

        # === Interpolation and Error Analysis ===
        # FEA coordinate array
        lag_dx = FEx_rbf_interp(np.column_stack([fem_xcoord, fem_ycoord]))
        lag_dy = FEy_rbf_interp(np.column_stack([fem_xcoord, fem_ycoord]))

        fem_xcoord_lag = fem_xcoord - lag_dx
        fem_ycoord_lag = fem_ycoord - lag_dy

        fem_points = np.column_stack((fem_xcoord_lag, fem_ycoord_lag)) 
        # fem_points = np.column_stack((fem_xcoord, fem_ycoord))

        # Collect all valid .sdic files in numerical order
        sundic_binary_files = sorted(
            [f for f in os.listdir(sundic_binary_folder) if f.endswith('results.sdic')],
            key=lambda x: int(x.split('_')[0])
        )
        print('\nSundic binary files:', sundic_binary_files)

        # Map file prefixes to Excel row positions
        all_expected_prefixesbin = imgan.expected_prefixes(sundic_binary_folder)
        prefix_positionsbin = {
            prefix: i for i, prefix in enumerate(all_expected_prefixesbin)
        }

        for prefix_num in all_expected_prefixesbin:
            m = prefix_positionsbin[prefix_num] + 2

            print(f'\nCurrent prefix number is {prefix_num}')
            print(f'Current row number is {m}')

            sunfile = f'{prefix_num}_results.sdic'
            sundic_data_path = os.path.join(sundic_binary_folder, sunfile)

            print(f'Reading DIC data: file {sundic_data_path}')

            try:
                if not os.path.exists(sundic_data_path):
                    raise Exception("Sundic.sdic file not found. Printing zeros")

                # Read DIC displacement data
                sundic_data, nRows, nCols = sdpp.getDisplacements(
                                                    sundic_data_path,
                                                    -1,
                                                    smoothWindow=5
                                                )

                nan_count = np.isnan(sundic_data).any(axis=1).sum()
                total_points = len(sundic_data)
                nan_ratio = nan_count / total_points
                nan_percentage = nan_ratio * 100

                print(f"NaN-to-total points ratio: {nan_ratio:.2%}")

                if nan_percentage >= 15:
                    raise Exception(f"NaN percentage too high: {nan_percentage:.4f}")

                # Extract and filter displacement components
                x_coord, y_coord = sundic_data[:, 0], sundic_data[:, 1]
                X_disp, Y_disp = sundic_data[:, 3], sundic_data[:, 4]
                dic_mag = np.sqrt(X_disp**2 + Y_disp**2)

                threshold = 15000
                filtered_indices = dic_mag < threshold

                x_coord_filtered = x_coord[filtered_indices]
                y_coord_filtered = y_coord[filtered_indices]
                X_disp_filtered = X_disp[filtered_indices]
                Y_disp_filtered = Y_disp[filtered_indices]
                dic_mag_filtered = dic_mag[filtered_indices]

                sundic_points = np.column_stack((x_coord_filtered, y_coord_filtered))

                # Set which displacement quantity to compare
                value = 2  # 0: X, 1: Y, 2: Magnitude

                if value == 0:
                    fem_value = fem_x_disp
                    dic_value = X_disp_filtered
                elif value == 1:
                    fem_value = fem_y_disp
                    dic_value = Y_disp_filtered
                elif value == 2:
                    fem_value = fem_mag
                    dic_value = dic_mag_filtered
                else:
                    raise ValueError("Invalid value")

                # Interpolate FEM displacements on DIC grid
                #------------------------------------------------------------------
                interpolated_FEM_values = griddata(
                    fem_points, 
                    fem_value, 
                    sundic_points, 
                    method='cubic')
                
                # https://stackoverflow.com/questions/2831516/isnotnan-functionality-in-numpy-can-this-be-more-pythonic
                valid_mask = ~np.isnan(interpolated_FEM_values) & ~np.isnan(dic_value)
                valid_interpolated_FEM = interpolated_FEM_values[valid_mask]
                valid_dic_data = dic_value[valid_mask]

                # Match lengths for FEM to DIC comparison (truncation)
                min_length = min(len(valid_interpolated_FEM), len(valid_dic_data))
                # Truncate to the smaller array
                valid_interpolated_FEM = valid_interpolated_FEM[:min_length]
                valid_dic_data = valid_dic_data[:min_length]

                # Compute error
                err_on_dic = valid_dic_data - valid_interpolated_FEM

                # Mean Absolute Percentage Error
                data = err_on_dic[~np.isnan(err_on_dic)].flatten()
                IQR_f2d = stats.iqr(data)

                # Error calculations per literature definitions
                MBE_f2d = np.mean(err_on_dic)
                SDE_f2d = np.sqrt(np.sum((err_on_dic - MBE_f2d) ** 2) / (len(err_on_dic) - 1))
                RMSE_f2d = np.sqrt(MBE_f2d**2 + SDE_f2d**2)

                # Interpolate DIC displacements on FEM grid
                #------------------------------------------------------------------
                interpolated_DIC_values = griddata(
                    sundic_points, 
                    dic_value, 
                    fem_points, 
                    method='cubic')

                # Handle NaN values in interpolated_DIC_values
                valid_mask = ~np.isnan(interpolated_DIC_values) & ~np.isnan(fem_value)
                valid_interpolated_DIC = interpolated_DIC_values[valid_mask]
                valid_fem_x_disp = fem_value[valid_mask]

                # Match lengths for DIC to FEM comparison (truncation)
                min_length = min(len(valid_interpolated_DIC), len(valid_fem_x_disp))
                valid_interpolated_DIC = valid_interpolated_DIC[:min_length]
                valid_fem_x_disp = valid_fem_x_disp[:min_length]

                # Compute error
                err_on_fem = valid_interpolated_DIC - valid_fem_x_disp

                # Error metrics
                MBE_d2f = np.mean(err_on_fem)
                SDE_d2f = np.sqrt(np.sum((err_on_fem - MBE_d2f) ** 2) / (len(err_on_fem) - 1))
                RMSE_d2f = np.sqrt(MBE_d2f**2 + SDE_d2f**2)

                # Interquartile range  - 14/09/2025
                data2 = err_on_fem[~np.isnan(err_on_fem)].flatten()
                IQR_d2f = stats.iqr(data2)

                

                # Save results to Excel document
                sheet[f'I{m}'] = RMSE_d2f
                sheet[f'J{m}'] = IQR_d2f
                sheet[f'K{m}'] = RMSE_f2d
                sheet[f'L{m}'] = IQR_f2d
                sheet[f'M{m}'] = MBE_d2f
                sheet[f'N{m}'] = MBE_f2d
                sheet[f'O{m}'] = SDE_d2f
                sheet[f'P{m}'] = SDE_f2d

                # Add NAN percentage
                sheet[f'Z{m}'] = nan_percentage

            except Exception as e:
                # If file is missing or error occurs, write zeroes
                sheet[f'I{m}'] = 0
                sheet[f'J{m}'] = 0
                sheet[f'K{m}'] = 0
                sheet[f'L{m}'] = 0
                sheet[f'M{m}'] = 0
                sheet[f'N{m}'] = 0
                sheet[f'O{m}'] = 0
                sheet[f'P{m}'] = 0
                # Add NAN percentage
                sheet[f'Z{m}'] = 0 
                print(f'Exception raised: Writing zeros to row {m}')
                message = (
                    f'Error reading {sundic_data_path}.\n'
                    f'Error details: {e}\nContinuing...\n'
                )

        # Save the updated workbook
        workbook.save(exl_file_path)
        print(f"Workbook saved at {exl_file_path}")

    '=======================ERROR DISTRIBUTION ANALYSIS============================'
    if FLAGS['error_dist']:
        print('\n8. Error distribution analysis...\n')
        # Loop through the numpy files (from the DIC analysis) and perform error analysis.
        # Ignore files that have f.split('_')[1] == '_T' as these are the subpixel translations

        # DIC binary files
        sundic_binary_files = sorted(
            [f for f in os.listdir(sundic_binary_folder)
            if f.endswith('results.sdic')],
            key=lambda x: int(x.split('_')[0])
        )
        print(sundic_binary_files)
        # Get prefixes for id/indexing purposes
        all_expected_prefixesbin = imgan.expected_prefixes(sundic_binary_folder,odd=False,skip=True)
        prefix_positionsbin = {prefix: i for i, prefix in enumerate(all_expected_prefixesbin)}

        # FE data
        fem_xcoord = nodes_2d[:, 0] * x_scale
        fem_ycoord = nodes_2d[:, 1] * x_scale
        fem_x_disp = displacements_2d[:, 0] * x_scale  # x-direction displacement
        fem_y_disp = displacements_2d[:, 1] * x_scale
        fem_mag = np.sqrt(fem_x_disp**2 + fem_y_disp**2)
        lag_dx = FEx_rbf_interp(np.column_stack([fem_xcoord, fem_ycoord]))
        lag_dy = FEy_rbf_interp(np.column_stack([fem_xcoord, fem_ycoord]))

        fem_xcoord_lag = fem_xcoord - lag_dx
        fem_ycoord_lag = fem_ycoord - lag_dy

        fem_points = np.column_stack((fem_xcoord_lag, fem_ycoord_lag))  
        # fem_points = np.column_stack((fem_xcoord, fem_ycoord))
        print(f'FEM points shape = {fem_points.shape}')

        # Array for storing the overall mean error value for each file.
        print(f'\nnumpy files found: {len(sundic_binary_files)}')
        meanmean = np.zeros(len(sundic_binary_files))
        print(f'\nMean array generated {meanmean}')

        # Initialising i manually. Struggling with enumerate() for some reason
        i = 0
        for prefix_num in all_expected_prefixesbin:
            try:

                print(f'\nCurrent prefix number is {prefix_num}')

                # Expected file name
                sunfile = f'{prefix_num}_results.sdic'
                file_number = prefix_num
                sundic_data_path = os.path.join(sundic_binary_folder, sunfile)
                print(f'Reading DIC data: file {sundic_data_path}')

                if not os.path.exists(sundic_data_path):
                    print(f"File path: {sundic_data_path} not found. Moving to next prefix\n")
                    continue

                # Load path and file name for DIC data
                sundic_data, nRows, nCols = sdpp.getDisplacements(sundic_data_path,-1, smoothWindow=25)

                # sundic_data = np.load(sundic_data_path)
                x_coord, y_coord = sundic_data[:, 0], sundic_data[:, 1]
                X_disp, Y_disp = sundic_data[:, 3], sundic_data[:, 4]
                dic_mag = np.sqrt(X_disp**2 + Y_disp**2)

                sundic_points = np.column_stack((x_coord,y_coord))
                print(f'DIC points shape = {sundic_points.shape}')
                print(f'reshaped DIC data (Z) = {(dic_mag.reshape(nCols,nRows)).shape}')

                # === Interpolation and Error Analysis ===
                value = 0
                if value == 0:
                    fem_value = fem_x_disp
                    dic_value = X_disp
                    string = 'X_disp'
                elif value == 1:
                    fem_value = fem_y_disp
                    dic_value = Y_disp
                    string = 'Y_disp'
                elif value == 2:
                    fem_value = fem_mag
                    dic_value = dic_mag
                    string = 'Magnitude'
                else:
                    raise ValueError("Invalid value")

                # Only interpolate FE-grid and subtract DIC data
                #----------------------------------------------------
                interpolated_FEM_values = griddata(fem_points, fem_value, sundic_points, method='cubic')

                # Reshape data
                dic_value = dic_value.reshape(nCols,nRows)
                interpolated_FEM_values = interpolated_FEM_values.reshape(nCols,nRows)

                errors2 = interpolated_FEM_values - dic_value
                numpath = os.path.join(numpy_files,f"{file_number}_errors.npy")         # For KD
                np.save(numpath, errors2)
                #----------------------------------------------------

                print('errors grid shape = ',errors2.shape)
                mean_err = np.nanmean(errors2)
                print(f'Mean error = {mean_err:.4f}')

                meanmean[i] = mean_err
                i = i + 1

                '''
                The higher and the narrower the central peak, the better the correspondence between
                the imposed and the calculated displacements, thus the more accurate the results are.

                All the images are numerically treated in the same way. The difference in
                displacements is, therefore, only related to the speckle morphology.
                '''
                #-------------------------------------------------------------------------------------------------
                # Create and save heatmap
                plt.figure(figsize=(10, 8))
                # errors2 = np.nan_to_num(errors2, nan=0.0) 

                # plt.imshow(errors2.T, cmap='jet', interpolation='none', 
                #     extent=(x_coord.min(), x_coord.max(), y_coord.min(), y_coord.max()), 
                #     origin='lower', vmax=0.035,vmin=-0.035)
                
                plt.imshow(errors2.T, cmap='jet', interpolation='none', 
                    extent=(x_coord.min(), x_coord.max(), y_coord.min(), y_coord.max()), 
                    origin='lower')
                
                plt.colorbar(label=f'Error {string}')
                plt.title(f'{string} Error Distribution: Pattern {file_number}')
                plt.xlabel('X')
                plt.ylabel('Y')
                plt.gca().invert_yaxis()            # Invert y to match image

                text = plt.figtext(0.5, 0.02, f'Mean error: {mean_err:.4f}', size=14)
                Heatmap_path = os.path.join(error_heatmap_path, f'{file_number}_heatmap_{string}.png')
                plt.savefig(Heatmap_path)
                plt.close()

                #-------------------------------------------------------------------------------------------------
                # Create and save histogram
                plt.figure(figsize=(12, 10))
                plt.hist(errors2[~np.isnan(errors2)].flatten(), bins=250, density=True)  # Flatten and use only non-NaN values
                plt.title(f'Error Distribution: Pattern {file_number}', fontsize=22)
                plt.xlabel(f'Error {string}', fontsize=24)
                plt.ylabel('Frequency', fontsize=24)
                plt.tick_params(axis='both', which='major', labelsize=20)
                plt.xlim(-0.05, 0.05)
                plt.ylim(0, 300)
                histogrampath = os.path.join(error_hist_path, f'{file_number}_histogram_{string}.png')
                plt.savefig(histogrampath)
                plt.close()

                #-------------------------------------------------------------------------------------------------
                # Collapse  grid
                x_grid = x_coord.reshape(nCols, nRows)
                collapsed_errorgrid = np.nanmean(errors2, axis=1)
                x_line = np.mean(x_grid, axis=1)
                print(f'\nX_line min = {np.min(x_line)}\nX_line max = {np.max(x_line)}')

                print(f'\nCollapsed shape = {collapsed_errorgrid.shape}\nx_line shape = {x_line.shape}')
                
                # Save numpy file to reload for processing
                slice_numpy = np.column_stack((collapsed_errorgrid,x_line))
                print(f'slice data = {slice_numpy.shape}')

                path_to_slice_bin = r"output\Slices\slice_binaries"
                if not os.path.exists(path_to_slice_bin):
                    os.makedirs(path_to_slice_bin)
                
                save_collape = os.path.join(path_to_slice_bin,f'{file_number}_slice_{string}.npy')
                np.save(save_collape,slice_numpy)

                plt.figure(figsize=(10, 6))
                plt.plot(x_line,collapsed_errorgrid, color='blue', linewidth=1.5)
                plt.title("Collapsed Error Grid")
                plt.xlabel("Pixels")
                plt.ylabel(f"Error {string}")
                plt.ylim(-0.25, 0.25)
                plt.grid(True)
                slice_save = os.path.join(slice_path, f'{file_number}_slice_{string}.png')
                plt.savefig(slice_save)
                plt.close()
                #-------------------------------------------------------------------------------------------------

            except Exception as e:
                # Extract from traceback object
                tb = traceback.extract_tb(sys.exc_info()[2])
                filename, line_number, function_name, text = tb[-1]  # Last traceback entry
                print(f'Error with file: {sundic_data_path},\nMessage {str(e)}\nLine: {line_number}')    

            # print('Mean batch error:', np.mean(meanmean[:]))


    '=================================READ EXCEL==================================='
    if FLAGS['Excel plots']:

        # Get excel data objects
        print('\n8. Reading excel document...')
        temp_excel_path = r'output\excel_docs\excel_6'

        #-------------------------------------------------------
        p_metrics,meas_error,p_param,nans,indicators = imgan.read_spec_excel(excel_path, doc_num=None)
        # p_metrics,meas_error,p_param,nans,indicators = imgan.read_spec_excel(temp_excel_path, doc_num=1)

        #-------------------------------------------------------
        print('\nIndicators:',indicators)
        # All masks and filters defined here
        # Filter if discrepency between DIC2FE and FE2DIC errors are too large
        threshold = 0.001
        outlier_idx = imgan.array_difference_outlier(meas_error[:,0], meas_error[:,2], threshold)
        mask = np.array([i not in outlier_idx for i in range(meas_error.shape[0])]) & (meas_error[:,0] != 0)
        # Filter out zero values before finding min/max
        nonzero_mask = meas_error[:,0] != 0

        # Plots and analysis
        figflag = {'2d scatter': True,
                '3d scatter': False,
                'xtreme_index': False}

        # 2D scatter plot
        if figflag['2d scatter']:
            metric_strings = ("Mean subset fluctuation (MSF)", "Mean intensity gradient (MIG)", "E_f", "Mean intensity of the second derivative (MIOSD)",
                            "Shannon entropy", "Power area", "SSSIG", "Autocorrelation peak radius")
            save_fig = True

            # Plot details
            for number,metric in enumerate(metric_strings):

                plot_title = f"Normal Perlin images (n = {len(meas_error[:,0])})"
                x_value =  p_metrics[mask, number]
                # x_value = meas_error[mask,4]
                x_label = f'{metric}'
                y_value = meas_error[mask,0]
                y_label = 'RMS error'

                # Correlation coefficient
                r = np.corrcoef(x_value, y_value)
                plt.figure(figsize=(7,6))
                plt.scatter(x_value,y_value, color='black')
                plt.xlabel(x_label)
                plt.ylabel(y_label)
                plt.title(plot_title)
                plt.text(
                    0.95, 1.02, 
                    f'R = {r[0,1]:.2f}', 
                    transform=plt.gca().transAxes, 
                    ha='right'
                )
                plt.ylim(0, 0.12)  # Set y-axis limit
                plt.grid('on')    
                plt.draw()
                plt.get_current_fig_manager().window.raise_()
                if save_fig:

                    # batch_scatter = r"C:\Users\General User\nokop\pattern2\output\plots\plots_3"
                    # old plot_path -> plot_path
                    plot_save = os.path.join(scatter_plots,f'{plot_title}_{x_label}vs{y_label}.png')
                    plt.savefig(plot_save)

                plt.pause(3)
                plt.close()

        # # 3D plot 
        if figflag['3d scatter']:

            # Size vs density vs RMSE
            fig = plt.figure()
            ax = fig.add_subplot(projection='3d')
            scatter = ax.scatter(p_metrics[mask,1], p_metrics[mask,4], meas_error[mask,0], c=meas_error[mask,0], cmap='jet', marker='o')
            colorbar = plt.colorbar(scatter, ax=ax, label='RMSE (DIC)')
            ax.set_xlabel('Speckle size')
            ax.set_ylabel('Speckle density')
            ax.set_zlabel('RMSE')
            ax.set_title('RMSE vs speckle size vs speckle density')
            ax.axes.set_zlim3d(bottom=-0.000,top=0.005)
            plt.show()

        # Get locations of the highest and lowest values in the defined matrices
        if figflag['xtreme_index']:

            # Locations of the extreme points in the set
            maxmsf = np.max(p_metrics[:,0])
            maxmig = np.max(p_metrics[:,1]) 
            maxef = np.max(p_metrics[:,2])
            maxmiosd = np.max(p_metrics[:,3])
            maxshannon = np.max(p_metrics[:,4])
            maxse = np.max(p_metrics[:,5])
            maxsssig = np.max(p_metrics[:,6])
            max_R = np.max(p_metrics[:,7])

            # Get indices of maximum values
            # np.where returns a tuple of arrays; get the first matching index using [0][0]
            maxmsf_idx = np.where(p_metrics[:,0] == maxmsf)[0][0]
            maxmig_idx = np.where(p_metrics[:,1] == maxmig)[0][0]
            maxef_idx = np.where(p_metrics[:,2] == maxef)[0][0] 
            maxmiosd_idx = np.where(p_metrics[:,3] == maxmiosd)[0][0]
            maxshannon_idx = np.where(p_metrics[:,4] == maxshannon)[0][0]
            maxse_idx = np.where(p_metrics[:,5] == maxse)[0][0]
            maxsssig_idx = np.where(p_metrics[:,6] == maxsssig)[0][0]
            max_R_idx = np.where(p_metrics[:,7] == max_R)[0][0]

            print(f'\nMaximum metric values:')
            print(f'MSF: Pattern {maxmsf_idx * 2}')
            print(f'MIG: Pattern {maxmig_idx * 2}')
            print(f'Ef: Pattern {maxef_idx * 2}')
            print(f'MIOSD: Pattern {maxmiosd_idx * 2}')
            print(f'Shannon: Pattern {maxshannon_idx * 2}')
            print(f'Power area: Pattern {maxse_idx * 2}')
            print(f'SSSIG: Pattern {maxsssig_idx * 2}')
            print(f'Peak radius: Pattern {max_R_idx * 2}')

            print(f'\nMax MSF: {maxmsf:.3f}\nMax MIG: {maxmig:.3f}\nMax Ef: {maxef:.3f}\nMax MIOSD: {maxmiosd:.3f}\nMax Shannon: {maxshannon:.3f}\nMax Power area: {maxse:.3f}\nMax SSSIG: {maxsssig:.3f}')
            print(f'Max R: {max_R:.3f}')
            print('---------------------------------------------------------------')

            # Minimum values
            minmsf = np.min(p_metrics[:,0])
            minmig = np.min(p_metrics[:,1]) 
            minef = np.min(p_metrics[:,2])
            minmiosd = np.min(p_metrics[:,3])
            minshannon = np.min(p_metrics[:,4])
            minse = np.min(p_metrics[:,5])
            minsssig = np.min(p_metrics[:,6])
            min_R = np.min(p_metrics[:,7])

            # Get indices of minimum values
            minmsf_idx = np.where(p_metrics[:,0] == minmsf)[0][0]
            minmig_idx = np.where(p_metrics[:,1] == minmig)[0][0]
            minef_idx = np.where(p_metrics[:,2] == minef)[0][0] 
            minmiosd_idx = np.where(p_metrics[:,3] == minmiosd)[0][0]
            minshannon_idx = np.where(p_metrics[:,4] == minshannon)[0][0]
            minse_idx = np.where(p_metrics[:,5] == minse)[0][0]
            minsssig_idx = np.where(p_metrics[:,6] == minsssig)[0][0]
            min_R_idx = np.where(p_metrics[:,7] == min_R)[0][0]

            print(f'\nMinimum metric values:')
            print(f'MSF: Pattern {minmsf_idx * 2}')
            print(f'MIG: Pattern {minmig_idx * 2}')
            print(f'Ef: Pattern {minef_idx * 2}')
            print(f'MIOSD: Pattern {minmiosd_idx * 2}')
            print(f'Shannon: Pattern {minshannon_idx * 2}')
            print(f'Power area: Pattern {minse_idx * 2}')
            print(f'SSSIG: Pattern {minsssig_idx * 2}')
            print(f'Peak radius: Pattern {min_R_idx * 2}')

            print(f'\nMin MSF: {minmsf:.3f}\nMin MIG: {minmig:.3f}\nMin Ef: {minef:.3f}\nMin MIOSD: {minmiosd:.3f}\nMin Shannon: {minshannon:.3f}\nMin Power area: {minse:.3f}\nMin SSSIG: {minsssig:.3f}')
            print(f'Min R: {min_R:.3f}')
            print('---------------------------------------------------------------')


            # Get error extremes
            #-------------------------------------------------------------------------
            # Locations of patterns with minimum and maximum RMSE, MBE and SDE
            filtered_error = meas_error[nonzero_mask]

            # Filter out None or NaN values from excel
            non_empty_mask = np.isfinite(filtered_error[:,0])
            ffiltered_error = filtered_error[non_empty_mask]

            minrmse = np.min(ffiltered_error[:,0])
            minsde = np.min(ffiltered_error[:,6]) 
            maxrmse = np.max(ffiltered_error[:,0])
            maxsde = np.max(ffiltered_error[:,6])

            # Find indices in the original array
            minrmse_idx = np.where(meas_error[:,0]==minrmse)[0][0]
            minsde_idx = np.where(meas_error[:,6]==minsde)[0][0]
            maxrmse_idx = np.where(meas_error[:,0]==maxrmse)[0][0]
            maxsde_idx = np.where(meas_error[:,6]==maxsde)[0][0]

            print(f'\nMinimum and maximum error indices:')
            print(f'Min RMSE: Pattern {minrmse_idx * 2}')
            print(f'Min SDE: Pattern {minsde_idx * 2}')
            print(f'Max RMSE: Pattern {maxrmse_idx * 2}')
            print(f'Max SDE: Pattern {maxsde_idx * 2}')

            print(f'\nMin RMSE: {minrmse:.6f}\nMin SDE: {minsde:.6f}')
            print(f'Max RMSE: {maxrmse:.6f}\nMax SDE: {maxsde:.6f}')
            print('---------------------------------------------------------------')


    "====================Polyfit + constrained surrogate model====================="
    if FLAGS['Optimisation_poly']:

        print('\n9. Optimising speckle patterns...')
        # Uses an RBF based surrogate model
        if not FLAGS['Excel plots']:
            print('\nError: Excel document not loaded.\nCannot run optimisation with multiple starting points')
            print('Set FLAGS["Excel plots"] = True\n')
            sys.exit()

        '''
        Excel document will be used for accessing filled-in values to act as training data 
        for optimisation. Will use maxrow from previous section.
        '''
        # Read second last column
        '''
        speckle
        lines
        checkerboard
        perlin
        '''
        #-----------------------------------------------------
        override_optimisation = False

        if override_optimisation:
            gen_value = 'speckle'
        else:
            gen_value  = indicators[0, 0]
        #-----------------------------------------------------

        print('Gen value =', gen_value)
        # z = meas_error[:,0]
        z = p_metrics[:,1]
        optimization_direction  = -1            # -1 for maximisation

        if optimization_direction  == 1:
            print('\nRunning minimisation\n---------------------------------------')
        else:
            print('\nRunning maximisation\n---------------------------------------')

        numb = 69
        polyy = 3
        opt_file_name = f'{numb}_max_MIG_POLY_{polyy}'

        #------------------------------------------------------------
        if gen_value.lower() == 'speckle':
            # Coordinate array
            x_1 = p_param[:,0]      # Speckle diameter
            x_2 = p_param[:,1]      # Speckle density
            x_3 = p_param[:,2]      # Position randomness
            x_4 = p_param[:,3]      # Grid-step
            x_5 = p_param[:,4]      # Size randomness
            # For nans
            x_6 = nans[:,0]

            coordinate_arr = np.vstack((x_1,x_3,x_4,x_5)).T

            # RBFInterpolator surrogate model
            poly = PolynomialFeatures(degree=polyy)  
            X_poly = poly.fit_transform(coordinate_arr)
            poly_model = LinearRegression()
            poly_model.fit(X_poly, z)
            
            # Surrage models here
            # results = imgan.surrogate(X,model = 1)
            def interp_value(X):
                '''
                Polynomial interpolation for optimization
                '''
                # For multiple points evaluation
                X_poly = poly.transform(X)
                results = poly_model.predict(X_poly)
                return results  

            # Define constraint inequality
            nan_model = LinearRegression().fit(coordinate_arr, x_6)

            def g1(x):
                nan_pred = nan_model.predict(x.reshape(1, -1))[0]
                nan_pred = np.clip(nan_pred, 0, 100)
                return 5 - nan_pred 
                
            # Bounds
            x1_max = np.max(x_1)
            x1_min = np.min(x_1)
            x2_max = np.max(x_2)
            x2_min = np.min(x_2)
            x3_max = np.max(x_3)
            x3_min = np.min(x_3)
            x4_max = np.max(x_4)
            x4_min = np.min(x_4)
            x5_max = np.max(x_5)
            x5_min = np.min(x_5)
            bounds = [(x1_min, x1_max), (x3_min, x3_max), (x4_min, x4_max), (x5_min, x5_max)]

            def objective(x):
                # convert the 1D array to a 2D array 
                return optimization_direction * interp_value(x.reshape(1, -1))[0]
                    
            best_result = None
            # Number of entries = number of starting points
            for i in range(len(p_param[:,0])):

                # Skipping the middle parameter because it must remain constant
                initial_guess = np.array([p_param[i,0],p_param[i,2],p_param[i,3],p_param[i,4]])
                cons = [{'type': 'ineq', 'fun': g1}]
                # Supress runtime warnings
                with warnings.catch_warnings():
                    warnings.filterwarnings("ignore", category=RuntimeWarning)

                    result = minimize(objective, initial_guess, method='SLSQP', bounds=bounds,
                            constraints=cons, options={'ftol': 1e-8, 'maxiter': 1000})
            
                if best_result is None or result.fun <= best_result.fun:               
                    best_result = result

            if best_result is None:
                raise RuntimeError("Optimization failed for all initial guesses.")
            
            print(f'\n\n SLSQP result: {best_result}')
            solution = best_result.x
            print("Constraint value (should be >= 0):", g1(best_result.x))

            # Generate and save
            custom_name = 'speckles.tif'
            filee_name = '_'.join([opt_file_name,custom_name])
            spec_opt_file_path = os.path.join(optimised_save,filee_name)
            Generated_speckles = generate_and_save(500, 2000, 25.4 , solution[0],
                            spec_opt_file_path,size_randomness=solution[3],position_randomness=solution[1], speckle_blur=1.5, grid_step=solution[2])
            
            plt.imshow(Generated_speckles, cmap='gray')
            plt.title(f'Optimised pattern: {filee_name}')
            plt.show()
            plt.get_current_fig_manager().window.raise_()

        #------------------------------------------------------------
        elif gen_value.lower() == 'lines':
            print('\nStraightline pattern optimisation coming soon...')
            # Coordinate array
            x_1 = p_param[:,0]      # Ignore for now. This is dpi. Opted to keep it constant
            x_2 = p_param[:,1]      # line width
            x_3 = p_param[:,2]      # Number of lines
            # For nans
            x_6 = nans[:,0]
            '''
            Dimensional changes: coordinate_arr
                                bounds
                                initial_guess
            '''
            coordinate_arr = np.vstack((x_2,x_3)).T
            # RBFInterpolator surrogate model
            poly = PolynomialFeatures(degree=polyy)  
            X_poly = poly.fit_transform(coordinate_arr)
            poly_model = LinearRegression()
            poly_model.fit(X_poly, z)
            # Surrage models here
            # results = imgan.surrogate(X,model = 1)
            def interp_value(X):
                '''
                Polynomial interpolation for optimization
                '''
                # For multiple points evaluation
                X_poly = poly.transform(X)
                results = poly_model.predict(X_poly)
                return results  
            # Define constraint inequality
            nan_model = LinearRegression().fit(coordinate_arr, x_6)
            def g1(x):
                nan_pred = nan_model.predict(x.reshape(1, -1))[0]
                nan_pred = np.clip(nan_pred, 0, 100)
                return 5 - nan_pred 
                
            # Bounds
            x2_max = np.max(x_2)
            x2_min = np.min(x_2)
            x3_max = np.max(x_3)
            x3_min = np.min(x_3)

            bounds = [(x2_min, x2_max), (x3_min, x3_max)]

            def objective(x):
                # convert the 1D array to a 2D array 
                return optimization_direction * interp_value(x.reshape(1, -1))[0]
                    
            best_result = None
            # Number of entries = number of starting points
            for i in range(len(p_param[:,0])):

                # Skipping the middle parameter because it must remain constant
                initial_guess = np.array([p_param[i,1],p_param[i,2]])
                cons = [{'type': 'ineq', 'fun': g1}]
                # Supress runtime warnings
                with warnings.catch_warnings():
                    warnings.filterwarnings("ignore", category=RuntimeWarning)

                    result = minimize(objective, initial_guess, method='SLSQP', bounds=bounds,
                            constraints=cons, options={'ftol': 1e-8, 'maxiter': 1000})
            
                if best_result is None or result.fun <= best_result.fun:               
                    best_result = result

            if best_result is None:
                raise RuntimeError("Optimization failed for all initial guesses.")
            
            print(f'\n\n SLSQP result: {best_result}')
            solution = best_result.x
            print("Constraint value (should be >= 0):", g1(best_result.x))

            # Image generation and save
            # Generate and save
            custom_name = 'lines.tif'
            filee_name = '_'.join([opt_file_name,custom_name])
            lines_opt_file_path = os.path.join(optimised_save, filee_name)
            generate_lines(image_height, image_width, 25.4, solution[1], lines_opt_file_path,
                            orientation='vertical', N_lines=solution[2] )
            generated_lines = cv2.imread(lines_opt_file_path)
            plt.imshow(generated_lines, cmap='gray')
            plt.title(f'Optimised pattern: {filee_name}')
            plt.show()
            plt.get_current_fig_manager().window.raise_()

        #------------------------------------------------------------
        elif gen_value.lower() == 'checkerboard': 
            print('\nCheckerboard optimisation coming soon...')
            # Coordinate array
            x_1 = p_param[:,0]      # Line width
            x_2 = p_param[:,1]      # rows

            # For nans
            x_6 = nans[:,0]

            coordinate_arr = np.vstack((x_1,x_2)).T

            # RBFInterpolator surrogate model
            poly = PolynomialFeatures(degree=polyy)  
            X_poly = poly.fit_transform(coordinate_arr)
            poly_model = LinearRegression()
            poly_model.fit(X_poly, z)
            
            # Surrage models here
            # results = imgan.surrogate(X,model = 1)
            def interp_value(X):
                '''
                Polynomial interpolation for optimization
                '''
                # For multiple points evaluation
                X_poly = poly.transform(X)
                results = poly_model.predict(X_poly)
                return results  

            # Define constraint inequality
            nan_model = LinearRegression().fit(coordinate_arr, x_6)

            def g1(x):
                nan_pred = nan_model.predict(x.reshape(1, -1))[0]
                nan_pred = np.clip(nan_pred, 0, 100)
                return 5 - nan_pred 
                
            # Bounds
            x1_max = np.max(x_1)
            x1_min = np.min(x_1)
            x2_max = np.max(x_2)
            x2_min = np.min(x_2)

            bounds = [(x1_min, x1_max), (x2_min, x2_max)]

            def objective(x):
                # convert the 1D array to a 2D array 
                return optimization_direction * interp_value(x.reshape(1, -1))[0]
                    
            best_result = None
            # Number of entries = number of starting points
            for i in range(len(p_param[:,0])):

                # Skipping the middle parameter because it must remain constant
                initial_guess = np.array([p_param[i,0],p_param[i,1]])
                cons = [{'type': 'ineq', 'fun': g1}]
                # Supress runtime warnings
                with warnings.catch_warnings():
                    warnings.filterwarnings("ignore", category=RuntimeWarning)

                    result = minimize(objective, initial_guess, method='SLSQP', bounds=bounds,
                            constraints=cons, options={'ftol': 1e-8, 'maxiter': 1000})
            
                if best_result is None or result.fun <= best_result.fun:               
                    best_result = result

            if best_result is None:
                raise RuntimeError("Optimization failed for all initial guesses.")
            
            print(f'\n\n SLSQP result: {best_result}')
            solution = best_result.x
            print("Constraint value (should be >= 0):", g1(best_result.x))

            # Generate and save
            custom_name = 'checkb.tif'
            filee_name = '_'.join([opt_file_name,custom_name])
            cb_opt_file_path = os.path.join(optimised_save, filee_name)
            generate_checkerboard(image_height, image_width, dpi = 25.4,path=cb_opt_file_path, 
                                line_width=solution[0], N_rows=solution[1])
            
            generated_cb = cv2.imread(cb_opt_file_path)
            plt.imshow(generated_cb, cmap='gray')
            plt.title(f'Optimised pattern: {filee_name}')
            plt.show()
            plt.get_current_fig_manager().window.raise_()

        #------------------------------------------------------------
        elif gen_value.lower() == 'perlin':

            # Coordinate array
            x_1 = p_param[:,0]      # Scale
            x_2 = p_param[:,1]      # Octaves
            x_3 = p_param[:,2]      # Persistence
            x_4 = p_param[:,3]      # Lacunarity
            # For nans
            x_6 = nans[:,0]         # Always measured

            coordinate_arr = np.vstack((x_1,x_2,x_3,x_4)).T
            
            # RBFInterpolator surrogate model
            poly = PolynomialFeatures(degree=polyy)  
            X_poly = poly.fit_transform(coordinate_arr)
            poly_model = LinearRegression()
            poly_model.fit(X_poly, z)
            
            # Surrage models here
            # results = imgan.surrogate(X,model = 1)
            def interp_value(X):
                '''
                Polynomial interpolation for optimization
                '''
                # For multiple points evaluation
                X_poly = poly.transform(X)
                results = poly_model.predict(X_poly)
                return results  

            # Define constraint inequality
            nan_model = LinearRegression().fit(coordinate_arr, x_6)

            def g1(x):
                nan_pred = nan_model.predict(x.reshape(1, -1))[0]
                nan_pred = np.clip(nan_pred, 0, 100)
                return 5 - nan_pred 
                
            # Bounds
            x1_max = np.max(x_1)
            x1_min = np.min(x_1)
            x2_max = np.max(x_2)
            x2_min = np.min(x_2)
            x3_max = np.max(x_3)
            x3_min = np.min(x_3)
            x4_max = np.max(x_4)
            x4_min = np.min(x_4)

            bounds = [(x1_min, x1_max), (x2_min, x2_max), (x3_min, x3_max) , (x4_min, x4_max)]

            def objective(x):
                # convert the 1D array to a 2D array 
                return optimization_direction * interp_value(x.reshape(1, -1))[0]
                    
            best_result = None
            # Number of entries = number of starting points
            for i in range(len(p_param[:,0])):

                initial_guess = np.array([p_param[i,0],p_param[i,1],p_param[i,2],p_param[i,3]])
                cons = [{'type': 'ineq', 'fun': g1}]
                # Supress runtime warnings
                with warnings.catch_warnings():
                    warnings.filterwarnings("ignore", category=RuntimeWarning)

                    result = minimize(objective, initial_guess, method='SLSQP', bounds=bounds,
                            constraints=cons, options={'ftol': 1e-8, 'maxiter': 1000})
            
                if best_result is None or result.fun <= best_result.fun:               
                    best_result = result

            if best_result is None:
                raise RuntimeError("Optimization failed for all initial guesses.")
            
            print(f'\n\n SLSQP result: {best_result}')
            solution = best_result.x
            print("Constraint value (should be >= 0):", g1(best_result.x))

            # Generate and save
            custom_name = f'{indicators[1, 0]}_perlin.tif'
            filee_name = '_'.join([opt_file_name,custom_name])
            perlin_opt_file_path = os.path.join(optimised_save, filee_name)

            optimised_perlin = imgan.generate_single_perlin_image(image_height,image_width,
                                                    scale=solution[0],
                                                    octaves=int(round(solution[1])),
                                                    persistence=solution[2],
                                                    lacunarity=solution[3], texture_function=indicators[1, 0])
            cv2.imwrite(perlin_opt_file_path, optimised_perlin)
            plt.imshow(optimised_perlin, cmap='gray')
            plt.title(f'Optimised pattern: {filee_name}')
            plt.show()
            plt.get_current_fig_manager().window.raise_()

        #------------------------------------------------------------
        elif gen_value == 5:
            print('\nTuring optimisation...')
            # Coordinate array
            x_1 = p_param[:,0]      # Speckle diameter
            x_2 = p_param[:,1]      # Speckle density
            x_3 = p_param[:,2]      # Position randomness
            x_4 = p_param[:,3]      # Grid-step
            x_5 = p_param[:,4]      # Size randomness
            # For nans
            x_6 = nans[:,0]
            '''
            Dimensional changes: coordinate_arr
                                bounds
                                initial_guess
            '''
            coordinate_arr = np.vstack((x_1,x_3,x_4,x_5)).T

            # RBFInterpolator surrogate model
            poly = PolynomialFeatures(degree=polyy)  
            X_poly = poly.fit_transform(coordinate_arr)
            poly_model = LinearRegression()
            poly_model.fit(X_poly, z)
            
            # Surrage models here
            # results = imgan.surrogate(X,model = 1)
            def interp_value(X):
                '''
                Polynomial interpolation for optimization
                '''
                # For multiple points evaluation
                X_poly = poly.transform(X)
                results = poly_model.predict(X_poly)
                return results  

            # Define constraint inequality
            nan_model = LinearRegression().fit(coordinate_arr, x_6)

            def g1(x):
                nan_pred = nan_model.predict(x.reshape(1, -1))[0]
                nan_pred = np.clip(nan_pred, 0, 100)
                return 5 - nan_pred 
                
            # Bounds
            x1_max = np.max(x_1)
            x1_min = np.min(x_1)
            x2_max = np.max(x_2)
            x2_min = np.min(x_2)
            x3_max = np.max(x_3)
            x3_min = np.min(x_3)
            x4_max = np.max(x_4)
            x4_min = np.min(x_4)
            x5_max = np.max(x_5)
            x5_min = np.min(x_5)
            bounds = [(x1_min, x1_max), (x3_min, x3_max), (x4_min, x4_max), (x5_min, x5_max)]

            def objective(x):
                # convert the 1D array to a 2D array 
                return optimization_direction * interp_value(x.reshape(1, -1))[0]
                    
            best_result = None
            # Number of entries = number of starting points
            for i in range(len(p_param[:,0])):

                # Skipping the middle parameter because it must remain constant
                initial_guess = np.array([p_param[i,0],p_param[i,2],p_param[i,3],p_param[i,4]])
                cons = [{'type': 'ineq', 'fun': g1}]
                # Supress runtime warnings
                with warnings.catch_warnings():
                    warnings.filterwarnings("ignore", category=RuntimeWarning)

                    result = minimize(objective, initial_guess, method='SLSQP', bounds=bounds,
                            constraints=cons, options={'ftol': 1e-8, 'maxiter': 1000})
            
                if best_result is None or result.fun <= best_result.fun:               
                    best_result = result

            if best_result is None:
                raise RuntimeError("Optimization failed for all initial guesses.")
            
            print(f'\n\n SLSQP result: {best_result}')
            solution = best_result.x
            print("Constraint value (should be >= 0):", g1(best_result.x))

            # Make turing
            # Generate and save
            custom_name = 'turing.tif'
            filee_name = '_'.join([opt_file_name,custom_name])
            turing_opt_file_path = os.path.join(optimised_save, filee_name)
            generate_and_save(500, 2000, 25.4 , solution[0],
                            turing_opt_file_path,size_randomness=solution[3],position_randomness=solution[1], speckle_blur=1.5, grid_step=solution[2])
            
            # Generated_speckles = cv2.imread(turing_opt_file_path)
            Generated_speckles_turing = imgan.make_turing_single(turing_opt_file_path,rep=50, radius=1, sharpen_percent=500, size=None)
            Generated_speckles_turing.save(turing_opt_file_path)
            plt.imshow(Generated_speckles_turing, cmap='gray')
            plt.title(f'Optimised pattern: {filee_name}')
            plt.show()
            plt.get_current_fig_manager().window.raise_()

# === Read all errors (mean RMS per iteration) ===

# IEEE formatting
plt.style.use(['science', 'no-latex','ieee', 'grid'])

plt.rcParams.update({
    'font.family': 'Calibri',
    'font.size': 12,
    'axes.labelsize': 11,
    'xtick.labelsize': 10,
    'ytick.labelsize': 10,
    'legend.fontsize': 7,
    'lines.linewidth': 1.25,
    'lines.markersize': 3,

    # Border styling
    'axes.linewidth': 1.0,     
    'xtick.major.width': 1.0,    
    'ytick.major.width': 1.0,    
    'xtick.minor.width': 0.6,    
    'ytick.minor.width': 0.6,

    'xtick.major.size': 5,      
    'ytick.major.size': 5,
    'xtick.minor.size': 3,      
    'ytick.minor.size': 3,
})


value_of_interest = []

error1 = []
error2 = [] 

for read_number in range(iterations_limit):
    the_path = rf"C:\Users\General User\nokop\pattern2\output\excel_docs\iterative_study_mesh_lele\iter_{read_number}"
    # the_path = rf"output\excel_docs\iterative_study\iter_{read_number}"
    _, err, _, _, _ = imgan.read_spec_excel(the_path, doc_num=None)


    error1.append(err[:,6])    
    error2.append(err[:,7])    

    value_of_interest.append(err[:, 0])  


half_iter = iterations_limit // 2
eval_shape_fun = False

if eval_shape_fun:
    n = len(changing_parameter) // 2

    affine_error = error1[:n]
    quadratic_error = error2[n:]
    error_ratio = np.array(error2[n:]) / np.array(error1[:n])

    domain_affine = changing_parameter[:n]
    domain_quadratic = changing_parameter[n:]

    plt.plot(domain_affine[1:], affine_error[1:], 'o-', label='First order shape functions')
    plt.plot(domain_quadratic[1:], quadratic_error[1:], '^-', label='Second order shape functions')
    plt.ylim([0,0.007])
    print(f'---------------\nMean error ratio = {np.mean(error_ratio)}')


else:

    # plt.plot(changing_parameter, np.abs(value_of_interest), '--', color='black', label = 'Difference')

    # Additional (will remove after mesh study to avoid confusion)
    plt.plot(changing_parameter, error1, '-^', label = 'RMSE evaluated on FE grid')
    plt.plot(changing_parameter, error2, '-o', label = 'RMSE evaluated on DIC grid')
    # plt.plot(changing_parameter, np.abs(value_of_interest), '-')
    # plt.ylim(0.00, 0.0055)


tit = 'meshel_size'
xlab = "FE mesh element size [pixel]"
ylab = "RMSE [pixel]"

plt.xlabel(xlab)
plt.ylabel(ylab)
plt.legend()
plt.grid(True)
plt.tight_layout(pad=1.5)
results_save = os.path.join(single_plot_path,f'{tit}.png')
plt.savefig(results_save, dpi = 300, bbox_inches='tight', pad_inches=0.2)
plt.show()



# Process runtime
tic2 = time.time()
now = datetime.now()
current_time = now.strftime("%H:%M:%S")
print("Process completed....")
print("-------------------------------------------------------------------------------------------------")
print(f'Time taken: {tic2-tic1:.3f} seconds')
print("\nCurrent Time =", current_time)


