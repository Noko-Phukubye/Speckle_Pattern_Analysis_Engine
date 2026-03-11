print("Importing libraries...\n---")

# region Imports
from scipy.interpolate import RBFInterpolator, griddata
from sklearn.preprocessing import PolynomialFeatures, StandardScaler
from sklearn.linear_model import LinearRegression
from sklearn.svm import SVR
from sklearn import linear_model
from mealpy import PSO, FloatVar
from pyNastran.bdf.bdf import BDF
from pyNastran.op2.op2 import OP2
from numpy.linalg import norm
from scipy.optimize import minimize
from pyDOE import lhs
import matplotlib.pyplot as plt
import numpy as np
import cv2
import openpyxl
import configparser
import traceback
import warnings
import subprocess
import scipy.stats as stats
import ray
import sys
import os
import time
from datetime import datetime
import image_process_tool_box as imgan
import file_paths as path
from speckle_pattern import generate_and_save, generate_lines, generate_checkerboard
import sundic.sundic as sdic
import sundic.post_process as sdpp
import sundic.settings as sdset
# endregion

'''
Workflow:
1. Create an excel sheet. Each block that writes data will open the existing sheet individually.
2. In a loop of size n, generate n speckle patterns using Latin Hypercube Sampling. 
   Save sample arrays to excel.
3. Copy DIC reference images to the planar images folder with prefixes 0, 2, 4...n
4. Load FEA data.
5. Deform reference images and save with prefixes 1, 3, 5...n
6. Run DIC and save results as .sdic binary files in dedicated folder.
7. Access .sdic files for error analysis by comparing against FEA data. 
   The FEA data is corrected first.
8. Perform global optimisation.
'''

# IEEE-style font settings for all matplotlib figures
plt.style.use(['science', 'no-latex', 'ieee', 'grid'])
plt.rcParams.update({
    'font.family': 'Calibri',
    'font.size': 12,
    'axes.labelsize': 18,
    'xtick.labelsize': 10,
    'ytick.labelsize': 10,
    'legend.fontsize': 7,
    'lines.linewidth': 1.25,
    'lines.markersize': 1.5,
    'axes.linewidth': 1.0,
    'xtick.major.width': 1.0,
    'ytick.major.width': 1.0,
    'xtick.minor.width': 0.6,
    'ytick.minor.width': 0.6,
    'xtick.major.size': 5,
    'ytick.major.size': 5,
    'xtick.minor.size': 3,
    'ytick.minor.size': 3,
})

FLAGS = {
    'Create_new_excel':  True,
    'generate_pattern':  True,
    'Perlin_images':     False,
    'Load FEA':          True,
    'Deform speckles':   True,
    'run_dic':           True,
    'Pattern analysis':  True,
    'run_error':         True,
    'error_dist':        True,
    'Excel plots':       True,
    'Optimisation_poly': False
}

imgan.flag_status(FLAGS, 3.5)

tic1 = time.time()

'=========================== PATHS AND GLOBAL VARIABLES ==========================='
# region

image_save_path         = 2
error_hist_path         = r"output\histograms"
DIC_contour_path        = r"output\DIC_contour"
znssd_figure_path       = r"output\DIC_contour\ZNSSD"
error_heatmap_path      = r"output\Heatmaps"
debugg_folder           = r"output\Debugging"
excel_path              = r"output\excel_docs"
deformed_image_path     = r"data\speckle_pattern_img\deformed_im"
reference_image_path    = r"data\speckle_pattern_img\reference_im"
difference_images       = r"output\plots\Difference_images"
DIC_settings_path       = r"settings.ini"
sundic_save             = r"output\pyth"
Contour_path            = r"output\DIC_contour"
spt_img_path            = r"data\speckle_pattern_img\subpixel_translation"
numpy_files             = r"output\numpy_files"
sundic_binary_folder    = r"output\sundic_bin"
plot_path               = r"output\plots"
slice_path              = r"output\Slices"
image_matrix            = r"output\image_matrix"
autocorrelation_path    = r"output\Autocorrelation"
power_spec              = r"output\spectral_analysis"
optimised_save          = r"data\speckle_pattern_img\Optimised"
scatter_plots           = r"output\plots\Scatter_plots"
planar_images           = r"planar_images"

dirs_to_create = [
    error_hist_path, DIC_contour_path, znssd_figure_path, error_heatmap_path,
    debugg_folder, excel_path, deformed_image_path, reference_image_path,
    difference_images, sundic_save, spt_img_path, sundic_binary_folder,
    plot_path, slice_path, numpy_files, image_matrix, autocorrelation_path,
    power_spec, optimised_save, scatter_plots, planar_images
]

for dir_path in dirs_to_create:
    os.makedirs(dir_path, exist_ok=True)

# endregion


# Simulation parameters
pattern_count   = 10        # Number of speckle patterns to generate and evaluate

# Paths to the Nastran FEA output files
op2_path        = path.flat30_4mm_op2
bdf_path        = path.flat30_4mm_bdf

# Synthetic image dimensions in pixels (aspect ratio should match 2D FE simulation mesh's aspect ratio)
image_width     = 2000
image_height    = 500

# Rigid body translation range used in subpixel translation tests (in pixels)
u_min           = 0.0
u_max           = 1.1
uinterval       = 0.1

image_type      = 'tif'   # File format for all saved images
deform_type     = 0       # Selects the image deformation method (see imgan for options)
study_type      = 0       # Selects the DIC analysis scenario
relative_error  = False   # Sets

'=========================== CREATE EXCEL SHEET ==========================='
if FLAGS['Create_new_excel']:
    print("\n1. Creating excel document...\n")

    os.makedirs(excel_path, exist_ok=True)
    existing_files = [f for f in os.listdir(excel_path)
                      if f.startswith("Pattern_evaluation_") and f.endswith(".xlsx")]
    doc_number = len(existing_files) + 1

    print("Current doc_number:", doc_number)
    exl_file_path = os.path.join(excel_path, f"Pattern_evaluation_{doc_number}.xlsx")

    workbook = openpyxl.Workbook()
    workbook.save(exl_file_path)
    del workbook


'=========================== GENERATE PATTERN IMAGES ==========================='
if FLAGS['generate_pattern']:
    # Generates n speckle (or alternative) reference images using Latin Hypercube Sampling
    # to vary pattern parameters across the design space. Images are saved to reference_image_path.
    # Set pattern_method to select the pattern type and with_turing to apply Turing post-processing.
    print('\n2. Generating speckle pattern images...\n')

    exl_file_path = imgan.excel_doc_path(excel_path)
    pattern_method = 1
    with_turing    = False

    if pattern_method == 1:
        imgan.ladisk_generator(exl_file_path, reference_image_path,
                               pattern_count, image_width=image_width, image_height=image_height)
    elif pattern_method == 2:
        imgan.ladisk_generator_lines(exl_file_path, reference_image_path,
                                     pattern_count, image_width=image_width, image_height=image_height)
    elif pattern_method == 3:
        imgan.ladisk_generator_cb(exl_file_path, reference_image_path,
                                  pattern_count, image_width=image_width, image_height=image_height)
    elif pattern_method == 4:
        # Texture options: none, thresholded, sinusoidal, bimodal, logarithmic,
        #                  cubic, leopard, perlin_blobs, gaussian_spots, worley
        texture_options = {
            1: 'none',        2: 'thresholded', 3: 'sinusoidal',
            4: 'bimodal',     5: 'logarithmic', 6: 'cubic',
            7: 'perlin_blobs'
        }
        apply_texture = texture_options.get(8)
        imgan.single_perlin(image_height, image_width, excel_path=exl_file_path,
                            ref_image_save=reference_image_path,
                            texture_function=apply_texture,
                            number_of_images=pattern_count)

    if with_turing:
        imgan.make_turing(reference_image_path, rep=25, radius=1,
                          sharpen_percent=250, size=None, replace=True)

    plt.close('all')


'=========================== PERLIN NOISE IMAGES ==========================='
if FLAGS['Perlin_images']:
    # Generates a paired set of reference and FEA-deformed Perlin noise images.
    # This deforms the images analytically

    exl_file_path = imgan.excel_doc_path(excel_path)
    nodes_2d, deformed_nodes_2d = imgan.load_fe_nodes(bdf_path, op2_path)

    # Load FE node positions and apply displacement to get deformed position

    # Scale FE data to match image dimensions. 
    max_x_FE = np.max(nodes_2d[:, 0])
    x_scale  = image_width / max_x_FE 
    nodes_2d          = nodes_2d * x_scale
    deformed_nodes_2d = deformed_nodes_2d * x_scale

    # Compute the displacement field over the image domain
    _, _, dx, dy = imgan.smooth_field(
        np.full((image_height, image_width), np.nan),
        nodes_2d,
        deformed_nodes_2d,
        3
    )

    # Generate and save paired Perlin images warped by the FEA displacement field.
    # Texture options: none, thresholded, sinusoidal, bimodal, logarithmic,
    #                  cubic, perlin_blobs
    imgan.generate_perlin_pair(image_height, image_width, dx, dy,
                               exl_file_path, reference_image_path,
                               deformed_image_path, number_of_images=pattern_count,
                               texture_fun='cubic')


'=========================== LOAD FEA DATA ==========================='
if FLAGS['Load FEA']:
    # Reads the Nastran OP2 and BDF files to extract node coordinates and 
    # displacements. FE coordinates are then scaled to match the image pixel 
    # dimensions. There are options to show meshes and smooth deformation fields 
    # but these are not necessary and slow down the process when turned on.
    print('\n4. Loading FEA data...\n')

    FE_flags = {
        'show_meshes':             False,  # Plot original vs deformed mesh side by side
        'show_displacement_field': False   # Plot the interpolated displacement field over the image
    }

    # Parse FEA files
    model = OP2()
    model.read_op2(op2_path)
    bdf = BDF()
    bdf.read_bdf(bdf_path)

    # Extract translational displacements from the first time step of subcase 1 (static)
    itime, isubcase = 0, 1
    disp   = model.displacements[isubcase]
    txyz   = disp.data[itime, :, :3]   # Shape: (nnodes, 3) => Tx, Ty, Tz per node
    nnodes = disp.data.shape[1]

    # Build 2D node arrays
    # Extract all node positions from the BDF and keep only X and Y (ignore Z)
    nodes      = np.array([bdf.nodes[nid].get_position() for nid in bdf.nodes])
    nodes_2d   = nodes[:, :2]
    print(f'\n-------\nNodes 2D shape: {nodes_2d.shape}\n-------')

    # Only X and Y displacements
    displacements_2d  = txyz[:, :2]          
    # Apply displacement to get deformed positions             
    deformed_nodes_2d = nodes_2d + displacements_2d   

    # Visualise mesh nodes (both underformed and deformed)
    if FE_flags['show_meshes']:
        fig, axes = plt.subplots(1, 2, figsize=(12, 6))
        axes[0].scatter(nodes_2d[:, 0], nodes_2d[:, 1], c='b', marker='o')
        axes[0].set(title='Original Mesh', xlabel='X', ylabel='Y')
        axes[1].scatter(deformed_nodes_2d[:, 0], deformed_nodes_2d[:, 1], c='r', marker='o')
        axes[1].set(title='Deformed Mesh', xlabel='X', ylabel='Y')
        plt.show(block=False)
        cfm = plt.get_current_fig_manager()
        cfm.window.activateWindow()
        cfm.window.raise_()
        plt.pause(2.5)
        plt.close()

    # Scale FE coordinates to image pixel space
    # The FE model uses real-world units. A uniform scale factor is derived from
    # the ratio of image width (px) to the maximum FE x-coordinate (mm), so that the
    # full specimen width fills the image. This assumes that the image in question has 
    # the same aspect ratio as the FE model. 
    max_x_FE = np.max(nodes_2d[:, 0])
    x_scale  = image_width / max_x_FE
    print(f'max x = {max_x_FE} | image width = {image_width} | scale = {x_scale}')

    original_points = nodes_2d * x_scale
    new_points      = deformed_nodes_2d * x_scale
    print('\nFEA data loaded successfully.')

    # Interpolate displacement field over full image grid ---
    # Generates a smooth_field through RBF interpolation. 
    # Produces continuous dx, dy displacement
    # maps that can be used to warp reference images in the Deform speckles block.
    dx, dy, FEx_rbf_interp, FEy_rbf_interp = imgan.smooth_field(
        np.zeros((image_height, image_width)),
        original_points,
        new_points,
        3
    )

    # Visualise displacement field
    if FE_flags['show_displacement_field']:
        field = imgan.show_field(
            np.zeros((image_height, image_width)),
            original_points, new_points, debugg_folder
        )
        plt.draw()
        plt.pause(2.5)
        plt.close('all')

'=========================== DEFORM IMAGES ==========================='
if FLAGS['Deform speckles']:
    # Warps each reference speckle image using the FEA displacement field (or a
    # rigid body translation) to produce a synthetic deformed image pair for DIC.
    # The deformed images are saved separately and loaded alongside the reference
    # images by the DIC algorithm in the downstream blocks.
    #
    # Deformation pipeline (FE-based):
    #   1. Read reference image
    #   2. Apply RBF-interpolated displacement field (dx, dy)
    #   3. Save deformed image and diagnostic difference plots

    deform_flags = {
        'gaussian_blur_ref_img': True,   # Apply Gaussian blur to reference images before deforming
        'gaussian_noise_ref_img': False,  # Add Gaussian noise to reference images before deforming
        'show_process': False,            # Display before/after comparison with FE mesh overlay
        'Inverse consistency': False      # Re-deform with -dx/-dy to check inverse consistency
    }
    imgan.flag_status(deform_flags, wait_time=1.5)

    # Optional image pre-filtering
    if deform_flags['gaussian_blur_ref_img']:
        print('\nApplying Gaussian blur to reference images...')
        imgan.gaussian_blur_images(reference_image_path, size=5, sig_y=1.0, par='even')

    if deform_flags['gaussian_noise_ref_img']:
        print('\nAdding Gaussian noise to reference images...')
        imgan.gaussian_noise_images(reference_image_path, par='even')

    print('\n5. Deforming images...')

    # FE-based deformation
    # Uses the dx, dy displacement maps computed from the FEA data in the Load FEA block.
    # Each reference image (even-numbered) produces one deformed image (odd-numbered).
    if deform_type == 0:
        print('\nDeformation type: FE-based\n')
        image_files = imgan.get_image_strings(reference_image_path)

        for k, ref_image_file in enumerate(image_files):

            # Load reference image
            ref_img_path = os.path.join(reference_image_path, ref_image_file)
            reference_image = cv2.imread(ref_img_path)
            if reference_image is None:
                print(f"Warning: Could not read image at {ref_img_path}. Skipping.")
                continue

            # Deformed image is numbered one higher than its reference counterpart
            ref_num = int(ref_image_file.split('_')[0])
            deform_num = ref_num + 1
            deformed_name = f"{deform_num}_Generated_spec_image.tif"

            # Apply deformation
            tic_def = time.time()
            print(f'Applying deformation: Image {k + 1}')

            # This one uses the map_coordinates deformation but simply
            # changing the call to cv2_deform will utilise the openCV version.
            transformed_image, difference_image = imgan.img_deform(
                reference_image,
                dx=dx,
                dy=dy
            )

            # Save deformed image
            deformed_path = os.path.join(deformed_image_path, deformed_name)
            cv2.imwrite(deformed_path, transformed_image)

            # Difference image (grey-level deviation between ref and deformed)
            if difference_image.ndim == 3:
                difference_image = np.mean(difference_image, axis=2)  # Collapse to 2D if colour
            normalised_difference = difference_image / difference_image.max()

            plt.figure(figsize=(6, 5))
            plt.imshow(normalised_difference, cmap='jet')
            plt.colorbar(label="Difference")
            plt.axis('off')
            plt.tight_layout()
            plt.savefig(
                os.path.join(difference_images, f'{ref_num}_difference_image.png'),
                bbox_inches='tight', pad_inches=0, dpi=600
            )
            plt.close()

            # Collapsed difference plot (mean along rows)
            # Averages the difference image along the vertical axis to show
            # how the grey-level error varies across the image width.
            collapsed_difference = np.mean(difference_image, axis=0)
            x_line = np.arange(difference_image.shape[1])

            fig, ax = plt.subplots(figsize=(10, 6))
            ax.plot(x_line[1:], collapsed_difference[1:], color='blue', linewidth=2.5)
            ax.set_xlabel("$x$-position [pixel]", fontsize=30)
            ax.set_ylabel("Grey-level difference [0-255]", fontsize=28)
            ax.tick_params(axis='both', which='major', labelsize=24, direction='in',
                           length=10, width=2, top=True, right=True)
            ax.tick_params(axis='both', which='minor', direction='in',
                           length=5, width=1.5, top=True, right=True)
            ax.minorticks_on()
            ax.grid(True, alpha=0.7)
            fig.tight_layout()
            fig.savefig(
                os.path.join(difference_images, f'{ref_num}_collapsed_difference_axis0.png'),
                bbox_inches='tight', pad_inches=0.4, dpi=600
            )
            plt.close(fig)

            # Horizontal slice plot (single row through image centre)
            row_index = difference_image.shape[0] // 2
            row_difference = difference_image[row_index, :]
            x_line = np.arange(difference_image.shape[1])

            plt.figure(figsize=(10, 6))
            plt.plot(x_line, row_difference, color='blue', linewidth=1.5)
            plt.title(f"Difference Image Slice at Row {row_index}")
            plt.xlabel("Pixel Column Index (X)")
            plt.ylabel("Pixel Difference")
            plt.grid(True)
            plt.savefig(
                os.path.join(difference_images, f'{ref_num}_slice_row_{row_index}.png')
            )
            plt.close()

            # Inverse consistency check
            # Reverses the deformation with negated displacements. The result should
            # closely match the original reference image if the deformation is consistent.
            # I did this in the early stages of the project but I never actually ended 
            # up using it.
            if deform_flags['Inverse consistency']:
                print('Applying reverse deformation...')
                reversed_image, _ = imgan.img_deform(transformed_image, dx=-dx, dy=-dy)
                cv2.imwrite(
                    os.path.join(debugg_folder, f'{ref_num}_reversed_deformation.tif'),
                    reversed_image
                )

            # Side-by-side comparison with FE mesh overlay
            if deform_flags['show_process']:
                fig, axs = plt.subplots(2, 1, figsize=(12, 6))
                axs[0].imshow(reference_image)
                axs[0].set_title("Reference")
                axs[0].scatter(*zip(*original_points), color='red', s=5)
                axs[1].imshow(transformed_image)
                axs[1].set_title("Deformed")
                axs[1].scatter(*zip(*new_points), color='red', s=5)
                plt.figtext(0.5, 0.02,
                            f'Reference: {reference_image.shape} | '
                            f'Deformed: {transformed_image.shape}',
                            ha='center')
                plt.tight_layout(pad=1.5)
                plt.savefig(
                    os.path.join(debugg_folder, f"comparison_{ref_num}_{deform_num}.png"),
                    dpi=300, bbox_inches='tight'
                )
                plt.close(fig)

            toc_def = time.time()
            print(f'Image {k + 1} complete — {toc_def - tic_def:.3f}s\n')

    # Rigid body translation
    # Applies known sub-pixel shifts to the reference images for ground-truth
    # accuracy testing of the DIC engine (no FEA data required).
    elif deform_type == 1:
        print('Deformation type: Rigid-body translation')
        imgan.subpixel_translation(
            reference_image_path=reference_image_path,
            savepath=spt_img_path,
            shift_method='fourier',   # 'fourier' or 'grid'
            imagetype=image_type,
            umin=u_min,
            umax=u_max,
            intervals=uinterval
        )

    else:
        print(f"Invalid deform_type '{deform_type}'. Expected 0 (FE-based) or 1 (rigid body).")

    plt.close('all')


'=========================== DIGITAL IMAGE CORRELATION ==========================='
if FLAGS['run_dic']:
    # Runs SUN-DIC on each reference/deformed image pair and saves the results as
    # binary files. Optionally processes sub-pixel translation results for accuracy
    # benchmarking. The study_type flag controls which image folder is used:
    #   0 -> FE-deformed images (main analysis)
    #   1 -> Rigidly translated images (sub-pixel accuracy benchmark)

    DIC_flags = {
        'Image_blur':            True,   # Apply Gaussian blur to both image sets before DIC
        'sub_pixel_noise':       False,  # Add noise to sub-pixel translated images
        'run_analysis':          True,   # Run the DIC engine
        'process_subpixel_data': True    # Post-process sub-pixel translation results
    }
    imgan.flag_status(DIC_flags, wait_time=1.5)

    # pre-filter images with Gaussian blur
    # Smoothing both image sets consistently reduces noise-induced correlation errors.
    if DIC_flags['Image_blur']:
        print('\nApplying Gaussian blur to image sets...')
        imgan.gaussian_blur_images(reference_image_path, size=5, sig_y=1.0, par='even')
        imgan.gaussian_blur_images(deformed_image_path,  size=5, sig_y=1.0, par='odd')

        # Add noise to sub-pixel images if benchmarking under noisy conditions
        if study_type == 1 and DIC_flags['sub_pixel_noise']:
            imgan.gaussian_noise_images(spt_img_path, par='none')

    print("\n6. Running DIC analysis...\n")

    # --- Load DIC settings ---
    if not os.path.exists(DIC_settings_path):
        print(f"Warning: Settings file not found at {DIC_settings_path}")

    settings = sdset.Settings.fromSettingsFile(DIC_settings_path)

    # ROI is inset by start_x/start_y on each side to avoid edge artefacts.
    # Set roi = 'auto' to derive width/height from image dimensions automatically.
    roi     = 'auto'
    start_x = 30
    start_y = 30

    if roi.lower() == 'auto':
        width  = image_width  - (2 * start_x)
        height = image_height - (2 * start_y)
    else:
        width  = 1990
        height = 490

    settings.ROI                  = [start_x, start_y, width, height]
    settings.SubsetSize           = 33          # Subset size in pixels
    settings.StepSize             = 11          # Step between subset centres in pixels
    settings.GaussianBlurSize     = 5
    settings.GaussianBlurStdDev   = 0.0
    settings.DebugLevel           = 2
    settings.CPUCount             = 8
    settings.OptimizationAlgorithm = "IC-GN"   # Inverse compositional Gauss-Newton
    settings.ShapeFunctions       = "Affine"
    settings.ReferenceStrategy    = "Absolute"

    print(settings.__repr__())
    print("Settings loaded successfully\n")

    # Select image folder based on study type
    if study_type == 0:
        deformed_images_dir    = deformed_image_path
        sundic_binary_file_loc = sundic_binary_folder
    elif study_type == 1:
        deformed_images_dir    = spt_img_path
        sundic_binary_file_loc = os.path.join(sundic_binary_folder, "subpixel")
        os.makedirs(sundic_binary_file_loc, exist_ok=True)

    # Initialise Ray for parallel processing
    # Ray distributes subset correlation across CPU cores. If not already running,
    # a head node is started automatically.
    try:
        print('Checking Ray status...')
        subprocess.run('ray status', shell=True, check=True)
        time.sleep(5)
    except subprocess.CalledProcessError:
        print('Ray not running — starting head node...')
        subprocess.run('ray start --head --num-cpus=4', shell=True)
        time.sleep(5)
        subprocess.run('ray status', shell=True)

    # Run DIC
    if DIC_flags['run_analysis']:
        imgan.run_dic(
            settings,
            reference_image_path,
            deformed_images_dir,
            sundic_binary_file_loc,
            debugg_folder,
            DIC_contour_path,
            znssd_figure_path,
            study_type,
            image_type,
            start_index=0,
            umin=u_min,
            umax=u_max,
            u_interval=uinterval
        )

    # Post-process sub-pixel translation results
    # Compares DIC-measured displacements against known applied translations
    # and saves the error matrix for use in downstream analysis.
    if study_type == 1 and DIC_flags['process_subpixel_data']:
        error_matrix, subpixel_trans_figure = imgan.subpixel_analysis(
            sundic_binary_file_loc,
            plot_path,
            u_min=u_min,
            u_max=u_max,
            u_interval=uinterval
        )
        np.save(os.path.join(numpy_files, 'error_matrix.npy'), error_matrix)

    plt.close('all')


'=========================== ANALYSE PATTERNS ==========================='
if FLAGS['Pattern analysis']:
    # Reads reference speckle images and computes a set of image quality metrics
    # for each pattern. Results are written to the active Excel workbook.
    #
    # Metrics computed (and their column in the sheet):
    #   A - MSF    : Mean Subset Fluctuation (gradient-based, 1st order)
    #   B - MIG    : Mean Intensity Gradient
    #   C - Ef     : E_f metric
    #   D - MIOSD  : Mean Intensity of Second Derivatives
    #   E - Shannon: Shannon entropy (information content)
    #   F - Power  : Integrated 2D power spectrum area (PSA)
    #   G - SSSIG  : Sum of Square Subset Intensity Gradients
    #   H - R_peak : Autocorrelation peak radius

    print('\n3. Analysing speckle patterns...\n')

    try:
        # Load the most recent workbook, or create one if none exists
        exl_file_path = imgan.excel_doc_path(excel_path)
        try:
            workbook = openpyxl.load_workbook(exl_file_path)
        except FileNotFoundError:
            workbook = openpyxl.Workbook()

        sheet = workbook.active

        # Write column headers
        sheet['A1'] = 'MSF'
        sheet['B1'] = 'MIG'
        sheet['C1'] = 'E_f'
        sheet['D1'] = 'MIOSD'
        sheet['E1'] = 'Shannon'
        sheet['F1'] = 'Power spectrum area'
        sheet['G1'] = 'SSSIG globe'
        sheet['H1'] = 'Correlation peak radius'

        # Standardise image filenames before processing
        imgan.rename_img(reference_image_path)
        image_files = imgan.get_image_strings(reference_image_path, imagetype=image_type)
        print(f"\n------\nImage files: {image_files}")

        # Derive the filename suffix from the first file so that filenames for
        # other prefixes can be reconstructed consistently (e.g. '0_speckle.tif' → '_speckle.tif')
        image_file_endswith = '_'.join(image_files[0].split('_')[1:])

        # Map each prefix to its row index in the Excel sheet (data starts at row 2)
        reference_image_prefixes = imgan.expected_prefixes(reference_image_path)
        prefix_index = {prefix: i for i, prefix in enumerate(reference_image_prefixes)}

        for prefix_num in reference_image_prefixes:
            j = prefix_index[prefix_num] + 2   # Excel row for this image
            ref_image_file = f'{prefix_num}_{image_file_endswith}'
            ref_img_path   = os.path.join(reference_image_path, ref_image_file)

            try:
                if not os.path.exists(ref_img_path):
                    raise FileNotFoundError(f'Image not found: {ref_image_file}')

                print(f'Processing: {ref_image_file}')
                image = imgan.readImage(ref_img_path)

                # --- Grey-level histogram ---
                hist_fig, ax = plt.subplots()
                ax.hist(image.ravel(), bins=256, color='black', edgecolor='white')
                ax.set(xlabel='Grey value', ylabel='Pixel count', title='Grey Value Histogram')
                hist_save_path = os.path.join(sundic_save, f'Histogram_{ref_image_file}.png')
                hist_fig.savefig(hist_save_path)
                plt.close(hist_fig)

                # --- Compute metrics ---
                msf    = imgan.MSF(image)
                mig    = imgan.MIG(image)
                ef     = imgan.Ef(image)
                miosd  = imgan.miosd(image)
                shannon = imgan.ShannonEnt(image)
                sss    = imgan.globalSSSIG(image, 33, 5)

                # Power spectrum — integrates the 2D FFT spectrum area
                mag, powr, fx, fy = imgan.compute_fft_and_freq(image)
                power_area = imgan.integrate_power_2d(powr, fx, fy)
                power_fig  = imgan.plot_1d_spectra(mag, powr, fx, fy, show=False)
                power_fig.savefig(os.path.join(power_spec, f'Power_spectrum_{ref_image_file}.png'))
                plt.close(power_fig)

                # Autocorrelation peak radius
                _, figR, R_peak = imgan.autocorr(image, meth=6, cardinality=5, autype='2d')
                figR.savefig(os.path.join(autocorrelation_path, f'Autocorrelation_{ref_image_file}.png'))
                plt.close(figR)

                # --- Write results to Excel ---
                sheet[f'A{j}'] = msf
                sheet[f'B{j}'] = mig
                sheet[f'C{j}'] = ef
                sheet[f'D{j}'] = miosd
                sheet[f'E{j}'] = shannon
                sheet[f'F{j}'] = power_area
                sheet[f'G{j}'] = sss
                sheet[f'H{j}'] = R_peak

            except Exception as e:
                # If any metric fails, write zeros for this row so the sheet remains complete
                for col in 'ABCDEFGH':
                    sheet[f'{col}{j}'] = 0
                print(f'Pattern analysis failed for {ref_image_file}: {e}')
                traceback.print_exc()

        workbook.save(exl_file_path)
        print(f"Workbook saved at {exl_file_path}")

    except FileNotFoundError as e:
        print(f"File or folder not found: {e}")

    except Exception as e:
        tb = traceback.extract_tb(sys.exc_info()[2])
        filename, line_number, _, _ = tb[-1]
        print(f'Pattern analysis error on line {line_number} in {filename}: {e}')

    finally:
        if 'workbook' in locals():
            workbook.close()

    plt.close('all')


'=========================== ERROR ANALYSIS TO EXCEL ==========================='
if FLAGS['run_error']:
    # Compares DIC-measured displacements against FEA ground-truth values and
    # writes error statistics to the active Excel workbook.
    #
    # Two comparisons are performed for each image pair:
    #   d2f (DIC-to-FEM) : DIC results interpolated onto the FEA grid, then differenced
    #   f2d (FEM-to-DIC) : FEA results interpolated onto the DIC grid, then differenced
    #
    # Error metrics written to Excel (columns I–P, Z):
    #   I - RMSE_d2f   J - IQR_d2f   K - RMSE_f2d   L - IQR_f2d
    #   M - MBE_d2f    N - MBE_f2d   O - SDE_d2f    P - SDE_f2d
    #   Z - NaN ratio (%)
    print('\n7. Running error analysis...')

    # --- Load or create the workbook ---
    os.makedirs(excel_path, exist_ok=True)
    existing_files = [f for f in os.listdir(excel_path)
                      if f.startswith("Pattern_evaluation_") and f.endswith(".xlsx")]
    doc_number    = len(existing_files)
    exl_file_path = os.path.join(excel_path, f"Pattern_evaluation_{doc_number}.xlsx")
    print(f"Current doc_number (Error analysis): {doc_number}")

    try:
        workbook = openpyxl.load_workbook(exl_file_path)
    except FileNotFoundError:
        workbook = openpyxl.Workbook()

    sheet = workbook.active

    # Write column headers
    sheet['I1'] = 'RMSE_d2f';  sheet['J1'] = 'IQR_d2f'
    sheet['K1'] = 'RMSE_f2d';  sheet['L1'] = 'IQR_f2d'
    sheet['M1'] = 'MBE_d2f';   sheet['N1'] = 'MBE_f2d'
    sheet['O1'] = 'SDE_d2f';   sheet['P1'] = 'SDE_f2d'

    #Prepare FEA reference data
    # Scale FEA coordinates and displacements from model units to image pixel space
    fem_xcoord = nodes_2d[:, 0] * x_scale
    fem_ycoord = nodes_2d[:, 1] * x_scale
    fem_x_disp = displacements_2d[:, 0] * x_scale
    fem_y_disp = displacements_2d[:, 1] * x_scale
    fem_mag    = np.sqrt(fem_x_disp**2 + fem_y_disp**2)

    # Apply Lagrangian correction: shift FEA coordinates back by the interpolated
    # displacement so they align with the reference (undeformed) image frame
    lag_dx = FEx_rbf_interp(np.column_stack([fem_xcoord, fem_ycoord]))
    lag_dy = FEy_rbf_interp(np.column_stack([fem_xcoord, fem_ycoord]))
    fem_points = np.column_stack((fem_xcoord - lag_dx, fem_ycoord - lag_dy))

    # Displacement component to use for error analysis
    # 0 = x-displacement, 1 = y-displacement, 2 = displacement magnitude
    value = 0

    # Iterate over all expected DIC result files
    all_expected_prefixes = imgan.expected_prefixes(sundic_binary_folder)
    prefix_positions      = {prefix: i for i, prefix in enumerate(all_expected_prefixes)}

    sundic_binary_files = sorted(
        [f for f in os.listdir(sundic_binary_folder) if f.endswith('results.sdic')],
        key=lambda x: int(x.split('_')[0])
    )
    print(f'\nSUN-DIC binary files found: {sundic_binary_files}')

    for prefix_num in all_expected_prefixes:
        m = prefix_positions[prefix_num] + 2   # Specific excel row for this result
        sunfile = f'{prefix_num}_results.sdic'
        sundic_data_path = os.path.join(sundic_binary_folder, sunfile)
        print(f'\nPrefix: {prefix_num} | Excel row: {m} | File: {sunfile}')

        try:
            if not os.path.exists(sundic_data_path):
                raise FileNotFoundError(f"DIC result file not found: {sunfile}")

            # Load DIC results
            sundic_data, nRows, nCols = sdpp.getDisplacements(
                sundic_data_path, -1, smoothWindow=11
            )

            # Reject files with too many NaN points (threshold is 15%)
            nan_count      = np.isnan(sundic_data).any(axis=1).sum()
            nan_percentage = (nan_count / len(sundic_data)) * 100
            print(f"NaN ratio: {nan_percentage:.2f}%")

            if nan_percentage >= 15:
                raise ValueError(f"NaN percentage too high: {nan_percentage:.2f}%")

            x_coord = sundic_data[:, 0]
            y_coord = sundic_data[:, 1]
            X_disp  = sundic_data[:, 3]
            Y_disp  = sundic_data[:, 4]

            # Filter out unrealistically large displacements
            threshold        = 15000
            valid            = np.sqrt(X_disp**2 + Y_disp**2) < threshold
            x_coord, y_coord = x_coord[valid], y_coord[valid]
            X_disp,  Y_disp  = X_disp[valid],  Y_disp[valid]
            sundic_points    = np.column_stack((x_coord, y_coord))

            # Select which displacement component to compare
            if value == 0:
                fem_value = fem_x_disp
                dic_value = X_disp
            elif value == 1:
                fem_value = fem_y_disp
                dic_value = Y_disp
            elif value == 2:
                fem_value = fem_mag
                dic_value = np.sqrt(X_disp**2 + Y_disp**2)
            else:
                raise ValueError(f"Invalid value selector: {value}. Expected 0, 1, or 2.")

            # f2d: FEA interpolated onto DIC grid
            # Error is computed at DIC measurement points
            interp_fem = griddata(fem_points, fem_value, sundic_points, method='cubic')
            mask_f2d   = ~np.isnan(interp_fem) & ~np.isnan(dic_value)
            err_f2d    = dic_value[mask_f2d] - interp_fem[mask_f2d]

            MBE_f2d  = np.mean(err_f2d)
            SDE_f2d  = np.sqrt(np.sum((err_f2d - MBE_f2d)**2) / (len(err_f2d) - 1))
            RMSE_f2d = np.sqrt(MBE_f2d**2 + SDE_f2d**2)
            IQR_f2d  = stats.iqr(err_f2d)

            # d2f: DIC interpolated onto FEA grid
            # Error is computed at FEA node positions
            interp_dic = griddata(sundic_points, dic_value, fem_points, method='cubic')
            mask_d2f   = ~np.isnan(interp_dic) & ~np.isnan(fem_value)
            err_d2f    = interp_dic[mask_d2f] - fem_value[mask_d2f]

            MBE_d2f  = np.mean(err_d2f)
            SDE_d2f  = np.sqrt(np.sum((err_d2f - MBE_d2f)**2) / (len(err_d2f) - 1))
            RMSE_d2f = np.sqrt(MBE_d2f**2 + SDE_d2f**2)
            IQR_d2f  = stats.iqr(err_d2f)

            # Write results to Excel
            sheet[f'I{m}'] = RMSE_d2f;  sheet[f'J{m}'] = IQR_d2f
            sheet[f'K{m}'] = RMSE_f2d;  sheet[f'L{m}'] = IQR_f2d
            sheet[f'M{m}'] = MBE_d2f;   sheet[f'N{m}'] = MBE_f2d
            sheet[f'O{m}'] = SDE_d2f;   sheet[f'P{m}'] = SDE_f2d
            sheet[f'Z{m}'] = nan_percentage

        except Exception as e:
            # Write zeros for this row so the sheet remains complete (empty cells cause issues downstream
            # zeros can be easily filteredx out)  

            for col in 'IJKLMNOP':
                sheet[f'{col}{m}'] = 0
            sheet[f'Z{m}'] = 0
            print(f'Error processing {sunfile} (row {m}): {e}')

    workbook.save(exl_file_path)
    print(f"Workbook saved at {exl_file_path}")


'=========================== ERROR DISTRIBUTION ANALYSIS ==========================='
if FLAGS['error_dist']:
    # Loads each DIC result file, computes the pointwise error against FEA ground truth,
    # and saves three outputs per pattern: an error heatmap, a histogram, and a
    # column-averaged error slice. Results are also saved as .npy files for later use.
    #
    # The FEA coordinate system is Eulerian; DIC reports in Lagrangian coordinates.
    # The RBF interpolators from the Load FEA block are used to convert FEA node
    # positions into the Lagrangian frame before interpolation.
    #
    # Displacement component used for error analysis:
    #   0 = x-displacement, 1 = y-displacement, 2 = magnitude
    print('\n8. Error distribution analysis...\n')

    # Binary save folder for column-averaged slice data
    slice_bin_path = r"output\Slices\slice_binaries\correct2"
    os.makedirs(slice_bin_path, exist_ok=True)

    # Collect and sort DIC result files by their numeric prefix
    sundic_binary_files = sorted(
        [f for f in os.listdir(sundic_binary_folder) if f.endswith('results.sdic')],
        key=lambda x: int(x.split('_')[0])
    )
    print(f'SUN-DIC binary files found: {sundic_binary_files}')

    # Expected prefixes are even numbers corresponding to reference image indices
    all_expected_prefixes = imgan.expected_prefixes(sundic_binary_folder, odd=False, skip=True)
    prefix_positions      = {prefix: i for i, prefix in enumerate(all_expected_prefixes)}

    # Scale FEA coordinates and displacements to image pixel space
    fem_xcoord = nodes_2d[:, 0] * x_scale
    fem_ycoord = nodes_2d[:, 1] * x_scale
    fem_x_disp = displacements_2d[:, 0] * x_scale
    fem_y_disp = displacements_2d[:, 1] * x_scale
    fem_mag    = np.sqrt(fem_x_disp**2 + fem_y_disp**2)

    # Convert FEA node positions from Eulerian to Lagrangian frame
    lag_dx     = FEx_rbf_interp(np.column_stack([fem_xcoord, fem_ycoord]))
    lag_dy     = FEy_rbf_interp(np.column_stack([fem_xcoord, fem_ycoord]))

    fem_points = np.column_stack((fem_xcoord - lag_dx, 
                                  fem_ycoord - lag_dy))
    
    print(f'FEM points shape: {fem_points.shape}')

    # Array to store mean error per file for summary reporting
    meanmean = np.zeros(len(sundic_binary_files))

    for i, prefix_num in enumerate(all_expected_prefixes):
        try:
            print(f'\nProcessing prefix: {prefix_num}')
            sunfile          = f'{prefix_num}_results.sdic'
            sundic_data_path = os.path.join(sundic_binary_folder, sunfile)

            if not os.path.exists(sundic_data_path):
                print(f'File not found: {sundic_data_path}. Skipping.')
                continue

            # Load DIC result file
            sundic_data, nRows, nCols = sdpp.getDisplacements(
                sundic_data_path, -1, smoothWindow=11
            )

            x_coord = sundic_data[:, 0]
            y_coord = sundic_data[:, 1]
            X_disp  = sundic_data[:, 3]
            Y_disp  = sundic_data[:, 4]
            dic_mag = np.sqrt(X_disp**2 + Y_disp**2)

            sundic_points = np.column_stack((x_coord, y_coord))
            print(f'DIC points shape: {sundic_points.shape}')

            # Select displacement component for comparison
            value = 0
            if value == 0:
                fem_value = fem_x_disp;  
                dic_value = X_disp;  
                string = 'X_disp'
            elif value == 1:
                fem_value = fem_y_disp;  
                dic_value = Y_disp;  
                string = 'Y_disp'
            elif value == 2:
                fem_value = fem_mag;     
                dic_value = dic_mag;  
                string = 'Magnitude'
            else:
                raise ValueError(f"Invalid value selector: {value}. Expected 0, 1, or 2.")

            # Interpolate FEA displacements onto the DIC measurement grid
            interp_fem = griddata(fem_points, fem_value, sundic_points, method='cubic')

            # Reshape to 2D grid (nCols x nRows) for spatial plotting
            dic_value  = dic_value.reshape(nCols, nRows)
            interp_fem = interp_fem.reshape(nCols, nRows)

            # Compute pointwise error (relative or absolute)
            if relative_error:
                errors2 = (dic_value - interp_fem) / interp_fem
            else:
                errors2 = dic_value - interp_fem

            # Replace inf and NaN with zero to avoid issues in KDE and plotting
            errors2 = np.nan_to_num(errors2, nan=0.0, posinf=0.0, neginf=0.0)
            print(f'Error grid shape: {errors2.shape}')

            # Save error grid as numpy file for offline analysis
            np.save(os.path.join(numpy_files, f'{prefix_num}_errors.npy'), errors2)

            mean_err       = np.nanmean(errors2)
            meanmean[i]    = mean_err
            print(f'Mean error: {mean_err:.4f}')

            # Collapse FEA interpolated values along each axis for tick labelling
            FE_collapse_x = np.nanmean(interp_fem, axis=1)
            FE_collapse_y = np.nanmean(interp_fem, axis=0)

            # Construct heat map with specific tick mark 
            n_ticks   = 9
            x_indices = np.linspace(0, len(FE_collapse_x) - 1, n_ticks, dtype=int)

            plt.figure(figsize=(10, 6), dpi=300)
            plt.imshow(errors2.T, cmap='jet', interpolation='none',
                       extent=(0, len(FE_collapse_x), 0, len(FE_collapse_y)),
                       origin='lower',
                       vmin=np.nanmin(errors2), vmax=np.nanmax(errors2))
            
            cbar = plt.colorbar(label='$e_{ij}$', shrink=0.6, format='%.4f')
            cbar.set_label('$e_{ij}$ [pixel]', fontsize=20)
            cbar.ax.tick_params(labelsize=16)
            plt.xticks(x_indices, [f'{j * 0.5:.1f}' for j in range(n_ticks)], fontsize=14)
            plt.yticks([])
            plt.xlabel('$u_{0}$ [pixel]', fontsize=18)
            plt.gca().invert_yaxis()
            plt.tight_layout()
            plt.savefig(os.path.join(error_heatmap_path, f'{prefix_num}_heatmap_{string}.png'))
            plt.close()

            # Histogram — frequency distribution of error values
            plt.figure(figsize=(12, 10))
            plt.hist(errors2[~np.isnan(errors2)].flatten(), bins=250, color='blue', density=True)
            plt.title(f'Error Distribution: Pattern {prefix_num}', fontsize=22)
            plt.xlabel(f'Error {string}', fontsize=24)
            plt.ylabel('Frequency', fontsize=24)
            plt.tick_params(axis='both', which='major', labelsize=20)
            plt.xlim(-0.005, 0.005)
            plt.ylim(0, 1000)
            plt.savefig(os.path.join(error_hist_path, f'{prefix_num}_histogram_{string}.png'))
            plt.close()

            # Column-averaged error slice — mean error as a function of x-position
            # Useful for identifying systematic trends across the field of view
            x_grid            = x_coord.reshape(nCols, nRows)
            x_line            = np.mean(x_grid, axis=1)
            collapsed_errors  = np.nanmean(errors2, axis=1)
            FE_collapse       = np.nanmean(interp_fem, axis=1)
            DIC_collapse      = np.nanmean(dic_value, axis=1)
            x_mask            = x_line >= 0

            fig, ax = plt.subplots(figsize=(6, 4), dpi=300)
            ax.plot(DIC_collapse[x_mask], collapsed_errors[x_mask], color='black', linewidth=1.5)
            ax.set_xlabel("$u_{0}$ [pixel]", fontsize=18)
            ax.set_ylabel("$e_{sys,i}$ [pixel]", fontsize=18)
            ax.set_ylim(-0.005, 0.005)
            ax.grid(True)
            ax.tick_params(axis='both', which='major', direction='in', length=6,
                           width=2, labelsize=14, color='black', top=True, right=True)
            ax.tick_params(axis='both', which='minor', direction='in', length=3,
                           width=1.5, color='black', top=True, right=True)
            ax.minorticks_on()
            for spine in ax.spines.values():
                spine.set_edgecolor('black')
                spine.set_linewidth(2)
            plt.tight_layout()
            plt.savefig(os.path.join(slice_path, f'{prefix_num}_slice_{string}.png'))
            plt.close()

            # Save column-averaged data as numpy array for downstream processing.
            # Columns: y-averaged error, x-position, imposed displacement, measured displacement
            slice_numpy = np.column_stack((
                collapsed_errors[x_mask],
                x_line[x_mask],
                FE_collapse[x_mask],
                DIC_collapse[x_mask]
            ))
            print(f'Slice data shape: {slice_numpy.shape}')
            np.save(os.path.join(slice_bin_path, f'{prefix_num}_slice_{string}.npy'), slice_numpy)

        except Exception as e:
            tb = traceback.extract_tb(sys.exc_info()[2])
            _, line_number, _, _ = tb[-1]
            print(f'Error processing {sundic_data_path}: {e} (line {line_number})')


'=========================== READ EXCEL AND PLOT ==========================='
if FLAGS['Excel plots']:
    # Reads pattern metrics and DIC error statistics from the Excel workbook and
    # produces scatter plots of each image quality metric against RMSE. Optionally
    # generates a 3D scatter plot and prints the patterns with extreme metric values.
    print('\n8. Reading excel document for scatter plots...')

    p_metrics, meas_error, p_param, nans, indicators = imgan.read_spec_excel(excel_path, doc_num=None)
    print(f'\nIndicators: {indicators}')

    # Filter 1: remove patterns where d2f and f2d errors diverge beyond threshold
    threshold    = 0.001
    outlier_idx  = imgan.array_difference_outlier(meas_error[:, 0], meas_error[:, 2], threshold)

    # Filter 2: remove patterns with unrealistically large error values
    error_threshold = 10
    outlier_idx_y = [i for i in range(meas_error.shape[0]) if abs(meas_error[i, 0]) > error_threshold]

    # Last filter: removes values above error standard deviation
    standard_deviation_thresh = 3
    combined_outliers = outlier_idx + outlier_idx_y
    mask = (
        np.array([i not in combined_outliers for i in range(meas_error.shape[0])]) &
        (meas_error[:, 0] != 0) &
        (np.abs(meas_error[:, 0] - np.nanmean(meas_error[:, 0])) <= standard_deviation_thresh * np.nanstd(meas_error[:, 0]))
    )

    figflag = {
        '2d scatter':    True,
        '3d scatter':    False,
        'xtreme_index':  False
    }

    # 2D scatter plots — one plot per metric against RMSE
    if figflag['2d scatter']:

        metric_strings = ("MIG", "MSF", "$E_f$", "MIOSD",
                          "Shannon entropy", "PSA", "SSSIG", "$R_{peak}$")
        err_col  = 1
        save_fig = True

        for number, metric in enumerate(metric_strings):
            x_value = p_metrics[mask, number]
            y_value = np.abs(meas_error[mask, err_col])

            fig, ax = plt.subplots(figsize=(4, 3))
            ax.scatter(x_value, y_value, color='black', s=4)
            ax.set_xlabel(metric, fontsize=11)
            ax.set_ylabel('RMSE [pixel]', fontsize=11)
            ax.minorticks_on()
            ax.grid(True, which='major', alpha=0.6, linewidth=0.8)
            ax.grid(True, which='minor', alpha=0.3, linestyle=':', linewidth=0.5)
            ax.tick_params(axis='both', which='major', width=1.0, length=5,
                           labelsize=10, direction='in')
            ax.tick_params(axis='both', which='minor', width=0.6, length=3, direction='in')
            for spine in ax.spines.values():
                spine.set_linewidth(1.0)
                spine.set_edgecolor('black')
            plt.tight_layout(pad=1.5)

            if save_fig:
                plot_title = f"Valid patterns = {len(meas_error[mask, err_col])}"
                plt.savefig(
                    os.path.join(scatter_plots, f'{plot_title}_{metric}_vs_RMSE.png'),
                    dpi=300, bbox_inches='tight', pad_inches=0.2
                )
            plt.close()

    # 3D scatter plot — RMSE as a function of speckle size and density
    if figflag['3d scatter']:
        fig = plt.figure()
        ax  = fig.add_subplot(projection='3d')
        sc  = ax.scatter(p_metrics[mask, 1], p_metrics[mask, 4], meas_error[mask, 0],
                         c=meas_error[mask, 0], cmap='jet', marker='o')
        plt.colorbar(sc, ax=ax, label='RMSE (DIC)')
        ax.set_xlabel('Speckle size')
        ax.set_ylabel('Speckle density')
        ax.set_zlabel('RMSE')
        ax.set_title('RMSE vs speckle size vs speckle density')
        ax.set_zlim(bottom=0.0, top=0.005)
        plt.show()

    # Print the pattern indices that produced the highest and lowest value for each metric
    if figflag['xtreme_index']:

        metric_labels = ['MSF', 'MIG', 'Ef', 'MIOSD', 'Shannon', 'Power area', 'SSSIG', 'Peak radius']

        print('\nMaximum metric values:')
        for col, label in enumerate(metric_labels):
            max_val = np.max(p_metrics[:, col])
            max_idx = np.where(p_metrics[:, col] == max_val)[0][0]
            print(f'  {label}: Pattern {max_idx * 2}  ({max_val:.3f})')

        print('\nMinimum metric values:')
        for col, label in enumerate(metric_labels):
            min_val = np.min(p_metrics[:, col])
            min_idx = np.where(p_metrics[:, col] == min_val)[0][0]
            print(f'  {label}: Pattern {min_idx * 2}  ({min_val:.3f})')

        # Patterns with the best and worst DIC error statistics
        filtered_error  = meas_error[mask]
        valid_mask      = np.isfinite(filtered_error[:, 0])
        ffiltered_error = filtered_error[valid_mask]

        # RMSE and standard deviation error
        minrmse = np.min(ffiltered_error[:, 0])
        maxrmse = np.max(ffiltered_error[:, 0])
        minsde  = np.min(ffiltered_error[:, 6])
        maxsde  = np.max(ffiltered_error[:, 6])

        # Associated indeces/locations
        minrmse_idx = np.where(meas_error[:, 0] == minrmse)[0][0]
        maxrmse_idx = np.where(meas_error[:, 0] == maxrmse)[0][0]
        minsde_idx  = np.where(meas_error[:, 6] == minsde)[0][0]
        maxsde_idx  = np.where(meas_error[:, 6] == maxsde)[0][0]

        print('\nExtreme error values:')
        print(f'  Min RMSE: Pattern {minrmse_idx * 2}  ({minrmse:.6f})')
        print(f'  Max RMSE: Pattern {maxrmse_idx * 2}  ({maxrmse:.6f})')
        print(f'  Min SDE:  Pattern {minsde_idx * 2}   ({minsde:.6f})')
        print(f'  Max SDE:  Pattern {maxsde_idx * 2}   ({maxsde:.6f})')



'=========================== SURROGATE-BASED PATTERN OPTIMISATION ==========================='
if FLAGS['Optimisation_poly']:
    # Fits an SVR surrogate model to the pattern metrics and error data from the Excel
    # workbook, then uses constrained SLSQP optimisation with multiple starting points
    # to find the pattern parameters that maximise (or minimise) the chosen metric.
    # A second SVR model is fitted to the NaN ratio and used as an inequality constraint
    # to reject solutions that would produce too many failed DIC subsets.
    #
    # Pattern type is read from the Excel indicators column and determines which
    # design parameters and generator function are used.

    # NB: Please note that the batch analysis script was the one used for 
    # optimisation in the project so it has the most up to date optimisation 
    # block. I kept this one here because it forms part of the framework 
    # from pattern generation all the way up to optimisation.
    
    print('\n9. Optimising speckle patterns...')

    if not FLAGS['Excel plots']:
        print('Error: Excel data not loaded. Set FLAGS["Excel plots"] = True before running optimisation.')
        sys.exit()

    # Optimisation target (iether a specific metric column or error column)
    z = p_metrics[:, 1]
    optimization_direction = -1     # -1 for maximisation, +1 for minimisation

    direction_label = 'maximisation' if optimization_direction == -1 else 'minimisation'
    print(f'Running {direction_label}')

    # Output file label
    numb          = 0
    polyy         = 3
    opt_file_name = f'{numb}_max_MIG_POLY_{polyy}'

    # Pattern generation parameters from Excel (columns of p_param)
    x_1 = p_param[:, 0]
    x_2 = p_param[:, 1]
    x_3 = p_param[:, 2]
    x_4 = p_param[:, 3]
    x_5 = p_param[:, 4]
    x_6 = nans[:, 0]       # NaN ratio used as inequality constraint

    # Excel provides indicators that tell the block what type of pattern to 
    # optimise for. Optimisation override is selected then block optimises
    # for traditional sepckle pattern.

    # The indicators are:
    # speckle
    # lines
    # checkerboard
    # perlin

    override_optimisation = False
    gen_value = 'speckle' if override_optimisation else indicators[0, 0]
    print(f'Pattern type: {gen_value}')

    # Shared SLSQP settings (if it is selected as the optimiser)
    slsqp_options = {'ftol': 1e-8, 'maxiter': 1000}

    # The following blocks are the class-specific optimisation
    # routines. The class type is determined by the 'indicators'
    # that are obtained from the excel sheet.
    
    # Speckle pattern optimisation
    if gen_value.lower() == 'speckle':
        coordinate_arr = np.vstack((x_1, x_3, x_4, x_5)).T

        # Fit SVR surrogate to the target metric
        scaler_X = StandardScaler()
        scaler_z = StandardScaler()
        X_scaled = scaler_X.fit_transform(coordinate_arr)
        z_scaled = scaler_z.fit_transform(z.reshape(-1, 1)).ravel()
        svr_model = SVR(kernel='rbf')
        svr_model.fit(X_scaled, z_scaled)

        # Fit a second SVR to predict the NaN ratio (used as constraint)
        scaler_nan = StandardScaler()
        x6_scaled  = scaler_nan.fit_transform(x_6.reshape(-1, 1)).ravel()
        svr_nan    = SVR(kernel='rbf')
        svr_nan.fit(X_scaled, x6_scaled)

        def interp_value(X):
            return scaler_z.inverse_transform(
                svr_model.predict(scaler_X.transform(X)).reshape(-1, 1)
            ).ravel()

        def g1(x):
            # Constraint: predicted NaN ratio must stay below chosen threshold (currently 15%)
            nan_pred = scaler_nan.inverse_transform(
                svr_nan.predict(scaler_X.transform(x.reshape(1, -1))).reshape(-1, 1)
            ).ravel()[0]
            return 15 - np.clip(nan_pred, 0, 100)

        # Search parameters
        bounds    = [(np.min(x_1), np.max(x_1)), (np.min(x_3), np.max(x_3)),
                     (np.min(x_4), np.max(x_4)), (np.min(x_5), np.max(x_5))]
        objective = lambda x: optimization_direction * interp_value(x.reshape(1, -1))[0]
        cons      = [{'type': 'ineq', 'fun': g1}]

        best_result = None
        for i in range(len(p_param[:, 0])):

            initial_guess = np.array([p_param[i, 0], 
                                      p_param[i, 2], 
                                      p_param[i, 3], 
                                      p_param[i, 4]])
            
            with warnings.catch_warnings():
                warnings.filterwarnings('ignore', category=RuntimeWarning)
                result = minimize(objective, initial_guess, method='SLSQP',
                                  bounds=bounds, constraints=cons, options=slsqp_options)
            if best_result is None or result.fun <= best_result.fun:
                best_result = result

        if best_result is None:
            raise RuntimeError('Optimisation failed for all initial guesses.')

        solution = best_result.x
        print(f'SLSQP result: {best_result}')
        print(f'Constraint value (should be >= 0): {g1(solution)}')

        file_name          = '_'.join([opt_file_name, 'speckles.tif'])
        spec_opt_file_path = os.path.join(optimised_save, file_name)
        generated_speckles = generate_and_save(
            image_height, image_width, 25.4, solution[0], spec_opt_file_path,
            size_randomness=solution[3], position_randomness=solution[1],
            speckle_blur=1.5, grid_step=solution[2]
        )
        plt.imshow(generated_speckles, cmap='gray')
        plt.title(f'Optimised pattern: {file_name}')
        plt.show(block=False)
        plt.get_current_fig_manager().window.raise_()
        plt.pause(2)
        plt.close()

    # Line pattern optimisation
    elif gen_value.lower() == 'lines':
        coordinate_arr = np.vstack((x_2, x_3)).T

        scaler_X = StandardScaler()
        scaler_z = StandardScaler()
        X_scaled = scaler_X.fit_transform(coordinate_arr)
        z_scaled = scaler_z.fit_transform(z.reshape(-1, 1)).ravel()
        svr_model = SVR(kernel='rbf')
        svr_model.fit(X_scaled, z_scaled)

        scaler_nan = StandardScaler()
        x6_scaled  = scaler_nan.fit_transform(x_6.reshape(-1, 1)).ravel()
        svr_nan    = SVR(kernel='rbf')
        svr_nan.fit(X_scaled, x6_scaled)

        def interp_value(X):
            return scaler_z.inverse_transform(
                svr_model.predict(scaler_X.transform(X)).reshape(-1, 1)
            ).ravel()

        def g1(x):
            # Constraint: predicted NaN ratio must stay below 5%
            nan_pred = scaler_nan.inverse_transform(
                svr_nan.predict(scaler_X.transform(x.reshape(1, -1))).reshape(-1, 1)
            ).ravel()[0]
            return 5 - np.clip(nan_pred, 0, 100)

        bounds    = [(np.min(x_2), np.max(x_2)), (np.min(x_3), np.max(x_3))]
        objective = lambda x: optimization_direction * interp_value(x.reshape(1, -1))[0]
        cons      = [{'type': 'ineq', 'fun': g1}]

        best_result = None
        for i in range(len(p_param[:, 0])):
            initial_guess = np.array([p_param[i, 1], p_param[i, 2]])
            with warnings.catch_warnings():
                warnings.filterwarnings('ignore', category=RuntimeWarning)
                result = minimize(objective, initial_guess, method='SLSQP',
                                  bounds=bounds, constraints=cons, options=slsqp_options)
            if best_result is None or result.fun <= best_result.fun:
                best_result = result

        if best_result is None:
            raise RuntimeError('Optimisation failed for all initial guesses.')

        solution = best_result.x
        print(f'SLSQP result: {best_result}')
        print(f'Constraint value (should be >= 0): {g1(solution)}')

        file_name           = '_'.join([opt_file_name, 'lines.tif'])
        lines_opt_file_path = os.path.join(optimised_save, file_name)
        generate_lines(image_height, image_width, 25.4, solution[1],
                       lines_opt_file_path, orientation='vertical', N_lines=solution[2])
        plt.imshow(cv2.imread(lines_opt_file_path), cmap='gray')
        plt.title(f'Optimised pattern: {file_name}')
        plt.show(block=False)
        plt.get_current_fig_manager().window.raise_()
        plt.pause(2)
        plt.close()

    # Checkerboard pattern optimisation
    elif gen_value.lower() == 'checkerboard':
        coordinate_arr = np.vstack((x_1, x_2)).T

        scaler_X = StandardScaler()
        scaler_z = StandardScaler()
        X_scaled = scaler_X.fit_transform(coordinate_arr)
        z_scaled = scaler_z.fit_transform(z.reshape(-1, 1)).ravel()
        svr_model = SVR(kernel='rbf')
        svr_model.fit(X_scaled, z_scaled)

        scaler_nan = StandardScaler()
        x6_scaled  = scaler_nan.fit_transform(x_6.reshape(-1, 1)).ravel()
        svr_nan    = SVR(kernel='rbf')
        svr_nan.fit(X_scaled, x6_scaled)

        def interp_value(X):
            return scaler_z.inverse_transform(
                svr_model.predict(scaler_X.transform(X)).reshape(-1, 1)
            ).ravel()

        def g1(x):
            nan_pred = scaler_nan.inverse_transform(
                svr_nan.predict(scaler_X.transform(x.reshape(1, -1))).reshape(-1, 1)
            ).ravel()[0]
            return 5 - np.clip(nan_pred, 0, 100)

        bounds    = [(np.min(x_1), np.max(x_1)), (np.min(x_2), np.max(x_2))]
        objective = lambda x: optimization_direction * interp_value(x.reshape(1, -1))[0]
        cons      = [{'type': 'ineq', 'fun': g1}]

        best_result = None
        for i in range(len(p_param[:, 0])):
            initial_guess = np.array([p_param[i, 0], p_param[i, 1]])
            with warnings.catch_warnings():
                warnings.filterwarnings('ignore', category=RuntimeWarning)
                result = minimize(objective, initial_guess, method='SLSQP',
                                  bounds=bounds, constraints=cons, options=slsqp_options)
            if best_result is None or result.fun <= best_result.fun:
                best_result = result

        if best_result is None:
            raise RuntimeError('Optimisation failed for all initial guesses.')

        solution = best_result.x
        print(f'SLSQP result: {best_result}')
        print(f'Constraint value (should be >= 0): {g1(solution)}')

        file_name        = '_'.join([opt_file_name, 'checkb.tif'])
        cb_opt_file_path = os.path.join(optimised_save, file_name)
        generate_checkerboard(image_height, image_width, dpi=25.4, path=cb_opt_file_path,
                              line_width=solution[0], N_rows=solution[1])
        plt.imshow(cv2.imread(cb_opt_file_path), cmap='gray')
        plt.title(f'Optimised pattern: {file_name}')
        plt.show(block=False)
        plt.get_current_fig_manager().window.raise_()
        plt.pause(2)
        plt.close()

    # Perlin noise pattern optimisation
    elif gen_value.lower() == 'perlin':
        coordinate_arr = np.vstack((x_1, x_2, x_3, x_4)).T

        scaler_X = StandardScaler()
        scaler_z = StandardScaler()
        X_scaled = scaler_X.fit_transform(coordinate_arr)
        z_scaled = scaler_z.fit_transform(z.reshape(-1, 1)).ravel()
        svr_model = SVR(kernel='rbf')
        svr_model.fit(X_scaled, z_scaled)

        scaler_nan = StandardScaler()
        x6_scaled  = scaler_nan.fit_transform(x_6.reshape(-1, 1)).ravel()
        svr_nan    = SVR(kernel='rbf')
        svr_nan.fit(X_scaled, x6_scaled)

        def interp_value(X):
            return scaler_z.inverse_transform(
                svr_model.predict(scaler_X.transform(X)).reshape(-1, 1)
            ).ravel()

        def g1(x):
            nan_pred = scaler_nan.inverse_transform(
                svr_nan.predict(scaler_X.transform(x.reshape(1, -1))).reshape(-1, 1)
            ).ravel()[0]
            return 5 - np.clip(nan_pred, 0, 100)

        bounds    = [(np.min(x_1), np.max(x_1)), (np.min(x_2), np.max(x_2)),
                     (np.min(x_3), np.max(x_3)), (np.min(x_4), np.max(x_4))]
        objective = lambda x: optimization_direction * interp_value(x.reshape(1, -1))[0]
        cons      = [{'type': 'ineq', 'fun': g1}]

        best_result = None
        for i in range(len(p_param[:, 0])):
            initial_guess = np.array([p_param[i, 0], p_param[i, 1], p_param[i, 2], p_param[i, 3]])
            with warnings.catch_warnings():
                warnings.filterwarnings('ignore', category=RuntimeWarning)
                result = minimize(objective, initial_guess, method='SLSQP',
                                  bounds=bounds, constraints=cons, options=slsqp_options)
            if best_result is None or result.fun <= best_result.fun:
                best_result = result

        if best_result is None:
            raise RuntimeError('Optimisation failed for all initial guesses.')

        solution = best_result.x
        print(f'SLSQP result: {best_result}')
        print(f'Constraint value (should be >= 0): {g1(solution)}')

        file_name            = '_'.join([opt_file_name, f'{indicators[1, 0]}_perlin.tif'])
        perlin_opt_file_path = os.path.join(optimised_save, file_name)
        optimised_perlin     = imgan.generate_single_perlin_image(
            image_height, image_width,
            scale=solution[0], octaves=int(round(solution[1])),
            persistence=solution[2], lacunarity=solution[3],
            texture_function=indicators[1, 0]
        )
        cv2.imwrite(perlin_opt_file_path, optimised_perlin)
        plt.imshow(optimised_perlin, cmap='gray')
        plt.title(f'Optimised pattern: {file_name}')
        plt.show(block=False)
        plt.get_current_fig_manager().window.raise_()
        plt.pause(2)
        plt.close()

    # I tried to include a block for the RTG patterns. But the way RTG patterns were used in this 
    # study involved constent parameter values and they were therefor not optimised.
    # Im not even sure if the code below actually works i just wrote it the same way 
    # as the other blocks above.

    elif gen_value.lower() == 'turing':
        print('Turing optimisation...')
        coordinate_arr = np.vstack((x_1, x_3, x_4, x_5)).T

        scaler_X = StandardScaler()
        scaler_z = StandardScaler()
        X_scaled = scaler_X.fit_transform(coordinate_arr)
        z_scaled = scaler_z.fit_transform(z.reshape(-1, 1)).ravel()
        svr_model = SVR(kernel='rbf')
        svr_model.fit(X_scaled, z_scaled)

        scaler_nan = StandardScaler()
        x6_scaled  = scaler_nan.fit_transform(x_6.reshape(-1, 1)).ravel()
        svr_nan    = SVR(kernel='rbf')
        svr_nan.fit(X_scaled, x6_scaled)

        def interp_value(X):
            return scaler_z.inverse_transform(
                svr_model.predict(scaler_X.transform(X)).reshape(-1, 1)
            ).ravel()

        def g1(x):
            nan_pred = scaler_nan.inverse_transform(
                svr_nan.predict(scaler_X.transform(x.reshape(1, -1))).reshape(-1, 1)
            ).ravel()[0]
            return 15 - np.clip(nan_pred, 0, 100)

        bounds    = [(np.min(x_1), np.max(x_1)), (np.min(x_3), np.max(x_3)),
                     (np.min(x_4), np.max(x_4)), (np.min(x_5), np.max(x_5))]
        objective = lambda x: optimization_direction * interp_value(x.reshape(1, -1))[0]
        cons      = [{'type': 'ineq', 'fun': g1}]

        best_result = None
        for i in range(len(p_param[:, 0])):
            initial_guess = np.array([p_param[i, 0], p_param[i, 2], p_param[i, 3], p_param[i, 4]])
            with warnings.catch_warnings():
                warnings.filterwarnings('ignore', category=RuntimeWarning)
                result = minimize(objective, initial_guess, method='SLSQP',
                                  bounds=bounds, constraints=cons, options=slsqp_options)
            if best_result is None or result.fun <= best_result.fun:
                best_result = result

        if best_result is None:
            raise RuntimeError('Optimisation failed for all initial guesses.')

        solution = best_result.x
        print(f'SLSQP result: {best_result}')
        print(f'Constraint value (should be >= 0): {g1(solution)}')

        file_name            = '_'.join([opt_file_name, 'turing.tif'])
        turing_opt_file_path = os.path.join(optimised_save, file_name)
        generate_and_save(500, 2000, 25.4, solution[0], turing_opt_file_path,
                          size_randomness=solution[3], position_randomness=solution[1],
                          speckle_blur=1.5, grid_step=solution[2])

        turing_image = imgan.make_turing_single(
            turing_opt_file_path, rep=50, radius=1, sharpen_percent=500, size=None
        )
        turing_image.save(turing_opt_file_path)
        plt.imshow(turing_image, cmap='gray')
        plt.title(f'Optimised pattern: {file_name}')
        plt.show(block=False)
        plt.get_current_fig_manager().window.raise_()
        plt.pause(2)
        plt.close()

    else:
        print(f"Unknown pattern type: '{gen_value}'. Expected: speckle, lines, checkerboard, perlin, or turing.")


# Pipeline complete
tic2 = time.time()
print(f'\nProcess completed.')
print(f'Time taken: {tic2 - tic1:.3f} seconds')
print(f'Current time: {datetime.now().strftime("%H:%M:%S")}')