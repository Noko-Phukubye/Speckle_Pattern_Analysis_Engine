"""
Speckle Pattern Evaluation Workflow
===================================

This script implements the full workflow used in the thesis study for
evaluating speckle pattern quality and its influence on Digital Image
Correlation (DIC) accuracy.

The pipeline includes:

1. Creating an Excel workbook used for storing pattern metrics and results.
2. Generating synthetic speckle patterns using several pattern generation
   approaches.
3. Performing speckle pattern analysis using multiple image quality metrics.
4. Loading Finite Element (FE) displacement fields from Nastran output.
5. Deforming reference images using the FE displacement field.
6. Running SUN-DIC to estimate displacement fields.
7. Performing error analysis between FE and DIC displacement results.
8. Performing optimisation studies on speckle pattern parameters.

Each block of the workflow is controlled by a flag to prevent unnecessary
re-computation and to reduce the risk of overwriting intermediate data.

The script is designed to run multiple *batches* of pattern classes in order
to compare different speckle morphologies under identical deformation and
DIC conditions.
"""

print(
    "Importing libraries...\n"
    "---------------------------------------------------------------------------------------"
)

# =============================================================================
# IMPORTS
# =============================================================================
# Third-party scientific libraries
import numpy as np
import matplotlib.pyplot as plt
import cv2
import scipy.stats as stats
from scipy.interpolate import griddata, RBFInterpolator
from scipy.optimize import minimize
from numpy.linalg import norm

# Machine learning tools
from sklearn.preprocessing import PolynomialFeatures, StandardScaler
from sklearn.linear_model import LinearRegression
from sklearn import linear_model
from sklearn.svm import SVR
from sklearn.model_selection import GridSearchCV, train_test_split

# Optimisation tools
from mealpy import PSO, FloatVar

# FE data readers
from pyNastran.bdf.bdf import BDF
from pyNastran.op2.op2 import OP2

# Design of experiments
from pyDOE import lhs

# File and system utilities
import os
import sys
import time
import warnings
import subprocess
import traceback
from datetime import datetime
import configparser

# Excel handling
import openpyxl

# Local project modules
import image_process_tool_box as imgan
import file_paths as path
from speckle_pattern import (
    generate_and_save,
    generate_lines,
    generate_checkerboard
)

# SUN-DIC modules
import sundic.sundic as sdic
import sundic.post_process as sdpp
import sundic.settings as sdset


# =============================================================================
# WORKFLOW DESCRIPTION
# =============================================================================
"""
General workflow implemented in this script.

1. Create an Excel workbook used to store all pattern metrics and analysis
   results. Each section of the script opens the current workbook
   independently to prevent accidental overwriting.

2. Generate a set of speckle patterns (size n) using Latin Hypercube
   Sampling to vary the pattern generation parameters. These images serve
   as the reference images for the DIC analysis.

3. Copy the reference images to the planar images folder using even-numbered
   prefixes (0, 2, 4, ..., n).

4. Load the finite element displacement field from Nastran BDF and OP2 files.

5. Deform the reference images using the FE displacement field. The resulting
   deformed images are saved with odd-numbered prefixes
   (1, 3, 5, ..., n).

6. Run Digital Image Correlation using SUN-DIC and store the displacement
   results as NumPy arrays.

7. Extract the displacement components from the DIC output and compare them
   against the FE displacements to compute error metrics.

8. Perform global optimisation of pattern parameters based on the resulting
   error metrics.
"""


# =============================================================================
# WORKFLOW FLAGS
# =============================================================================
# Each flag activates a block of the workflow. This allows sections of the
# pipeline to be executed independently without repeating previous steps.
FLAGS = {
    "Create_new_excel": False,
    "generate_pattern": False,
    "Perlin_images": False,
    "Pattern analysis": False,
    "Load FEA": False,
    "Deform speckles": False,
    "run_dic": False,
    "run_error": False,
    "error_dist": False,
    "Excel plots": False,
    "Optimisation_poly": False,
}

# Display the status of all flags before execution
imgan.flag_status(FLAGS, 3.5)


tic1 = time.time()

# =============================================================================
# GLOBAL PARAMETERS
# =============================================================================
"""
Global parameters used throughout the workflow.

These variables control:
- image resolution
- number of patterns generated
- deformation parameters
- DIC study configuration
"""

pattern_count = 10

# Image dimensions (pixels)
image_width = 2000
image_height = 500

# Image save path index used by the pattern generation functions
image_save_path = 5

# Translation study parameters
u_min = 0.0
u_max = 1.05
uinterval = 0.05

# Image file format used throughout the workflow
image_type = "tif"


# =============================================================================
# PLOTTING STYLE
# =============================================================================
# Configure matplotlib to produce IEEE-style figures suitable for papers
plt.style.use(["science", "no-latex", "ieee", "grid"])

plt.rcParams.update({
    "font.family": "Calibri",
    "font.size": 12,
    "axes.labelsize": 8,
    "xtick.labelsize": 8,
    "ytick.labelsize": 8,
    "legend.fontsize": 6,
    "lines.linewidth": 1,
    "lines.markersize": 3,
})


# =============================================================================
# STUDY CONFIGURATION
# =============================================================================

# Image deformation scenario
# 0 → FE-based deformation
# 1 → Rigid-body translation
deform_type = 0

# DIC analysis scenario
study_type = 0

# Enable relative error calculations
relative_error = False

# Batch indices representing different pattern classes
batches_range = [0, 1, 2, 3, 4, 5]

# Prefix used for naming optimised images
first_pref = 0
# =============================================================================
# BATCH LOOP
# =============================================================================
# Each batch corresponds to a specific speckle pattern class. The full
# evaluation pipeline is executed for each batch independently so that
# patterns with different morphologies can be analysed under identical
# deformation and DIC conditions.

for batch in batches_range:

    # Optional skip logic for debugging specific batches
    # if batch not in [6]:
    #     first_pref = first_pref + 2
    #     continue

    # Display batch information
    print(f"\nOptimised image name prefix = {first_pref}\n")
    print(f"\nBatch: {batch}...\n")

    # =========================================================================
    # DIRECTORY PATH DEFINITIONS
    # =========================================================================
    """
    Define all directories required for the current batch.

    The workflow generates a large number of intermediate files including:
        - reference images
        - deformed images
        - DIC results
        - error maps
        - plots and debugging figures
        - NumPy data files

    To prevent file conflicts between batches, most directories are created
    using the current batch index.
    """

    rrref                = r"data\speckle_pattern_img\reference_im"
    dddef                = r"data\speckle_pattern_img\deformed_im"
    ooopt                = r"data\speckle_pattern_img\Optimised"

    error_hist_path      = rf"output\histograms\hist_{batch}"
    DIC_contour_path     = rf"output\DIC_contour\contour_{batch}"
    znssd_figure_path    = rf"output\DIC_contour\ZNSSD\znssdbatch_{batch}"
    error_heatmap_path   = rf"output\Heatmaps\heat_{batch}"

    debugg_folder        = rf"output\Debugging"
    excel_path           = rf"output\excel_docs\excel_{batch}"

    deformed_image_path  = rf"data\speckle_pattern_img\deformed_im\def_{batch}"
    reference_image_path = rf"data\speckle_pattern_img\reference_im\ref_{batch}"

    difference_images    = rf"output\plots\Difference_images\diff_{batch}"

    DIC_settings_path    = rf"settings.ini"

    sundic_save          = rf"output\pyth"
    Contour_path         = rf"output\DIC_contour\contour_{batch}"

    spt_img_path         = rf"data\speckle_pattern_img\subpixel_translation"

    numpy_files          = rf"C:\Users\General User\nokop\pattern2\output\numpy_files\npy_{batch}"

    sundic_binary_folder = rf"output\sundic_bin\batch_{batch}"

    slice_path           = rf"output\Slices\slice_{batch}"
    image_matrix         = rf"output\image_matrix"

    autocorr             = rf"output\Autocorrelation\aut_{batch}"
    power_spec           = rf"output\spectral_analysis\spec_{batch}"

    optimised_save       = rf"data\speckle_pattern_img\Optimised\optbatch_{batch}"

    scatter_plots        = rf"output\plots\Scatter_plots\scatter_{batch}"


    # =========================================================================
    # CREATE DIRECTORIES
    # =========================================================================
    """
    Ensure that all directories required by the workflow exist before
    processing begins. Missing directories are created automatically.
    """

    dirs_to_create = [
        rrref,
        dddef,
        ooopt,
        error_hist_path,
        DIC_contour_path,
        znssd_figure_path,
        error_heatmap_path,
        debugg_folder,
        excel_path,
        deformed_image_path,
        reference_image_path,
        difference_images,
        sundic_save,
        spt_img_path,
        sundic_binary_folder,
        slice_path,
        numpy_files,
        image_matrix,
        autocorr,
        power_spec,
        optimised_save,
        scatter_plots
    ]

    for dir_path in dirs_to_create:
        if not os.path.exists(dir_path):
            os.makedirs(dir_path)


    # =========================================================================
    # FINITE ELEMENT DATA PATHS
    # =========================================================================
    """
    Paths to the finite element files used to generate the deformation field.
    These files contain:
        - nodal coordinates (BDF file)
        - nodal displacement results (OP2 file)
    """

    op2_path = path.flat30_4mm_op2
    bdf_path = path.flat30_4mm_bdf


    # Image file extension used throughout the workflow
    image_type = "tif"


    try:

        # =========================================================================
        # CREATE EXCEL DOCUMENT
        # =========================================================================
        if FLAGS['Create_new_excel']:

            """
            Create a new Excel document used to store pattern metrics,
            DIC results, and error statistics.

            Important:
            This flag must be disabled after execution to prevent accidental
            overwriting of existing data.
            """

            print("\n1. Creating excel document...\n")

            if not os.path.exists(excel_path):

                os.makedirs(excel_path)
                doc_number = 1

            else:

                # Determine the next document number based on existing files
                existing_files = [
                    f for f in os.listdir(excel_path)
                    if f.startswith("Pattern_evaluation_") and f.endswith(".xlsx")
                ]

                doc_number = len(existing_files) + 1

            print("Current doc_number:", doc_number)

            exl_file_name = f"Pattern_evaluation_{doc_number}.xlsx"
            exl_file_path = os.path.join(excel_path, exl_file_name)

            workbook = openpyxl.Workbook()
            sheet = workbook.active
            workbook.save(exl_file_path)

            del workbook


        # =========================================================================
        # GENERATE SPECKLE PATTERNS
        # =========================================================================
        if FLAGS['generate_pattern']:

            print("\n2. Generating speckle pattern images...\n")

            # Retrieve the most recent Excel document
            exl_file_path = imgan.excel_doc_path(excel_path)

            texx = "none"
            make_turing = True

            # ---------------------------------------------------------------------
            # Pattern class
            # ---------------------------------------------------------------------

            if batch == 0:
                pattern_method = 1

            elif batch == 1:
                pattern_method = 3

            elif batch == 2:
                pattern_method = 4
                texx = "none"

            elif batch == 3:
                pattern_method = 4
                texx = "cubic"

            elif batch == 4:
                pattern_method = 4
                texx = "perlin_blobs"

            elif batch == 5:
                pattern_method = 4
                texx = "sinusoidal"


            # Same patterns but without RTG transformation

            if batch == 6:
                pattern_method = 1
                make_turing = False

            elif batch == 7:
                pattern_method = 3
                make_turing = False

            elif batch == 8:
                pattern_method = 4
                texx = "none"
                make_turing = False

            elif batch == 9:
                pattern_method = 4
                texx = "cubic"
                make_turing = False

            elif batch == 10:
                pattern_method = 4
                texx = "perlin_blobs"
                make_turing = False

            elif batch == 11:
                pattern_method = 4
                texx = "sinusoidal"
                make_turing = False


            # ---------------------------------------------------------------------
            # Pattern generation
            # ---------------------------------------------------------------------

            if pattern_method == 1:

                # Traditional speckle generator
                imgan.ladisk_generator(
                    exl_file_path,
                    reference_image_path,
                    pattern_count,
                    image_width=image_width,
                    image_height=image_height
                )

            elif pattern_method == 2:

                # Parallel line pattern
                imgan.ladisk_generator_lines(
                    exl_file_path,
                    reference_image_path,
                    pattern_count,
                    image_width=image_width,
                    image_height=image_height
                )

            elif pattern_method == 3:

                # Checkerboard pattern
                imgan.ladisk_generator_cb(
                    exl_file_path,
                    reference_image_path,
                    pattern_count,
                    image_width=image_width,
                    image_height=image_height
                )

            elif pattern_method == 4:

                """
                Perlin noise based speckle patterns.

                Texture options include:
                    none
                    thresholded
                    sinusoidal
                    bimodal
                    logarithmic
                    cubic
                    perlin_blobs
                """

                imgan.single_perlin(
                    image_height,
                    image_width,
                    excel_path=exl_file_path,
                    ref_image_save=reference_image_path,
                    texture_function=texx,
                    number_of_images=pattern_count
                )

            # ---------------------------------------------------------------------
            # Generate Turing-like patterns
            # ---------------------------------------------------------------------

            if make_turing:

                imgan.make_turing(
                    reference_image_path,
                    rep=25,
                    radius=1,
                    sharpen_percent=250,
                    size=None,
                    replace=True,
                    excel_document=exl_file_path
                )

            plt.close("all")
            

        # =========================== IMPORT FINITE ELEMENT MESH =========================
        if FLAGS['Load FEA']:

            """
            Load finite element mesh and displacement results.

            The workflow extracts:
                • nodal coordinates from the BDF file
                • nodal displacements from the OP2 results file

            These are converted into 2D fields (x,y only) because the DIC workflow
            operates on planar images.
            """

            FE_flags = {
                'show meshes': False,
                'show_displacement_field': False
            }

            print('\n4. Loading FEA data...\n')

            # -------------------------------------------------------------------------
            # Load FEA result files
            # -------------------------------------------------------------------------
            model = OP2()
            model.read_op2(op2_path)

            bdf = BDF()
            bdf.read_bdf(bdf_path)

            # -------------------------------------------------------------------------
            # Extract displacement results
            # -------------------------------------------------------------------------
            itime = 0
            isubcase = 1

            disp = model.displacements[isubcase]

            # Extract translational displacement components
            txyz = disp.data[itime, :, :3]

            # Calculate displacement magnitude
            total_xyz = norm(txyz, axis=1)

            nnodes = disp.data.shape[1]

            # -------------------------------------------------------------------------
            # Extract node coordinates from the BDF mesh
            # -------------------------------------------------------------------------
            nodes = np.array([
                bdf.nodes[nid].get_position()
                for nid in bdf.nodes
            ])

            # Convert to 2D (DIC assumes planar deformation)
            nodes_2d = nodes[:, :2]
            displacements_2d = txyz[:, :2]

            # Calculate deformed node positions
            deformed_nodes_2d = nodes_2d + displacements_2d

            # -------------------------------------------------------------------------
            # Scale coordinates so FE mesh aligns with image pixel coordinates
            # -------------------------------------------------------------------------
            x_scale = 1000
            original_points = nodes_2d * x_scale
            new_points = deformed_nodes_2d * x_scale

            # -------------------------------------------------------------------------
            # Optional mesh visualization
            # -------------------------------------------------------------------------
            if FE_flags['show meshes']:

                plt.figure(figsize=(12, 6))

                # Original mesh
                plt.subplot(1, 2, 1)
                plt.scatter(nodes_2d[:, 0], nodes_2d[:, 1],
                            c='b', marker='o', label='Original')
                plt.title('Original Mesh')
                plt.xlabel('X')
                plt.ylabel('Y')
                plt.grid(True)
                plt.axis('equal')

                # Deformed mesh
                plt.subplot(1, 2, 2)
                plt.scatter(deformed_nodes_2d[:, 0], deformed_nodes_2d[:, 1],
                            c='r', marker='o', label='Deformed')
                plt.title('Deformed Mesh')
                plt.xlabel('X')
                plt.ylabel('Y')
                plt.grid(True)
                plt.axis('equal')

                plt.show(block=False)
                plt.pause(5)
                plt.close()

            # -------------------------------------------------------------------------
            # Determine FE-to-image scaling
            # -------------------------------------------------------------------------
            max_x_FE = np.max(nodes_2d[:, 0])
            max_y_FE = np.max(nodes_2d[:, 1])

            print(f'max x = {max_x_FE}\nmax y = {max_y_FE}')

            FE_to_img = image_width / max_x_FE
            print(f'scale = {FE_to_img}')

            x_scale = FE_to_img
            original_points = nodes_2d * x_scale
            new_points = deformed_nodes_2d * x_scale

            print('\nFEA data loaded successfully.')

            # -------------------------------------------------------------------------
            # Interpolate displacement field over image grid
            # -------------------------------------------------------------------------
            dx, dy, FEx_rbf_interp, FEy_rbf_interp = imgan.smooth_field(
                np.zeros((image_height, image_width)),
                original_points,
                new_points,
                3
            )

            # -------------------------------------------------------------------------
            # Optional displacement field visualization
            # -------------------------------------------------------------------------
            if FE_flags['show_displacement_field']:

                matrix = np.zeros((image_height, image_width))

                field = imgan.show_field(
                    matrix,
                    original_points,
                    new_points,
                    debugg_folder
                )

                plt.draw()
                plt.pause(5)
                plt.close('all')

                print('\nFEA data loaded successfully.')


        # ============================== DEFORM IMAGES ===================================
        if FLAGS['Deform speckles']:

            """
            Apply deformation to generated speckle patterns using the interpolated
            displacement field derived from the FEA mesh.

            Deformation pipeline:

                1. Load reference speckle image
                2. Apply displacement field using RBF interpolation
                3. Remap pixel coordinates using scipy / ndimage mapping
                4. Save resulting deformed image
                5. Compute difference images for deformation verification
            """

            deform_flag = {
                'gaussian_blur_ref_img': True,
                'show process': False,
                'Inverse consistency': False
            }

            imgan.flag_status(deform_flag, wait_time=1.5)

            # -------------------------------------------------------------------------
            # Optional Gaussian blur pre-filter
            # -------------------------------------------------------------------------
            if deform_flag['gaussian_blur_ref_img']:

                print('\nReference image prefilter')

                imgan.gaussian_blur_images(
                    reference_image_path,
                    size=5,
                    sig_y=1.0,
                    par='even'
                )

            wait_time = 0.01

            """
            Deformation types:

                0 → FE-based deformation
                1 → Rigid body translation (implemented in the pattern_analysis.py script. Ive never used it here)
            """

            print('\n5. Deforming images...')

            if deform_type == 0:

                print('\nFE-based deformation\n')

                image_files = imgan.get_image_strings(
                    reference_image_path,
                    imagetype=image_type
                )

                for k, ref_image_file in enumerate(image_files):

                    # -----------------------------------------------------------------
                    # Load reference image
                    # -----------------------------------------------------------------
                    ref_img_path = os.path.join(reference_image_path, ref_image_file)

                    reference_image = cv2.imread(ref_img_path)

                    if reference_image is None:
                        print(f"Warning: Could not read reference image {ref_img_path}")
                        continue

                    # Determine numbering for deformed image
                    ref_num = int(ref_image_file.split('_')[0])
                    deform_num = ref_num + 1

                    deformed_name = f"{deform_num}_Generated_spec_image.tif"

                    print(f'\nApplying deformation: Image {k + 1}')

                    tic_def = time.time()

                    # -----------------------------------------------------------------
                    # Apply image deformation
                    # -----------------------------------------------------------------
                    transformed_image, difference_image = imgan.img_deform(
                        reference_image,
                        dx=dx,
                        dy=dy
                    )

                    # Save deformed image
                    deformed_path = os.path.join(deformed_image_path, deformed_name)

                    cv2.imwrite(deformed_path, transformed_image)

                    toc_def = time.time()

                    # -----------------------------------------------------------------
                    # Process difference image
                    # -----------------------------------------------------------------
                    if difference_image.ndim == 3:
                        difference_image = np.mean(difference_image, axis=2)

                    normalised_difference = difference_image / difference_image.max()

                    plt.figure(figsize=(6, 5))
                    plt.title("Difference image")

                    img = plt.imshow(normalised_difference, cmap='jet')
                    plt.colorbar(label="Difference")

                    plt.axis('off')
                    plt.tight_layout()

                    difference_img_save_path = os.path.join(
                        difference_images,
                        f'{ref_num}_difference_image.png'
                    )

                    plt.savefig(
                        difference_img_save_path,
                        bbox_inches='tight',
                        pad_inches=0,
                        dpi=300
                    )

                    plt.show(block=False)
                    plt.pause(wait_time)
                    plt.close()

                    # -----------------------------------------------------------------
                    # Collapsed difference profile
                    # -----------------------------------------------------------------
                    collapsed_difference = np.mean(difference_image, axis=0)

                    x_line = np.arange(difference_image.shape[1])

                    plt.figure(figsize=(10, 6))

                    plt.plot(x_line, collapsed_difference, linewidth=1.5)

                    plt.title("Collapsed Difference Image Along Axis 0")
                    plt.xlabel("Pixel Column Index (X)")
                    plt.ylabel("Mean Pixel Difference")
                    plt.grid(True)

                    slice_save = os.path.join(
                        difference_images,
                        f'{ref_num}_collapsed_difference_image_axis0.png'
                    )

                    plt.savefig(slice_save)
                    plt.close()

                    # -----------------------------------------------------------------
                    # Row slice difference profile
                    # -----------------------------------------------------------------
                    row_index = difference_image.shape[0] // 2

                    row_difference = difference_image[row_index, :]

                    x_line = np.arange(difference_image.shape[1])

                    plt.figure(figsize=(10, 6))

                    plt.plot(x_line, row_difference, linewidth=1.5)

                    plt.title(f"Difference Image Slice at Row {row_index}")
                    plt.xlabel("Pixel Column Index (X)")
                    plt.ylabel("Pixel Difference")

                    plt.grid(True)

                    slice_save = os.path.join(
                        difference_images,
                        f'{ref_num}_slice_row_{row_index}_difference_image.png'
                    )

                    plt.savefig(slice_save)
                    plt.close()

                    print(
                        f'Image {k + 1} deformation complete. '
                        f'{toc_def - tic_def:.3f} seconds\n'
                    )

            elif deform_type == 1:

                print('Rigid-body translation not yet implemented in batch study')

            else:

                print('Invalid input')

            plt.close('all')
            
        # ======================== RUN DIGITAL IMAGE CORRELATION ========================
        if FLAGS['run_dic']:

            """
            Execute Digital Image Correlation (DIC) using the SunDIC solver.

            This block performs three possible tasks depending on the flags:
                1. Pre-filter reference and deformed images (optional)
                2. Run the DIC solver on image pairs
                3. Post-process subpixel translation results (only used on pattern_analysis.py script, not here but could be 
                used potentially)

            The SunDIC solver generates binary `.sdic` result files that contain
            displacement vectors for each subset location in the image.
            """

            # Control which parts of the DIC pipeline should execute
            DIC_flags = {
                "Image_blur": False,
                "run_analysis": True,
                "process_subpixel_data": True
            }

            imgan.flag_status(DIC_flags, wait_time=1.5)

            # -------------------------------------------------------------------------
            # Optional Gaussian pre-filter
            # -------------------------------------------------------------------------
            if DIC_flags['Image_blur']:

                print('\nReference image prefilter')

                imgan.gaussian_blur_images(
                    reference_image_path,
                    size=5,
                    sig_y=1.0,
                    par='even'
                )

                print('\nDeformed image prefilter')

                imgan.gaussian_blur_images(
                    deformed_image_path,
                    size=5,
                    sig_y=1.0,
                    par='odd'
                )

            print("\n6. Running DIC analysis...\n")

            # -------------------------------------------------------------------------
            # Load DIC configuration settings
            # -------------------------------------------------------------------------
            if not os.path.exists(DIC_settings_path):
                print(f"Settings file not found at: {DIC_settings_path}")

            settings = sdset.Settings.fromSettingsFile(DIC_settings_path)

            # -------------------------------------------------------------------------
            # Region of Interest (ROI)
            #
            # ROI is automatically scaled based on the image size. A small margin is
            # removed from each side to prevent subsets extending outside the image.
            # -------------------------------------------------------------------------
            roi = 'auto'
            start_x = 30
            start_y = 30

            if roi.lower() == 'auto':
                width = image_width - (2 * start_x)
                height = image_height - (2 * start_y)
            else:
                width = 1990
                height = 490

            settings.ROI = [start_x, start_y, width, height]

            # Core DIC parameters
            settings.SubsetSize = 67
            settings.StepSize = 23
            settings.GaussianBlurSize = 5
            settings.GaussianBlurStdDev = 0.0
            settings.DebugLevel = 1
            settings.CPUCount = 8
            settings.OptimizationAlgorithm = "IC-GN"
            settings.ReferenceStrategy = "Absolute"

            # -------------------------------------------------------------------------
            # Select deformation dataset depending on study type
            # -------------------------------------------------------------------------
            if study_type == 0:

                Deformed_images = deformed_image_path
                sundic_binary_file_loc = sundic_binary_folder

            elif study_type == 1:

                Deformed_images = spt_img_path
                sundic_binary_file_loc = os.path.join(
                    sundic_binary_folder,
                    "subpixel"
                )

                if not os.path.exists(sundic_binary_file_loc):
                    os.makedirs(sundic_binary_file_loc)

            print(settings.__repr__())
            print("Settings loaded successfully\n")

            # -------------------------------------------------------------------------
            # Ensure Ray parallel processing service is running
            # -------------------------------------------------------------------------
            try:

                print('Checking ray...')
                subprocess.run('ray status', shell=True, check=True)
                time.sleep(5)

            except subprocess.CalledProcessError:

                print('Starting Ray...')
                subprocess.run('ray start --head --num-cpus=8', shell=True)

                time.sleep(5)
                subprocess.run('ray status', shell=True)

            # -------------------------------------------------------------------------
            # Run DIC solver
            # -------------------------------------------------------------------------
            if DIC_flags["run_analysis"]:

                DIC_log_name = f"DIC_log_{batch}.txt"

                imgan.run_dic(
                    settings,
                    reference_image_path,
                    Deformed_images,
                    sundic_binary_file_loc,
                    debugg_folder,
                    DIC_contour_path,
                    znssd_figure_path,
                    study_type,
                    image_type,
                    start_index=0,
                    umin=u_min,
                    umax=u_max,
                    u_interval=uinterval,
                    dic_log_file_name=DIC_log_name
                )

            # -------------------------------------------------------------------------
            # Subpixel translation analysis (for rigid translation studies)
            # -------------------------------------------------------------------------
            if study_type == 1 and DIC_flags["process_subpixel_data"]:

                error_matrix, subpixel_trans_figure = imgan.subpixel_analysis(
                    sundic_binary_file_loc,
                    plot_path,
                    u_min=u_min,
                    u_max=u_max,
                    u_interval=uinterval
                )

                matrix_save = os.path.join(numpy_files, 'error_matrix.npy')
                np.save(matrix_save, error_matrix)

            plt.close('all')


        # ============================= ANALYSE PATTERNS ================================
        if FLAGS['Pattern analysis']:

            """
            Compute speckle pattern quality metrics for each generated reference image.

            These metrics quantify:
                • gradient information
                • spatial frequency content
                • entropy
                • morphological structure

            Results are written to the active Excel workbook.
            """

            print('\n3. Analysing speckle patterns...\n')

            try:

                exl_file_path = imgan.excel_doc_path(excel_path)

                try:
                    workbook = openpyxl.load_workbook(exl_file_path)
                except FileNotFoundError:
                    workbook = openpyxl.Workbook()

                sheet = workbook.active

                # Column headers
                sheet['A1'] = 'MSF'
                sheet['B1'] = 'MIG'
                sheet['C1'] = 'E_f'
                sheet['D1'] = 'MIOSD'
                sheet['E1'] = 'Shannon'
                sheet['F1'] = 'Power spectrum area'
                sheet['G1'] = 'SSSIG global'
                sheet['H1'] = 'Correlation peak radius'

                # Ensure consistent naming of image files
                imgan.rename_img(reference_image_path)

                image_files = imgan.get_image_strings(
                    reference_image_path,
                    imagetype=image_type
                )

                first_file = image_files[0]
                image_file_endswith = '_'.join(first_file.split('_')[1:])

                reference_image_prefixes = imgan.expected_prefixes(reference_image_path)
                prefix_index = {
                    prefix: i
                    for i, prefix in enumerate(reference_image_prefixes)
                }

                for prefix_num in reference_image_prefixes:

                    j = prefix_index[prefix_num] + 2

                    ref_image_file = f'{prefix_num}_{image_file_endswith}'
                    ref_img_path = os.path.join(reference_image_path, ref_image_file)

                    try:

                        if not os.path.exists(ref_img_path):
                            raise Exception(
                                f'Image file {ref_image_file} not found'
                            )

                        Generated_speckles = imgan.readImage(ref_img_path)

                        print(f'{ref_image_file}:')

                        msf = imgan.MSF(Generated_speckles)
                        mean_intensity_gradient = imgan.MIG(Generated_speckles)

                        subset_entropy = imgan.meanSE(Generated_speckles, 33, 5)

                        mag, powr, fx, fy = imgan.compute_fft_and_freq(
                            Generated_speckles
                        )

                        power_area = imgan.integrate_power_2d(powr, fx, fy)

                        power_fig = imgan.plot_1d_spectra(
                            mag,
                            powr,
                            fx,
                            fy,
                            show=False
                        )

                        power_figsave_path = os.path.join(
                            power_spec,
                            f'Power_spectrum_{ref_image_file}.png'
                        )

                        power_fig.savefig(power_figsave_path)
                        plt.close(power_fig)

                        sss = imgan.globalSSSIG(Generated_speckles, 33, 10)
                        shannon = imgan.ShannonEnt(Generated_speckles)
                        miosd = imgan.miosd(Generated_speckles)
                        ef = imgan.Ef(Generated_speckles)

                        # Autocorrelation analysis
                        _, figR, R_peak = imgan.autocorr(
                            Generated_speckles,
                            meth=6,
                            cardinality=3,
                            autype='2d'
                        )

                        figsave_path = os.path.join(
                            autocorr,
                            f'Autocorrelation_{ref_image_file}.png'
                        )

                        plt.close('all')

                        sheet[f'A{j}'] = msf
                        sheet[f'B{j}'] = mean_intensity_gradient
                        sheet[f'C{j}'] = ef
                        sheet[f'D{j}'] = miosd
                        sheet[f'E{j}'] = shannon
                        sheet[f'F{j}'] = power_area
                        sheet[f'G{j}'] = sss

                    except Exception:

                        sheet[f'A{j}'] = 0
                        sheet[f'B{j}'] = 0
                        sheet[f'C{j}'] = 0
                        sheet[f'D{j}'] = 0
                        sheet[f'E{j}'] = 0
                        sheet[f'F{j}'] = 0
                        sheet[f'G{j}'] = 0
                        sheet[f'H{j}'] = 0

                workbook.save(exl_file_path)
                print(f"Workbook saved at {exl_file_path}")

            except FileNotFoundError as fnf_error:

                print(f"File or folder not found: {fnf_error}")

            except Exception as e:

                tb = traceback.extract_tb(sys.exc_info()[2])
                filename, line_number, function_name, text = tb[-1]

                print(f'Error in pattern analysis on line {line_number}: {str(e)}')
                print(f'Look at {filename}')

            finally:

                if 'workbook' in locals():
                    workbook.close()

            plt.close('all')

       # ======================== ERROR DISTRIBUTION ANALYSIS =========================
        if FLAGS['error_dist']:

            """
            Perform error distribution analysis between DIC results and FEM displacements.

            Steps:
                1. Loop through all DIC binary result files (.sdic)
                2. Interpolate FEM displacement values onto DIC points
                3. Compute error grids (absolute or relative)
                4. Generate heatmaps, histograms, and collapsed X-axis plots
                5. Save all results to numpy files and figures
            """

            print('\n8. Error distribution analysis...\n')

            # -------------------------------------------------------------------------
            # Collect DIC result files
            # -------------------------------------------------------------------------
            sundic_binary_files = sorted(
                [f for f in os.listdir(sundic_binary_folder) if f.endswith('results.sdic')],
                key=lambda x: int(x.split('_')[0])
            )
            print(f'DIC files found: {sundic_binary_files}')

            # Get expected prefixes for indexing
            all_expected_prefixesbin = imgan.expected_prefixes(sundic_binary_folder, odd=False, skip=True)
            prefix_positionsbin = {prefix: i for i, prefix in enumerate(all_expected_prefixesbin)}

            # -------------------------------------------------------------------------
            # Prepare FEM displacement data
            # -------------------------------------------------------------------------
            fem_xcoord = nodes_2d[:, 0] * x_scale
            fem_ycoord = nodes_2d[:, 1] * x_scale
            fem_x_disp = displacements_2d[:, 0] * x_scale
            fem_y_disp = displacements_2d[:, 1] * x_scale
            fem_mag = np.sqrt(fem_x_disp**2 + fem_y_disp**2)

            # Apply RBF lag correction
            lag_dx = FEx_rbf_interp(np.column_stack([fem_xcoord, fem_ycoord]))
            lag_dy = FEy_rbf_interp(np.column_stack([fem_xcoord, fem_ycoord]))
            fem_points = np.column_stack((fem_xcoord - lag_dx, fem_ycoord - lag_dy))
            print(f'FEM points shape: {fem_points.shape}')

            # Initialize array to store mean error per file
            meanmean = np.zeros(len(sundic_binary_files))
            print(f'Mean error array initialized: {meanmean}')

            # -------------------------------------------------------------------------
            # Loop through each prefix / DIC result file
            # -------------------------------------------------------------------------
            i = 0
            for prefix_num in all_expected_prefixesbin:
                try:

                    print(f'\nProcessing pattern: {prefix_num}')

                    # Construct full path to DIC result file
                    sunfile = f'{prefix_num}_results.sdic'
                    sundic_data_path = os.path.join(sundic_binary_folder, sunfile)
                    if not os.path.exists(sundic_data_path):
                        print(f'File not found: {sundic_data_path}, skipping.')
                        continue

                    # Load DIC displacement data
                    sundic_data, nRows, nCols = sdpp.getDisplacements(sundic_data_path, -1, smoothWindow=25)
                    x_coord, y_coord = sundic_data[:, 0], sundic_data[:, 1]
                    X_disp, Y_disp = sundic_data[:, 3], sundic_data[:, 4]
                    dic_mag = np.sqrt(X_disp**2 + Y_disp**2)
                    sundic_points = np.column_stack((x_coord, y_coord))
                    print(f'DIC points shape: {sundic_points.shape}')

                    # ---------------------------------------------------------------------
                    # Select displacement component for analysis
                    # ---------------------------------------------------------------------
                    value = 0  # 0->X, 1->Y, 2->Magnitude
                    if value == 0:
                        fem_value, dic_value, string = fem_x_disp, X_disp, 'X_disp'
                    elif value == 1:
                        fem_value, dic_value, string = fem_y_disp, Y_disp, 'Y_disp'
                    elif value == 2:
                        fem_value, dic_value, string = fem_mag, dic_mag, 'Magnitude'
                    else:
                        raise ValueError("Invalid displacement value selection")

                    # ---------------------------------------------------------------------
                    # Interpolate FEM onto DIC points and compute errors
                    # ---------------------------------------------------------------------
                    interpolated_FEM_values = griddata(fem_points, fem_value, sundic_points, method='cubic')
                    dic_value = dic_value.reshape(nCols, nRows)
                    interpolated_FEM_values = interpolated_FEM_values.reshape(nCols, nRows)
                    errors2 = ((dic_value - interpolated_FEM_values) / interpolated_FEM_values) if relative_error else (dic_value - interpolated_FEM_values)

                    # Save error numpy file
                    np.save(os.path.join(numpy_files, f"{prefix_num}_{batch}_errors.npy"), errors2)

                    mean_err = np.nanmean(errors2)
                    meanmean[i] = mean_err
                    i += 1
                    print(f'Mean error for pattern {prefix_num}: {mean_err:.4f}')

                    # ---------------------------------------------------------------------
                    # Generate heatmap
                    # ---------------------------------------------------------------------
                    plt.figure(figsize=(10, 8))
                    plt.imshow(
                        errors2.T, cmap='jet', interpolation='none',
                        extent=(150, x_coord.max(), y_coord.min(), y_coord.max()),
                        origin='lower'
                    )
                    plt.colorbar(label=f'Error {string}')
                    plt.title(f'{string} Error Distribution: Pattern {prefix_num}')
                    plt.xlabel('X')
                    plt.ylabel('Y')
                    plt.gca().invert_yaxis()
                    plt.figtext(0.5, 0.02, f'Mean error: {mean_err:.4f}', size=14)
                    plt.savefig(os.path.join(error_heatmap_path, f'{prefix_num}_heatmap_{string}.png'))
                    plt.close()

                    # ---------------------------------------------------------------------
                    # Generate histogram
                    # ---------------------------------------------------------------------
                    plt.figure(figsize=(12, 10))
                    plt.hist(errors2[~np.isnan(errors2)].flatten(), bins=250, density=True)
                    plt.title(f'Error Distribution: Pattern {prefix_num}')
                    plt.xlabel(f'Error {string}')
                    plt.ylabel('Frequency')
                    plt.xlim(-0.02, 0.02)
                    plt.ylim(0, 1000)
                    plt.savefig(os.path.join(error_hist_path, f'{prefix_num}_histogram_{string}.png'))
                    plt.close()

                    # ---------------------------------------------------------------------
                    # Collapsed error along X-axis
                    # ---------------------------------------------------------------------
                    x_grid = x_coord.reshape(nCols, nRows)
                    collapsed_errorgrid = np.nanmean(errors2, axis=1)
                    x_line = np.mean(x_grid, axis=1)
                    x_line_mask = x_line >= 150

                    slice_numpy = np.column_stack((collapsed_errorgrid, x_line))
                    os.makedirs(r"output\Slices\slice_binaries", exist_ok=True)
                    np.save(os.path.join(r"output\Slices\slice_binaries", f'{prefix_num}_slice_{string}.npy'), slice_numpy)

                    plt.figure(figsize=(10, 6))
                    plt.plot(x_line[x_line_mask], collapsed_errorgrid[x_line_mask], color='blue', linewidth=1.5)
                    plt.title("Collapsed Error Grid")
                    plt.xlabel("Pixels")
                    plt.ylabel(f"Error {string}")
                    plt.ylim(-0.02, 0.02)
                    plt.grid(True)
                    plt.savefig(os.path.join(slice_path, f'{prefix_num}_slice_{string}.png'))
                    plt.close()

                except Exception as e:
                    tb = traceback.extract_tb(sys.exc_info()[2])[-1]
                    print(f'Error with file {sundic_data_path}: {e}\nLine: {tb.lineno}')


       # =========================== READ EXCEL & PLOT ===============================
        if FLAGS['Excel plots']:

            """
            Read the evaluation Excel document and generate plots for pattern metrics.

            Steps:
                1. Read pattern metrics, measurement errors, and parameters from Excel
                2. Apply filters to exclude outliers
                3. Generate 2D, random, and 3D scatter plots as configured
                4. Identify maximum and minimum values for all metrics
            """

            print('\n8. Reading Excel document...')
            custom_excel = rf'output\excel_docs\excel_{batch}'

            # ---------------------------------------------------------
            # Read metrics, measurement errors, and parameters
            # ---------------------------------------------------------
            p_metrics, meas_error, p_param, nans, indicators = imgan.read_spec_excel(
                custom_excel, doc_num=None
            )
            print('\nIndicators:', indicators)

            # ---------------------------------------------------------
            # Define masks and filters
            # ---------------------------------------------------------
            # Exclude entries where discrepancy between DIC2FE and FE2DIC exceeds threshold
            threshold = 0.01
            outlier_idx = imgan.array_difference_outlier(meas_error[:, 0], meas_error[:, 2], threshold)

            # Remove values above a y-value threshold
            threshold_y = 100  
            outlier_idx_y = [i for i in range(meas_error.shape[0]) if abs(meas_error[i, 0]) > threshold_y]

            # Combine outlier indices
            combined_outliers = outlier_idx + outlier_idx_y

            # Final mask: valid entries within 3-sigma of mean and not zero
            mask = np.array([
                i not in combined_outliers for i in range(meas_error.shape[0])
            ]) & (meas_error[:, 0] != 0) & (np.abs(meas_error[:, 0] - np.nanmean(meas_error[:, 0])) <= 3 * np.nanstd(meas_error[:, 0]))

            # ---------------------------------------------------------
            # Plot configuration
            # ---------------------------------------------------------
            figflag = {
                '2d scatter': True,
                'random_scatter_plot': False,
                '3d scatter': False,
                'xtreme_index': False
            }
            save_fig = True

            pattern_class = ('RS','RCh','RP','RPc','RPbl','RPS','S','Ch','P','PC','Pbl','PS')
            metric_strings = (
                "MSF", 
                "MIG", 
                "$E_f$", 
                "MIOSD",
                "Shannon entropy", 
                "PSA", 
                "SSSIG", 
                "$R_{peak}$"
            )
            error_strings = ('RMSE [pixel]', 'IQR', 'Systematic error', 'Standard deviation')
            error_index = [0]  # Only RMSE for plotting

            # ====================== 2D SCATTER PLOTS ======================
            if figflag['2d scatter']:
                for number, metric in enumerate(metric_strings):
                    err_col = 0  # Use RMSE

                    x_value = p_metrics[mask, number]
                    y_value = np.abs(meas_error[mask, err_col])
                    x_label = f'{metric}'
                    y_label = 'RMSE [pixel]'
                    plot_title = f"Valid patterns = {len(y_value)}"

                    # Plot configuration
                    fig, ax = plt.subplots(figsize=(4, 3))
                    ax.scatter(x_value, y_value, color='black', s=4)
                    ax.set_xlabel(x_label, fontsize=13)
                    ax.set_ylabel(y_label, fontsize=13)
                    ax.minorticks_on()
                    ax.grid(True, which='major', alpha=0.3, linewidth=0.8)
                    ax.grid(True, which='minor', alpha=0.3, linestyle=':', linewidth=0.5)

                    # Configure all spines
                    for spine in ax.spines.values():
                        spine.set_linewidth(1.0)
                        spine.set_edgecolor('black')
                        spine.set_visible(True)

                    # Configure ticks
                    ax.tick_params(axis='both', which='major', width=1.0, length=5, labelsize=10, direction='in')
                    ax.tick_params(axis='both', which='minor', width=0.6, length=3, direction='in')

                    plt.tight_layout(pad=1.5)

                    # Save figure
                    if save_fig:
                        plot_save = os.path.join(
                            scatter_plots,
                            f'{pattern_class[batch]}_{x_label}_vs_{y_label}.png'
                        )
                        plt.savefig(plot_save, dpi=300, bbox_inches='tight', pad_inches=0.2)

                    plt.close()

            # ====================== RANDOM SCATTER PLOT ======================
            if figflag['random_scatter_plot']:
                x_value = p_metrics[mask, number]
                y_value = meas_error[mask, 4]
                x_label = f'{metric}'
                y_label = error_strings[1]
                plot_title = "Batch_images"

                r = np.corrcoef(x_value, y_value)
                plt.figure(figsize=(7, 6))
                plt.scatter(x_value, y_value, color='black')
                plt.xlabel(x_label)
                plt.ylabel(y_label)
                plt.title(plot_title)
                plt.text(0.95, 1.02, f'R = {r[0, 1]:.2f}', transform=plt.gca().transAxes, ha='right')
                plt.grid(True)

                if save_fig:
                    plot_save = os.path.join(scatter_plots, f'{plot_title}_{x_label}vs{y_label}.png')
                    plt.savefig(plot_save)
                plt.close()

            # ====================== 3D SCATTER PLOT ======================
            if figflag['3d scatter']:
                fig = plt.figure()
                ax = fig.add_subplot(projection='3d')
                scatter = ax.scatter(
                    p_metrics[mask, 1], p_metrics[mask, 4], meas_error[mask, 0],
                    c=meas_error[mask, 0], cmap='jet', marker='o'
                )
                plt.colorbar(scatter, ax=ax, label='RMSE (DIC)')
                ax.set_xlabel('Speckle size')
                ax.set_ylabel('Speckle density')
                ax.set_zlabel('RMSE')
                ax.set_title('RMSE vs speckle size vs speckle density')
                ax.set_zlim(0, 0.005)
                plt.show()

            # ====================== EXTREME METRIC INDEX ======================
            if figflag['xtreme_index']:
                # Maximum values
                max_values = [np.max(p_metrics[:, i]) for i in range(p_metrics.shape[1])]
                max_indices = [np.argmax(p_metrics[:, i]) for i in range(p_metrics.shape[1])]
                print('\nMaximum metric values:')
                for i, val in enumerate(max_values):
                    print(f'{metric_strings[i] if i < len(metric_strings) else "Other"}: Pattern {max_indices[i]*2} -> {val:.3f}')

                print('---------------------------------------------------------------')

                # Minimum values
                min_values = [np.min(p_metrics[:, i]) for i in range(p_metrics.shape[1])]
                min_indices = [np.argmin(p_metrics[:, i]) for i in range(p_metrics.shape[1])]
                print('\nMinimum metric values:')
                for i, val in enumerate(min_values):
                    print(f'{metric_strings[i] if i < len(metric_strings) else "Other"}: Pattern {min_indices[i]*2} -> {val:.3f}')
                print('---------------------------------------------------------------')


        '============================PATTERN OPTIMISATION================================='
        if FLAGS['Optimisation_poly']:

            print('\n9. Optimising speckle patterns...')

            # 0 -> SVR(); 
            # 1 -> polynomial regression
            metamodel = 0

            # 0 -> SLSQP with ftol
            # 1 -> trust-constr with gtol
            optimiser = 1
            


            # Read data
            custom_excel = rf'output\excel_docs\excel_{batch}'

            # -------------------------------------------------------
            p_metrics, meas_error, p_param, nans, indicators = imgan.read_spec_excel(custom_excel, doc_num = None)
            # -------------------------------------------------------
            print('\nIndicators:', indicators)

            """
            Excel document will be used for accessing filled-in values to act as 
            training data for optimisation. Will use maxrow from previous section.

            Zero values, NANs and outtliers filtered out
            """

            errors = meas_error[:, 0]
            mymask = (~np.isnan(errors)) & (errors != 0) & (np.abs(errors - np.nanmean(errors)) <= 3 * np.nanstd(errors))

            # mymask = (np.abs(errors - np.nanmean(errors)) <= 3 * np.nanstd(errors))
            # meas_error[meas_error[:, 2] == 0, 0] = 0.1

            mp_param = p_param[mymask, :]
            mmeas_error = meas_error[mymask, :]
            mp_metrics = p_metrics[mymask, :]
            mnans = nans[mymask, :]

            
            print('P_param with mask', mp_param.shape)
            print('meas_error with mask', mmeas_error.shape)
            print('p_metrics with mask', mp_metrics.shape)

            # Read second last column for one of the following terms
            """
            speckle
            lines
            checkerboard
            perlin
            """
            # -----------------------------------------------------
            override_optimisation = False

            if override_optimisation:
                gen_value = 'speckle'
            else:
                gen_value = indicators[0, 0]

            p_grid = {
                'kernel': ['rbf'],  
                'C': np.logspace(-3, 3, 7),
                'gamma' : ('auto','scale')
            }

            # -----------------------------------------------------

            print('Gen value =', gen_value)
            # z = np.abs(mmeas_error[:, 0])  # Make absolute if looking at mean bias error
            '''
            Indexes     :0 MSF
                        1 MIG
                        2 E_f
                        3 MIOSD
                        4 Shannon entropy
                        5 Power spectrum area
                        6 SSSIG
                        7 R_peak
            ''' 
            metric_objective = 5
            z = mp_metrics[:,metric_objective]

            optimization_direction = 1  # -1 for maximisation

            if optimiser==0:
                print("--------------------------------------")
                print("SLSQP optimiser")
            else:
                print("--------------------------------------")
                print("trust-constr optimiser")

            if optimization_direction == 1:

                print('\nMinimising objective function\n'
                    '---------------------------------------')
            else:
                print('\nMaximising objective function\n'
                    '---------------------------------------')

            # Naming
            opt_file_name = f'{first_pref}_Generated_spec_image.tif'

            # Get metadata (masked)
            x_1 = mp_param[:,0]      
            x_2 = mp_param[:,1]    
            x_3 = mp_param[:,2]   
            x_4 = mp_param[:,3]     
            x_5 = mp_param[:,4]

            # Extra parameter (s) for turing (RTG)      
            # if Turing_pattern_transform:
            #     x_7 = indicators[0,2]
            # For nans
            x_6 = mnans[:,0]
            nan_threshold = 0.75

            # ------------------------------------------------------------
            if gen_value.lower() == 'speckle':
                # Coordinate array
                coordinate_arr = np.vstack((x_1, x_3, x_4, x_5)).T

                # Choose between metamodels
                if metamodel == 0:
                    # SVR() model
                    scaler_X = StandardScaler()
                    X_scaled = scaler_X.fit_transform(coordinate_arr)

                    scaler_z = StandardScaler()
                    # Scaler expects a 2D array
                    z_scaled = scaler_z.fit_transform(z.reshape(-1, 1)).ravel()

                    svr_model = SVR()
                    grid = GridSearchCV(svr_model, param_grid=p_grid, cv=5, n_jobs=-1)
                    grid.fit(X_scaled, z_scaled)
                    svr_model = grid.best_estimator_

                    def interp_value(X):
                        """
                        SVR interpolation. Returns unscaled value.
                        """
                        X_scaled = scaler_X.transform(X)
                        z_pred_scaled = svr_model.predict(X_scaled)
                        return scaler_z.inverse_transform(z_pred_scaled.reshape(-1, 1)).ravel()

                    # Define constraint inequality
                    scaler_nan = StandardScaler()
                    x6_scaled = scaler_nan.fit_transform(x_6.reshape(-1, 1)).ravel()

                    svr_nan = SVR()
                    grid_nan = GridSearchCV(svr_nan, param_grid=p_grid, cv=5, n_jobs=-1)
                    grid_nan.fit(X_scaled, x6_scaled)
                    svr_nan = grid_nan.best_estimator_

                    def g1(x):
                        nan_scaled = svr_nan.predict(scaler_X.transform(x.reshape(1, -1)))
                        nan_pred = scaler_nan.inverse_transform(nan_scaled.reshape(-1, 1)).ravel()[0]
                        nan_pred = np.clip(nan_pred, 0, 100)
                        # returns g(x) = Threshold - predicted_NAN_ratio(X) 
                        return nan_threshold - nan_pred
                    
                    def g2(x):
                        """
                        Constraint to ensure the interpolated objective is positive.
                        Returns positive when interp_value(x) > 0.
                        """
                        return interp_value(x.reshape(1, -1))[0]
                    
                elif metamodel == 1:
                    # Polynomial regression (Need to change to Gaussian process)
                    orderr = 3
                    polyfeatures = PolynomialFeatures(degree=orderr)
                    polyfit_X_values = polyfeatures.fit_transform(coordinate_arr)
                    polynomial_model = LinearRegression()
                    polynomial_model.fit(polyfit_X_values, z)

                    # Callable function
                    def interp_value(X):
                        """
                        Polynomial interpolation for optimization
                        """
                        # For multiple points evaluation
                        X_poly = polyfeatures.transform(X)
                        results = polynomial_model.predict(X_poly)
                        return results
                    
                    # Optimisation constraint
                    nan_model = LinearRegression().fit(coordinate_arr, x_6)

                    def g1(x):
                        """
                        Constraint function for optimization: ensures NaN percentage stays below 5%.
                        Returns positive value when constraint is satisfied (NaN% < 5%).
                        """
                        nan_pred = nan_model.predict(x.reshape(1, -1))[0]
                        # nan_pred = np.clip(nan_pred, 0, 100)
                        return nan_threshold - nan_pred

                # Bounds defined within parameter space
                bounds = [
                    (np.min(x_1), np.max(x_1)),
                    (np.min(x_3), np.max(x_3)),
                    (np.min(x_4), np.max(x_4)),
                    (np.min(x_5), np.max(x_5))
                ]

                def objective(x):
                    return optimization_direction * interp_value(x.reshape(1, -1))[0]

                best_result = None
                lower_bounds = np.array([b[0] for b in bounds])
                upper_bounds = np.array([b[1] for b in bounds])

                for i in range(len(mp_param[:, 0])):

                    # base guess from mp_param
                    base_guess = np.array([
                        mp_param[i, 0],
                        mp_param[i, 2],
                        mp_param[i, 3],
                        mp_param[i, 4]
                    ])

                    # add small random noise (±5% of parameter range)
                    noise = np.random.uniform(-0.05, 0.05, size=base_guess.shape) * (upper_bounds - lower_bounds)
                    initial_guess = np.clip(base_guess + noise, lower_bounds, upper_bounds)

                    cons = [{'type': 'ineq', 'fun': g1}]

                    with warnings.catch_warnings():
                        warnings.filterwarnings("ignore", category=RuntimeWarning)
                        warnings.filterwarnings("ignore", category=UserWarning)

                        if optimiser == 0:
                            result = minimize(
                                objective,
                                initial_guess,
                                method='SLSQP',
                                bounds=bounds,
                                constraints=cons,
                                options={
                                    'ftol': 1e-8,
                                    'gtol': 1e-8,
                                    'xtol': 1e-8,
                                    'maxiter': 5000,
                                    'disp': False
                                })

                        elif optimiser == 1:
                            result = minimize(
                                objective,
                                initial_guess,
                                method='trust-constr',
                                bounds=bounds,
                                constraints=cons,
                                options={
                                    'gtol': 1e-6,
                                    'maxiter': 1000
                                })

                    if best_result is None or result.fun <= best_result.fun:
                        best_result = result

                if best_result is None:
                    raise RuntimeError("Optimization failed for all initial guesses.")

                print(f" Result: {best_result}")
                solution = best_result.x
                print("Constraint equation (g1(X)) value (should be >= 0):", g1(solution))

                # Generate and save
                custom_name = 'speckles.tif'                            # If necessary
                file_name = '_'.join([opt_file_name, custom_name])

                spec_opt_file_path = os.path.join(optimised_save, opt_file_name)
                Generated_speckles = generate_and_save(
                    image_height, image_width, 25.4, solution[0], spec_opt_file_path,
                    size_randomness=solution[3], position_randomness=solution[1],
                    speckle_blur=1.5, grid_step=solution[2])


                optimig = imgan.MIG(Generated_speckles)
                print(f"\n\nimage shape = {Generated_speckles.shape}\nOptimised MIG value: {optimig}\n")
                plt.imshow(Generated_speckles, cmap='gray')
                plt.title(f'Optimised pattern: {file_name}')
                plt.show(block=False)
                plt.get_current_fig_manager().window.raise_()
                plt.pause(2)
                plt.close()


            # ------------------------------------------------------------
            elif gen_value.lower() == 'lines':
                # Coordinate array
                coordinate_arr = np.vstack((x_2, x_3)).T

                # SVR surrogate model
                scaler_X = StandardScaler()
                X_scaled = scaler_X.fit_transform(coordinate_arr)

                scaler_z = StandardScaler()
                z_scaled = scaler_z.fit_transform(z.reshape(-1, 1)).ravel()

                svr_model = SVR()
                p_grid = {
                    'kernel': ['rbf'],  
                    'C': np.logspace(-2, 2, 5),
                }
                grid = GridSearchCV(svr_model, param_grid=p_grid, cv=5, n_jobs=-1)
                grid.fit(X_scaled, z_scaled)
                svr_model = grid.best_estimator_

                def interp_value(X):
                    """
                    Metamodel call function 
                    """
                    X_scaled = scaler_X.transform(X)
                    z_pred_scaled = svr_model.predict(X_scaled)
                    return scaler_z.inverse_transform(z_pred_scaled.reshape(-1, 1)).ravel()

                # Define constraint inequality
                scaler_nan = StandardScaler()
                x6_scaled = scaler_nan.fit_transform(x_6.reshape(-1, 1)).ravel()

                svr_nan = SVR()
                grid_nan = GridSearchCV(svr_nan, param_grid=p_grid, cv=5, n_jobs=-1)
                grid_nan.fit(X_scaled, x6_scaled)
                svr_nan = grid_nan.best_estimator_

                def g1(x):
                    nan_scaled = svr_nan.predict(scaler_X.transform(x.reshape(1, -1)))
                    nan_pred = scaler_nan.inverse_transform(nan_scaled.reshape(-1, 1)).ravel()[0]
                    nan_pred = np.clip(nan_pred, 0, 100)
                    return nan_threshold - nan_pred
                
                bounds = [
                    (np.min(x_2), np.max(x_2)),
                    (np.min(x_3), np.max(x_3))
                ]

                def objective(x):
                    return optimization_direction * interp_value(x.reshape(1, -1))[0]

                best_result = None
                for i in range(len(p_param[:, 0])):
                    initial_guess = np.array([
                        p_param[i, 1],
                        p_param[i, 2]
                    ])
                    cons = [{'type': 'ineq', 'fun': g1}]

                    with warnings.catch_warnings():
                        warnings.filterwarnings("ignore", category=RuntimeWarning)
                        warnings.filterwarnings("ignore", category=UserWarning)
                        if optimiser == 0:
                            result = minimize(
                                objective,
                                initial_guess,
                                method='SLSQP',
                                bounds=bounds,
                                constraints=cons,
                                options={
                                    'ftol': 1e-8,
                                    'maxiter': 1000}
                            )
                            
                        elif optimiser == 1:
                            result = minimize(
                                objective,
                                initial_guess,
                                method='trust-constr',
                                bounds=bounds,
                                constraints=cons,
                                options={
                                    'gtol': 1e-6,
                                    'maxiter': 1000}
                            )
                            
                    if best_result is None or result.fun <= best_result.fun:
                        best_result = result

                if best_result is None:
                    raise RuntimeError("Optimization failed for all initial guesses.")

                print(f"\n Result: {best_result}")
                solution = best_result.x
                print("Constraint value (should be >= 0):", g1(solution))

                custom_name = 'lines.tif'
                file_name = '_'.join([opt_file_name, custom_name])
                lines_opt_file_path = os.path.join(optimised_save, file_name)

                generate_lines(
                    image_height,
                    image_width,
                    25.4,
                    solution[1],
                    lines_opt_file_path,
                    orientation='vertical',
                    N_lines=solution[2]
                )

                generated_lines = cv2.imread(lines_opt_file_path)
                plt.imshow(generated_lines, cmap='gray')
                plt.title(f'Optimised pattern: {file_name}')
                plt.show(block=False)
                plt.get_current_fig_manager().window.raise_()
                plt.pause(2)
                plt.close()

            # ------------------------------------------------------------
            elif gen_value.lower() == 'checkerboard':
                # Coordinate array
                coordinate_arr = np.vstack((x_1, x_2)).T

                # SVR surrogate model
                scaler_X = StandardScaler()
                X_scaled = scaler_X.fit_transform(coordinate_arr)

                scaler_z = StandardScaler()
                z_scaled = scaler_z.fit_transform(z.reshape(-1, 1)).ravel()

                svr_model = SVR()

                grid = GridSearchCV(svr_model, param_grid=p_grid, cv=5, n_jobs=-1)
                grid.fit(X_scaled, z_scaled)
                svr_model = grid.best_estimator_

                def interp_value(X):
                    """
                    Polynomial interpolation for optimization
                    """
                    X_scaled = scaler_X.transform(X)
                    z_pred_scaled = svr_model.predict(X_scaled)
                    return scaler_z.inverse_transform(z_pred_scaled.reshape(-1, 1)).ravel()

                # Define constraint inequality
                scaler_nan = StandardScaler()
                x6_scaled = scaler_nan.fit_transform(x_6.reshape(-1, 1)).ravel()

                svr_nan = SVR()
                grid_nan = GridSearchCV(svr_nan, param_grid=p_grid, cv=5, n_jobs=-1)
                grid_nan.fit(X_scaled, x6_scaled)
                svr_nan = grid_nan.best_estimator_

                def g1(x):
                    nan_scaled = svr_nan.predict(scaler_X.transform(x.reshape(1, -1)))
                    nan_pred = scaler_nan.inverse_transform(nan_scaled.reshape(-1, 1)).ravel()[0]
                    nan_pred = np.clip(nan_pred, 0, 100)
                    return nan_threshold - nan_pred

                # Bounds
                bounds = [
                    (np.min(x_1), np.max(x_1)),
                    (np.min(x_2), np.max(x_2))
                ]

                def objective(x):
                    return optimization_direction * interp_value(x.reshape(1, -1))[0]

                best_result = None
                # Number of entries = number of starting points
                for i in range(len(p_param[:, 0])):
                            initial_guess = np.array([p_param[i, 0], p_param[i, 1]])
                            cons = [{'type': 'ineq', 'fun': g1}]

                            with warnings.catch_warnings():
                                warnings.filterwarnings("ignore", category=RuntimeWarning)
                                warnings.filterwarnings("ignore", category=UserWarning)

                                if optimiser == 0:
                                    result = minimize(
                                        objective,
                                        initial_guess,
                                        method='SLSQP',
                                        bounds=bounds,
                                        constraints=cons,
                                        options={
                                            'ftol': 1e-8,
                                            'maxiter': 1000}
                                    )
                                    
                                elif optimiser == 1:
                                    result = minimize(
                                        objective,
                                        initial_guess,
                                        method='trust-constr',
                                        bounds=bounds,
                                        constraints=cons,
                                        options={
                                            'gtol': 1e-5,
                                            'maxiter': 1000}
                                    )

                            if best_result is None or result.fun <= best_result.fun:
                                best_result = result

                if best_result is None:
                    raise RuntimeError("Optimization failed for all initial guesses.")

                print(f'\n\n Result: {best_result}')
                solution = best_result.x
                print("Constraint value (should be >= 0):", g1(best_result.x))

                # Generate and save
                custom_name = 'checkb.tif'
                file_name = '_'.join([opt_file_name, custom_name])
                cb_opt_file_path = os.path.join(optimised_save, opt_file_name)

                generate_checkerboard(
                    image_height,
                    image_width,
                    dpi=25.4,
                    path=cb_opt_file_path,
                    line_width=solution[0],
                    N_rows=solution[1]
                )

                generated_cb = cv2.imread(cb_opt_file_path)
                plt.imshow(generated_cb, cmap='gray')
                plt.title(f'Optimised pattern: {file_name}')
                plt.show(block=False)
                plt.get_current_fig_manager().window.raise_()
                plt.pause(2)
                plt.close()

            # ------------------------------------------------------------
            elif gen_value.lower() == 'perlin':

                # Coordinate array
                coordinate_arr = np.vstack((x_1, x_2, x_3, x_4)).T

                # SVR surrogate model
                scaler_X = StandardScaler()
                X_scaled = scaler_X.fit_transform(coordinate_arr)

                scaler_z = StandardScaler()
                z_scaled = scaler_z.fit_transform(z.reshape(-1, 1)).ravel()

                svr_model = SVR()

                grid = GridSearchCV(svr_model, param_grid=p_grid, cv=5, n_jobs=-1)
                grid.fit(X_scaled, z_scaled)
                svr_model = grid.best_estimator_

                def interp_value(X):
                    """
                    Polynomial interpolation for optimization
                    """
                    X_scaled = scaler_X.transform(X)
                    z_pred_scaled = svr_model.predict(X_scaled)
                    return scaler_z.inverse_transform(z_pred_scaled.reshape(-1, 1)).ravel()

                # Define constraint inequality
                scaler_nan = StandardScaler()
                x6_scaled = scaler_nan.fit_transform(x_6.reshape(-1, 1)).ravel()

                svr_nan = SVR()
                grid_nan = GridSearchCV(svr_nan, param_grid=p_grid, cv=5, n_jobs=-1)
                grid_nan.fit(X_scaled, x6_scaled)
                svr_nan = grid_nan.best_estimator_

                def g1(x):
                    nan_scaled = svr_nan.predict(scaler_X.transform(x.reshape(1, -1)))
                    nan_pred = scaler_nan.inverse_transform(nan_scaled.reshape(-1, 1)).ravel()[0]
                    nan_pred = np.clip(nan_pred, 0, 100)
                    return nan_threshold - nan_pred

                # Bounds
                bounds = [
                    (np.min(x_1), np.max(x_1)),
                    (np.min(x_2), np.max(x_2)),
                    (np.min(x_3), np.max(x_3)),
                    (np.min(x_4), np.max(x_4))
                ]

                def objective(x):
                    return optimization_direction * interp_value(x.reshape(1, -1))[0]

                best_result = None
                # Number of data points = number of starting points
                for i in range(len(mp_param[:, 0])):

                    initial_guess = np.array([mp_param[i, 0], mp_param[i, 1],
                                            mp_param[i, 2], mp_param[i, 3]])
                    
                    cons = [{'type': 'ineq', 'fun': g1}]

                    # Suppress runtime warnings
                    with warnings.catch_warnings():
                        warnings.filterwarnings("ignore", category=RuntimeWarning)
                        warnings.filterwarnings("ignore", category=UserWarning)

                        if optimiser == 0:
                            result = minimize(
                                objective,
                                initial_guess,
                                method='SLSQP',
                                bounds=bounds,
                                constraints=cons,
                                options={
                                    'ftol': 1e-12,   
                                    'gtol': 1e-10,   
                                    'xtol': 1e-12,   
                                    'maxiter': 5000, 
                                    'disp': False      
                                })
                        elif optimiser == 1:
                            result = minimize(
                                objective,
                                initial_guess,
                                method='trust-constr',
                                bounds=bounds,
                                constraints=cons,
                                options={
                                    'gtol': 1e-5,
                                    'maxiter': 1000}
                            )

                    if best_result is None or result.fun <= best_result.fun:
                        best_result = result

                if best_result is None:
                    raise RuntimeError("Optimization failed for all initial guesses.")

                print(f'\n\n Result: {best_result}')
                solution = best_result.x
                print("Constraint value (should be >= 0):", g1(best_result.x))

                # Generate and save
                custom_name = f'{indicators[1, 0]}_perlin.tif'
                file_name = '_'.join([opt_file_name, custom_name])
                perlin_opt_file_path = os.path.join(optimised_save, opt_file_name)

                optimised_perlin = imgan.generate_single_perlin_image(
                    image_height, 
                    image_width, 
                    scale=solution[0],
                    octaves=int(round(solution[1])), 
                    persistence=solution[2],
                    lacunarity=solution[3], 
                    texture_function=indicators[1, 0]
                )
                
                cv2.imwrite(perlin_opt_file_path, optimised_perlin)
                optimig = imgan.MIG(optimised_perlin)
                print(f"Optimised MIG value: {optimig}\n")
                plt.imshow(optimised_perlin, cmap='gray')
                plt.title(f'Optimised pattern: {file_name}')
                plt.show(block=False)
                plt.get_current_fig_manager().window.raise_()
                plt.pause(2)
                plt.close()

                # Get MIG
                optimig = imgan.MIG(imgan.readImage(spec_opt_file_path))
                print(f"\n\nimage shape = {Generated_speckles.shape}\nOptimised MIG value: {optimig}\n")

        first_pref = first_pref + 2

        # Confrim correct optimisation
        opt_obj = ["MSF","MIG","E_f","MIOSD","Shannon","PSA","SSSIG", "R_peak"]

        if metric_objective is not None:
            print(f'\n----------------------\n{opt_obj[metric_objective]} optimised')
        else:
            print(f'\n----------------------\nRMSE optimised')



    except Exception as batch_error:
        tb = traceback.extract_tb(batch_error.__traceback__)[-1]
        filename, line_number, func_name, _ = tb
        print(f"\nBatch {batch} error in {filename}, line {line_number}, in {func_name}:\n{batch_error}")



print("Process completed....")
tic2 = time.time()
now = datetime.now()
current_time = now.strftime("%H:%M:%S")
print("-------------------------------------------------------------------------------------------------")
print(f'Time taken: {tic2-tic1:.3f} seconds')
print("\nCurrent Time =", current_time)


